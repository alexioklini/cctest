# Data Workbench tool bodies.
#
# Cross-module deps (artifact-folder resolution, thread-locals) are imported
# lazily from `brain` at call time to avoid a circular import — brain.py
# imports these near the end of its module evaluation.
#
# All bodies operate on the per-session DuckDB file `_data.duckdb` inside the
# session's artifact folder (CLAUDE.md invariant: that folder is python_exec's
# cwd; here we resolve it explicitly because the daemon process doesn't chdir).

import json
import os
import re

_DUCKDB_FILENAME = "_data.duckdb"

# Read-only SQL only. We allow statements that start with one of these
# keywords (case-insensitive, after stripping comments/whitespace); anything
# else — INSERT/UPDATE/DELETE/CREATE/ATTACH/COPY/INSTALL/LOAD/... — is refused.
_READ_SQL_PREFIXES = ("select", "with", "describe", "desc", "show", "pragma", "table", "values", "summarize", "explain")
# DuckDB identifier — table/column names we'll accept for `register_as`.
_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _err(msg: str) -> str:
    return json.dumps({"error": msg}, ensure_ascii=False)


def _ok(result: dict) -> str:
    return json.dumps(result, ensure_ascii=False, default=str)


def data_session_dir() -> str | None:
    """Absolute path to the current session's artifact folder, or None.

    Mirrors `tool_python_exec` / `tool_generate_image`: derived from the
    `current_session_id` + `current_agent` thread-locals, NOT from process
    cwd (the server daemon never chdir's).
    """
    import brain
    session_id = getattr(brain._thread_local, "current_session_id", None)
    agent = getattr(brain._thread_local, "current_agent", None) or brain._current_agent
    if not (session_id and agent):
        return None
    folder = brain._get_artifact_session_folder(session_id)
    return os.path.join(brain.AGENTS_DIR, agent.agent_id, "artifacts", folder)


def data_db_path(create_dir: bool = False) -> str | None:
    d = data_session_dir()
    if not d:
        return None
    if create_dir:
        os.makedirs(d, exist_ok=True)
    return os.path.join(d, _DUCKDB_FILENAME)


def _strip_sql_comments(sql: str) -> str:
    # Drop /* ... */ blocks and -- line comments so the prefix check sees the
    # real first keyword. Cheap, not a parser — DuckDB still validates.
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql.strip()


def _is_read_only_sql(sql: str) -> bool:
    head = _strip_sql_comments(sql).lstrip("(").lstrip().lower()
    return any(head.startswith(p) for p in _READ_SQL_PREFIXES)


def tool_data_query(args: dict) -> str:
    """Run a read-only SQL query against the session's DuckDB tables.

    args: {sql: str, register_as?: str}
    returns: {columns, rows[:200], n_total, registered_as?}
    """
    sql = (args.get("sql") or "").strip()
    if not sql:
        return _err("data_query: `sql` is required")
    register_as = (args.get("register_as") or "").strip()
    if register_as and not _IDENT_RE.match(register_as):
        return _err(f"data_query: invalid table name {register_as!r} (letters, digits, underscores; must not start with a digit)")
    if not _is_read_only_sql(sql):
        return _err("data_query: only read-only statements are allowed (SELECT / WITH / DESCRIBE / SHOW / PRAGMA / SUMMARIZE). Use python_exec for anything that modifies the database.")

    db_path = data_db_path(create_dir=True)
    if not db_path:
        return _err("data_query: no active workbench session (missing session/agent context)")

    try:
        import duckdb
    except ImportError:
        return _err("data_query: duckdb is not installed on the server (pip install --break-system-packages duckdb)")

    con = None
    try:
        con = duckdb.connect(db_path)
        if register_as:
            # CREATE TABLE … AS materialises the result so the rest of the
            # turn can reference it. Read-only-ness of the *input* SQL is
            # already checked above; this CREATE is the one write we permit.
            con.execute(f'CREATE OR REPLACE TABLE "{register_as}" AS {sql}')
            rel = con.execute(f'SELECT * FROM "{register_as}"')
        else:
            rel = con.execute(sql)
        columns = [d[0] for d in (rel.description or [])]
        all_rows = rel.fetchall()
        n_total = len(all_rows)
        rows = [list(r) for r in all_rows[:200]]
        out = {"columns": columns, "rows": rows, "n_total": n_total}
        if register_as:
            out["registered_as"] = register_as
        if n_total > 200:
            out["note"] = f"showing first 200 of {n_total} rows"
        return _ok(out)
    except Exception as e:
        return _err(f"data_query: {type(e).__name__}: {e}")
    finally:
        if con is not None:
            try:
                con.close()
            except Exception:
                pass


# ── Anonymisation (data_anonymise) ────────────────────────────────────────────
#
# One body, two modes. anonymise: per-column de-identification with deterministic
# Python (no LLM — Rule 5); never mutates the source (new DuckDB table + new
# output file); writes a 3-sheet index file when a reversible strategy is used;
# re-runs the 71-detector scanner on the output (residual_scan); k-anon warning
# on generalisation. deanonymise: file + its index file in → reversible
# strategies restored, one-way ones left as-is.

import hashlib
import secrets

_INDEX_FORMAT_VERSION = "1.0"
_K_ANON_DEFAULT = 5
# Strategies whose effect can be undone with the index file.
_REVERSIBLE = {"tokenise", "hash"}  # hash only when salt is kept (default: kept in index)
_ONE_WAY = {"nullify", "redact", "generalise", "shuffle", "noise"}
_ALL_STRATEGIES = _REVERSIBLE | _ONE_WAY


def _norm_header(h) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(h).strip().lower())


def _redact_free_text(val):
    """Run the 71-detector regex set over a string, replace each match with
    [REDACTED:<rule_id>]. Returns (new_value, n_redactions)."""
    import brain
    if val is None:
        return None, 0
    s = str(val)
    if not s:
        return s, 0
    n = 0
    for rule in brain._pii_rules():
        rid = rule.get("id")
        pat = rule.get("re")
        if pat is None:
            continue
        new_s, k = pat.subn(f"[REDACTED:{rid}]", s)
        if k:
            n += k
            s = new_s
    return s, n


def _generalise_value(val, opts: dict):
    """Coarsen a value. opts.kind: date_month | date_quarter | age_band |
    postcode_district | bucket (numeric, opts.size). Falls back to a generic
    bucketing for numbers, returns the value unchanged if it can't coarsen."""
    if val is None:
        return None
    kind = (opts or {}).get("kind") or ""
    sval = str(val)
    try:
        if kind in ("date_month", "date_quarter", ""):
            m = re.search(r"(\d{4})-(\d{1,2})", sval) or re.search(r"(\d{4})/(\d{1,2})", sval)
            if m:
                y, mo = int(m.group(1)), int(m.group(2))
                if kind == "date_quarter":
                    q = (mo - 1) // 3 + 1
                    return f"{y}-Q{q}"
                return f"{y}-{mo:02d}"
        if kind == "age_band":
            n = int(float(sval))
            lo = (n // 10) * 10
            return f"{lo}-{lo+9}"
        if kind == "postcode_district":
            # Keep leading alphanumerics up to the first space / digit-group break.
            m = re.match(r"\s*([A-Za-z]*\d{1,3}|\d{1,3})", sval)
            if m:
                return m.group(1)
        if kind == "bucket":
            size = float((opts or {}).get("size") or 0) or 0
            n = float(sval)
            if size > 0:
                lo = (int(n // size)) * size
                # integer-ish formatting
                if size == int(size):
                    return f"{int(lo)}-{int(lo+size)}"
                return f"{lo}-{lo+size}"
    except (ValueError, TypeError):
        pass
    # Generic numeric coarsening to one significant figure of magnitude.
    try:
        n = float(sval)
        if n == 0:
            return "0"
        import math
        mag = 10 ** int(math.floor(math.log10(abs(n))))
        lo = (int(n // mag)) * mag
        return f"{int(lo)}-{int(lo+mag)}"
    except (ValueError, TypeError):
        return val


def _apply_strategy(series_values: list, strategy: str, opts: dict, salt: str):
    """Apply a strategy to one column's list of values.

    Returns (new_values, mapping) where `mapping` is {original_str: surrogate}
    for reversible strategies (else {}). `series_values` is a plain Python list.
    """
    opts = opts or {}
    mapping: dict = {}
    out = []
    if strategy == "nullify":
        return [None] * len(series_values), {}
    if strategy == "shuffle":
        import random
        idx = list(range(len(series_values)))
        random.shuffle(idx)
        return [series_values[i] for i in idx], {}
    if strategy == "noise":
        import random
        eps = float(opts.get("epsilon") or 0.05)
        for v in series_values:
            try:
                f = float(v)
                jitter = f * eps * (random.random() * 2 - 1)
                out.append(f + jitter if "." in str(v) or isinstance(v, float) else int(round(f + jitter)))
            except (ValueError, TypeError):
                out.append(v)
        return out, {}
    if strategy == "generalise":
        return [_generalise_value(v, opts) for v in series_values], {}
    if strategy == "redact":
        for v in series_values:
            nv, _ = _redact_free_text(v)
            out.append(nv)
        return out, {}
    if strategy == "hash":
        keep_prefix = int(opts.get("keep_prefix") or 0)
        for v in series_values:
            if v is None:
                out.append(None); continue
            s = str(v)
            digest = hashlib.sha256((salt + s).encode("utf-8")).hexdigest()[:16]
            pref = s[:keep_prefix] if keep_prefix > 0 else ""
            surrogate = (pref + digest) if pref else digest
            out.append(surrogate)
            mapping[s] = surrogate  # reversible only if salt kept
        return out, mapping
    if strategy == "tokenise":
        prefix = str(opts.get("prefix") or "ID")
        counter = {"n": 0}
        seen: dict = {}
        for v in series_values:
            if v is None:
                out.append(None); continue
            s = str(v)
            if s not in seen:
                counter["n"] += 1
                seen[s] = f"{prefix}_{counter['n']:05d}"
                mapping[s] = seen[s]
            out.append(seen[s])
        return out, mapping
    # Unknown strategy → leave column unchanged.
    return list(series_values), {}


def _write_index_xlsx(path: str, *, columns_spec: list, mapping_by_col: dict,
                      salt: str, embed_salt: bool, source_file: str,
                      source_sheets, residual_status: str, session_id: str,
                      audit_event_id: str = "") -> None:
    """Write the 3-sheet index workbook (mapping / schema / info) — modelled on
    the standalone Excel-Anonymisierung tool's mapping workbook so it's
    self-describing and round-trippable."""
    import openpyxl
    wb = openpyxl.Workbook()
    # --- mapping sheet ---
    ws_m = wb.active
    ws_m.title = "mapping"
    ws_m.append(["column_id", "column_name", "normalized_header", "original_value", "pseudonym"])
    for spec in columns_spec:
        col = spec["name"]
        m = mapping_by_col.get(col) or {}
        nh = _norm_header(col)
        for orig, surr in m.items():
            ws_m.append([col, col, nh, orig, surr])
    # --- schema sheet ---
    ws_s = wb.create_sheet("schema")
    ws_s.append(["column_id", "column_index", "header_original", "header_normalized",
                 "strategy", "opts", "created_at"])
    import time as _t
    now = _t.strftime("%Y-%m-%dT%H:%M:%S")
    for i, spec in enumerate(columns_spec):
        ws_s.append([spec["name"], i, spec["name"], _norm_header(spec["name"]),
                     spec.get("strategy", ""), json.dumps(spec.get("opts") or {}), now])
    # --- info sheet ---
    ws_i = wb.create_sheet("info")
    ws_i.append(["key", "value"])
    info = {
        "format_version": _INDEX_FORMAT_VERSION,
        "generated_at": now,
        "source_file": source_file or "",
        "source_sheets": ",".join(source_sheets) if isinstance(source_sheets, (list, tuple)) else (source_sheets or ""),
        "salt_present": "yes" if embed_salt else "no",
        "salt": salt if embed_salt else "",
        "residual_scan_status": residual_status,
        "brain_session_id": session_id or "",
        "audit_event_id": audit_event_id or "",
    }
    for k, v in info.items():
        ws_i.append([k, v])
    wb.save(path)


def _read_index_xlsx(path: str) -> dict:
    """Read a 3-sheet index workbook → {schema: [{column_name, normalized_header,
    strategy, opts}], mapping: {col -> {pseudonym -> original}}, info: {...}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"schema": [], "mapping": {}, "info": {}}
    if "schema" in wb.sheetnames:
        rows = list(wb["schema"].iter_rows(values_only=True))
        if rows:
            hdr = [str(h) for h in rows[0]]
            for r in rows[1:]:
                rec = dict(zip(hdr, r))
                opts = rec.get("opts")
                try:
                    opts = json.loads(opts) if isinstance(opts, str) and opts.strip() else (opts or {})
                except Exception:
                    opts = {}
                out["schema"].append({
                    "column_name": rec.get("header_original") or rec.get("column_id"),
                    "normalized_header": rec.get("header_normalized") or _norm_header(rec.get("column_id") or ""),
                    "strategy": rec.get("strategy") or "",
                    "opts": opts,
                })
    if "mapping" in wb.sheetnames:
        rows = list(wb["mapping"].iter_rows(values_only=True))
        if rows:
            hdr = [str(h) for h in rows[0]]
            for r in rows[1:]:
                rec = dict(zip(hdr, r))
                col = rec.get("column_name") or rec.get("column_id")
                if col is None:
                    continue
                orig = rec.get("original_value")
                surr = rec.get("pseudonym")
                if surr is None:
                    continue
                out["mapping"].setdefault(str(col), {})[str(surr)] = orig
    if "info" in wb.sheetnames:
        for r in wb["info"].iter_rows(values_only=True):
            if r and r[0] is not None and str(r[0]).lower() != "key":
                out["info"][str(r[0])] = r[1]
    wb.close()
    return out


def _audit(action: str, *, session_id: str, args_summary: str, result_summary: str) -> str:
    import brain
    try:
        if brain._audit_log:
            brain._audit_log.log_action(
                agent="main", action_type=action, tool_name=action,
                args_summary=args_summary, result_summary=result_summary,
                session_id=session_id or "", source="data_workbench")
    except Exception:
        pass
    return action


def _pii_scan_df_sample(df, sample_rows: int = 200) -> dict:
    """Run the 71-detector scanner over a sample of a DataFrame. Returns
    {clean: bool, findings: [{column, category, count, examples}]}."""
    import brain
    findings_by_col: dict = {}
    head = df.head(sample_rows)
    for col in head.columns:
        for val in head[col].tolist():
            sval = "" if val is None else str(val)
            if not sval:
                continue
            try:
                fs = brain._pii_scan_text(sval, max_findings=3)
            except Exception:
                fs = []
            for f in fs:
                key = (col, f.get("category", "personal"))
                rec = findings_by_col.setdefault(key, {"column": col, "category": f.get("category", "personal"),
                                                       "count": 0, "examples": []})
                rec["count"] += 1
                if len(rec["examples"]) < 2:
                    rec["examples"].append(sval[:40])
    findings = list(findings_by_col.values())
    return {"clean": len(findings) == 0, "findings": findings}


def tool_data_anonymise(args: dict) -> str:
    """Anonymise (or deanonymise) a DuckDB table / uploaded file. Deterministic
    Python only — the caller (agent or GUI) supplies the column→strategy plan.

    args (anonymise): {table, columns: [{name, strategy, opts?}],
        output_format?: "preserve"|"csv"|"xlsx"|"markdown", new_table?, mapping_out?,
        source_file?, embed_salt?: bool, k?: int}
    args (deanonymise): {mode:"deanonymise", source_file, index_file, output_format?}
    """
    import brain
    mode = (args.get("mode") or "anonymise").strip()
    db_path = data_db_path(create_dir=True)
    sess_dir = data_session_dir()
    if not db_path or not sess_dir:
        return _err("data_anonymise: no active workbench session")
    session_id = getattr(brain._thread_local, "current_session_id", "") or ""

    try:
        import duckdb
        import pandas as pd
    except ImportError as e:
        return _err(f"data_anonymise: required package missing ({e})")

    # ── DEANONYMISE ──────────────────────────────────────────────────────────
    if mode == "deanonymise":
        src_name = (args.get("source_file") or "").strip()
        idx_name = (args.get("index_file") or "").strip()
        if not src_name or not idx_name:
            return _err("data_anonymise(deanonymise): `source_file` and `index_file` are required")
        src_path = os.path.join(sess_dir, os.path.basename(src_name))
        idx_path = os.path.join(sess_dir, os.path.basename(idx_name))
        if not os.path.exists(src_path):
            return _err(f"data_anonymise(deanonymise): source file not found: {os.path.basename(src_name)}")
        if not os.path.exists(idx_path):
            return _err(f"data_anonymise(deanonymise): index file not found: {os.path.basename(idx_name)}")
        try:
            idx = _read_index_xlsx(idx_path)
        except Exception as e:
            return _err(f"data_anonymise(deanonymise): cannot read index file: {e}")
        # Load the anonymised file into a DataFrame (xlsx or csv).
        ext = os.path.splitext(src_path)[1].lower()
        try:
            if ext in (".csv", ".tsv"):
                df = pd.read_csv(src_path, sep=None, engine="python")
            else:
                df = pd.read_excel(src_path)
        except Exception as e:
            return _err(f"data_anonymise(deanonymise): cannot read source file: {e}")
        # Match treated columns by normalized header, then position.
        norm_to_actual = {_norm_header(c): c for c in df.columns}
        actual_cols = list(df.columns)
        not_reversible = []
        restored_cols = []
        for i, sch in enumerate(idx["schema"]):
            strat = sch.get("strategy", "")
            target = norm_to_actual.get(sch.get("normalized_header"))
            if target is None and i < len(actual_cols):
                target = actual_cols[i]
            if target is None:
                continue
            if strat not in _REVERSIBLE:
                not_reversible.append({"column": target, "strategy": strat})
                continue
            inv = idx["mapping"].get(sch.get("column_name")) or idx["mapping"].get(target) or {}
            if not inv:
                not_reversible.append({"column": target, "strategy": strat, "reason": "no mapping rows in index"})
                continue
            df[target] = df[target].map(lambda v: inv.get(str(v), v) if v is not None else v)
            restored_cols.append(target)
        # Write restored file in the source format (default = same as source).
        out_fmt = (args.get("output_format") or ("csv" if ext in (".csv", ".tsv") else "xlsx")).strip()
        base = os.path.splitext(os.path.basename(src_name))[0]
        if base.endswith("_anon"):
            base = base[:-5]
        if out_fmt == "csv" or ext in (".csv", ".tsv"):
            out_path = os.path.join(sess_dir, f"{base}_restored.csv")
            df.to_csv(out_path, index=False)
        elif out_fmt == "markdown":
            out_path = os.path.join(sess_dir, f"{base}_restored.md")
            with open(out_path, "w") as f:
                f.write(df.head(500).to_markdown(index=False))
        else:
            out_path = os.path.join(sess_dir, f"{base}_restored.xlsx")
            df.to_excel(out_path, index=False)
        try:
            brain._after_file_write(out_path, "created", "main")
        except Exception:
            pass
        _audit("data_deanonymise", session_id=session_id,
               args_summary=f"source={os.path.basename(src_name)} index={os.path.basename(idx_name)}",
               result_summary=f"restored {len(restored_cols)} cols, {len(not_reversible)} not reversible")
        return _ok({
            "output_artifact": os.path.basename(out_path),
            "columns_restored": restored_cols,
            "not_reversible": not_reversible,
            "audit_event": "data_deanonymise",
        })

    # ── ANONYMISE ────────────────────────────────────────────────────────────
    table = (args.get("table") or "").strip()
    columns = args.get("columns") or []
    if not table:
        return _err("data_anonymise: `table` is required")
    if not _IDENT_RE.match(table):
        return _err(f"data_anonymise: invalid table name {table!r}")
    if not columns or not isinstance(columns, list):
        return _err("data_anonymise: `columns` must be a non-empty list of {name, strategy, opts?}")
    for spec in columns:
        if not isinstance(spec, dict) or not spec.get("name"):
            return _err("data_anonymise: each column spec needs a `name`")
        strat = (spec.get("strategy") or "").strip()
        if strat not in _ALL_STRATEGIES:
            return _err(f"data_anonymise: unknown strategy {strat!r} for column {spec.get('name')!r}. "
                        f"Allowed: {sorted(_ALL_STRATEGIES)}")
        spec["strategy"] = strat
    output_format = (args.get("output_format") or "preserve").strip()
    new_table = (args.get("new_table") or f"{table}_anon").strip()
    if not _IDENT_RE.match(new_table):
        return _err(f"data_anonymise: invalid new_table name {new_table!r}")
    mapping_out = (args.get("mapping_out") or f"{table}_anon_map.xlsx").strip()
    mapping_out = os.path.basename(mapping_out)
    if not mapping_out.lower().endswith(".xlsx"):
        mapping_out += ".xlsx"
    source_file = (args.get("source_file") or "").strip()
    embed_salt = bool(args.get("embed_salt"))
    k_threshold = int(args.get("k") or _K_ANON_DEFAULT)

    con = None
    try:
        con = duckdb.connect(db_path)
        existing = {r[0] for r in con.execute("SHOW TABLES").fetchall()}
        if table not in existing:
            return _err(f"data_anonymise: table {table!r} not found in this workbench")
        df = con.execute(f'SELECT * FROM "{table}"').df()
        col_set = set(df.columns)
        for spec in columns:
            if spec["name"] not in col_set:
                return _err(f"data_anonymise: column {spec['name']!r} not in table {table!r}")

        salt = secrets.token_hex(16)
        mapping_by_col: dict = {}
        used_reversible = False
        for spec in columns:
            col = spec["name"]; strat = spec["strategy"]; opts = spec.get("opts") or {}
            new_vals, mapping = _apply_strategy(df[col].tolist(), strat, opts, salt)
            df[col] = new_vals
            if mapping:
                mapping_by_col[col] = mapping
            if strat in _REVERSIBLE:
                used_reversible = True

        # Materialise the new DuckDB table (source untouched).
        t = new_table; i = 2
        while t in existing:
            t = f"{new_table}_{i}"; i += 1
        con.register("_anon_tmp", df)
        con.execute(f'CREATE TABLE "{t}" AS SELECT * FROM _anon_tmp')
        con.unregister("_anon_tmp")
        new_table = t

        # Residual scan over the anonymised data.
        residual = _pii_scan_df_sample(df)
        residual_status = "clean" if residual["clean"] else "dirty"

        # k-anonymity warning: if any column was generalised, check quasi-identifier
        # group sizes over the set of generalised columns.
        k_warning = None
        gen_cols = [s["name"] for s in columns if s["strategy"] == "generalise"]
        if gen_cols:
            try:
                grp = df.groupby(gen_cols).size()
                min_k = int(grp.min())
                small = int((grp < k_threshold).sum())
                if min_k < k_threshold:
                    k_warning = {"min_group_size": min_k, "k": k_threshold,
                                 "groups_below_k": small,
                                 "note": f"{small} quasi-identifier group(s) have fewer than {k_threshold} rows"}
            except Exception:
                pass

        # ── emit the output artifact in the requested shape ──
        out_artifact = None
        out_fmt_effective = output_format
        if output_format == "preserve" and source_file and source_file.lower().endswith((".xlsx", ".xlsm")):
            # Rewrite the original workbook, preserving sheets/formatting; only
            # the table's sheet's treated cells change. Phase A: openpyxl
            # roundtrip (a faithful in-place chunked-OOXML rewrite — the
            # translate.py path — lands in a later PR for docx/pptx too).
            src_path = os.path.join(sess_dir, os.path.basename(source_file))
            if not os.path.exists(src_path):
                return _err(f"data_anonymise: source_file {os.path.basename(source_file)} not found for preserve output")
            out_artifact = _rewrite_xlsx_preserve(src_path, df, columns, sess_dir)
        elif output_format == "preserve" and source_file and source_file.lower().endswith((".csv", ".tsv")):
            base = os.path.splitext(os.path.basename(source_file))[0]
            out_artifact = os.path.join(sess_dir, f"{base}_anon.csv")
            df.to_csv(out_artifact, index=False)
        elif output_format == "xlsx":
            out_artifact = os.path.join(sess_dir, f"{table}_anon.xlsx")
            df.to_excel(out_artifact, index=False)
        elif output_format == "markdown":
            out_artifact = os.path.join(sess_dir, f"{table}_anon.md")
            with open(out_artifact, "w") as f:
                f.write(df.head(500).to_markdown(index=False))
        else:
            # default / csv / preserve-without-source → fresh csv from the table
            out_fmt_effective = "csv"
            out_artifact = os.path.join(sess_dir, f"{table}_anon.csv")
            df.to_csv(out_artifact, index=False)
        try:
            brain._after_file_write(out_artifact, "created", "main")
        except Exception:
            pass

        # Index file (only when a reversible strategy was used).
        mapping_path = None
        if used_reversible:
            mapping_path = os.path.join(sess_dir, mapping_out)
            try:
                _write_index_xlsx(
                    mapping_path, columns_spec=columns, mapping_by_col=mapping_by_col,
                    salt=salt, embed_salt=embed_salt, source_file=source_file,
                    source_sheets="", residual_status=residual_status, session_id=session_id)
                brain._after_file_write(mapping_path, "created", "main")
            except Exception as e:
                return _err(f"data_anonymise: failed to write index file: {e}")

        # Update the DataSessionDB table index.
        try:
            from server_lib.db import DataSessionDB
            all_tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
            DataSessionDB.set_tables(session_id, all_tables)
        except Exception:
            pass

        _audit("data_anonymise", session_id=session_id,
               args_summary=f"table={table} cols={[c['name'] for c in columns]} strat={[c['strategy'] for c in columns]} fmt={out_fmt_effective}",
               result_summary=f"new_table={new_table} residual={residual_status} reversible={used_reversible}")

        out = {
            "new_table": new_table,
            "output_artifact": os.path.basename(out_artifact),
            "output_format": out_fmt_effective,
            "columns_treated": [{"name": c["name"], "strategy": c["strategy"]} for c in columns],
            "residual_scan": residual,
            "audit_event": "data_anonymise",
        }
        if mapping_path:
            out["mapping_file"] = os.path.basename(mapping_path)
        if k_warning:
            out["k_anon_warning"] = k_warning
        return _ok(out)
    except Exception as e:
        return _err(f"data_anonymise: {type(e).__name__}: {e}")
    finally:
        if con is not None:
            try:
                con.close()
            except Exception:
                pass


def _rewrite_xlsx_preserve(src_path: str, df, columns_spec: list, sess_dir: str) -> str:
    """Rewrite an .xlsx in place, preserving sheets/formatting — only the cells
    in the treated columns change. Phase A: openpyxl roundtrip.

    Strategy: load the workbook with openpyxl; find the sheet whose header row
    matches the DataFrame's columns (normalized); overwrite the treated columns'
    cells from the anonymised DataFrame, row by row. Other sheets pass through.
    """
    import openpyxl
    wb = openpyxl.load_workbook(src_path)  # keep formatting
    treated = {_norm_header(s["name"]): s["name"] for s in columns_spec}
    df_norm = {_norm_header(c): c for c in df.columns}
    target_ws = None
    header_row_idx = 1
    col_map = {}  # sheet column index (1-based) -> df column name
    for ws in wb.worksheets:
        # Find a header row in the first 5 rows that contains all treated headers.
        for hr in range(1, min(6, ws.max_row + 1)):
            headers = {}
            for cell in ws[hr]:
                if cell.value is not None:
                    headers[_norm_header(cell.value)] = cell.column
            if treated and all(nh in headers for nh in treated):
                target_ws = ws
                header_row_idx = hr
                # map sheet columns to df columns by normalized header
                for nh, sheet_col in headers.items():
                    if nh in df_norm:
                        col_map[sheet_col] = df_norm[nh]
                break
        if target_ws is not None:
            break
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_path = os.path.join(sess_dir, f"{base}_anon.xlsx")
    if target_ws is None:
        # Couldn't line up the sheet — fall back to a fresh single-sheet xlsx
        # rather than silently emitting an unchanged copy.
        df.to_excel(out_path, index=False)
        return out_path
    # Overwrite treated columns row by row (data starts after the header row).
    treated_sheet_cols = [sc for sc, dfc in col_map.items() if _norm_header(dfc) in treated]
    for r_off, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row_idx + 1 + r_off
        if excel_row > target_ws.max_row + len(df):  # safety
            break
        for sc in treated_sheet_cols:
            dfc = col_map[sc]
            val = row[dfc]
            target_ws.cell(row=excel_row, column=sc, value=(None if (val is None or (isinstance(val, float) and val != val)) else val))
    wb.save(out_path)
    return out_path


# ── File-level GDPR scan (data_scan_files) ───────────────────────────────────
#
# PR3 scope: scan the workbench's DuckDB tables (tabular sources). Reports which
# tables leak, where (which column), how bad, and a suggested per-column
# strategy. Modifies nothing. Pure code, no LLM. One audit line (files_scanned).
# docx/pptx/pdf file scanning lands in PR4 alongside the format arms.

# Per-detector-category → suggested anonymise strategy for the auto-fix path.
_SUGGEST_STRATEGY = {
    "contact": "redact",          # emails / phones inside free text
    "personal": "tokenise",       # names
    "national_id": "tokenise",
    "national_id_ctx": "tokenise",
    "bare_id": "tokenise",
    "financial": "hash",          # IBANs / card PANs
    "secrets": "redact",
    "network": "redact",
}


def _scan_table(con, table: str, sample_rows: int = 500) -> dict:
    """Run the 71-detector scanner over a table's columns (a sampled prefix).

    Returns {name, type:"table", status, findings:[{where, category, count,
    examples, suggested_strategy}], worst_category, total_hits}.
    """
    import brain
    try:
        df = con.execute(f'SELECT * FROM "{table}" LIMIT {sample_rows}').df()
    except Exception as e:
        return {"name": table, "type": "table", "status": "error",
                "findings": [], "worst_category": None, "total_hits": 0,
                "error": f"{type(e).__name__}: {e}"}
    findings_by_col: dict = {}
    for col in df.columns:
        for val in df[col].tolist():
            sval = "" if val is None else str(val)
            if not sval:
                continue
            try:
                fs = brain._pii_scan_text(sval, max_findings=3)
            except Exception:
                fs = []
            for f in fs:
                cat = f.get("category", "personal")
                rec = findings_by_col.setdefault((col, cat), {
                    "where": f"column {col}", "column": col, "category": cat,
                    "count": 0, "examples": [],
                    "suggested_strategy": _SUGGEST_STRATEGY.get(cat, "tokenise"),
                })
                rec["count"] += 1
                if len(rec["examples"]) < 2:
                    rec["examples"].append(sval[:40])
    findings = list(findings_by_col.values())
    total = sum(f["count"] for f in findings)
    worst = None
    if findings:
        # "worst" = the category with the most hits; ties broken by a rough severity order.
        sev = {"secrets": 5, "national_id": 4, "national_id_ctx": 4, "financial": 3,
               "personal": 2, "contact": 1, "network": 1, "bare_id": 2}
        worst = max(findings, key=lambda f: (f["count"], sev.get(f["category"], 0)))["category"]
    status = "clean" if not findings else "dirty"
    return {"name": table, "type": "table", "status": status, "findings": findings,
            "worst_category": worst, "total_hits": total}


def tool_data_scan_files(args: dict) -> str:
    """File-level GDPR scan over the workbench's DuckDB tables.

    args: {tables?: [str]} — if omitted, scans every table in the workbench.
    Returns {files: [...per-table report...], summary: {scanned, clean, dirty, error}}.
    Modifies nothing.
    """
    import brain
    db_path = data_db_path(create_dir=False)
    if not db_path or not os.path.exists(db_path):
        return _ok({"files": [], "summary": {"scanned": 0, "clean": 0, "dirty": 0, "error": 0},
                    "note": "no tables in this workbench yet"})
    session_id = getattr(brain._thread_local, "current_session_id", "") or ""
    requested = args.get("tables") or []

    try:
        import duckdb
    except ImportError:
        return _err("data_scan_files: duckdb is not installed on the server")

    con = None
    try:
        con = duckdb.connect(db_path, read_only=True)
        all_tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
        targets = [t for t in (requested or all_tables) if t in all_tables]
        if requested and not targets:
            return _err(f"data_scan_files: none of {requested} are tables in this workbench (have: {all_tables})")
        reports = [_scan_table(con, t) for t in targets]
    except Exception as e:
        return _err(f"data_scan_files: {type(e).__name__}: {e}")
    finally:
        if con is not None:
            try:
                con.close()
            except Exception:
                pass

    summary = {
        "scanned": len(reports),
        "clean": sum(1 for r in reports if r["status"] == "clean"),
        "dirty": sum(1 for r in reports if r["status"] == "dirty"),
        "error": sum(1 for r in reports if r["status"] == "error"),
    }
    _audit("files_scanned", session_id=session_id,
           args_summary=f"tables={[r['name'] for r in reports]}",
           result_summary=f"{summary['dirty']} dirty / {summary['scanned']} scanned")
    return _ok({"files": reports, "summary": summary})


# ── Chart rendering (data_render_chart) ──────────────────────────────────────
#
# Validate a Vega-Lite spec, bind it to one of the workbench's DuckDB tables
# (inlining a ≤5k-row sample as `data.values`), render a PNG server-side via
# vl-convert (a self-contained Rust wheel — no node, no headless browser).
# Raises on an invalid spec / a field the table doesn't have → the agentic loop
# hands the error back, the model retries.

_CHART_SAMPLE_ROWS = 5000


def _collect_encoded_fields(spec) -> set:
    """Walk a Vega-Lite spec and collect every `field` referenced in any
    `encoding` block (top-level, layered, faceted, concatenated)."""
    fields: set = set()
    def walk(node):
        if isinstance(node, dict):
            enc = node.get("encoding")
            if isinstance(enc, dict):
                for ch in enc.values():
                    chans = ch if isinstance(ch, list) else [ch]
                    for c in chans:
                        if isinstance(c, dict) and isinstance(c.get("field"), str):
                            fields.add(c["field"])
                        # transform-derived fields (aggregate/calculate) won't be in the table
                        # but neither will they error on render — leave them; we only flag
                        # bare `field` refs that look like raw columns.
            for key in ("layer", "concat", "hconcat", "vconcat", "spec"):
                child = node.get(key)
                if isinstance(child, list):
                    for ch in child:
                        walk(ch)
                elif isinstance(child, dict):
                    walk(child)
            facet = node.get("facet")
            if isinstance(facet, dict):
                for ch in facet.values():
                    if isinstance(ch, dict) and isinstance(ch.get("field"), str):
                        fields.add(ch["field"])
        elif isinstance(node, list):
            for ch in node:
                walk(ch)
    walk(spec)
    return fields


def _spec_uses_calculate_or_aggregate(spec) -> bool:
    """True if the spec has transforms / aggregate channels that synthesise
    field names not present in the source table — we then skip the strict
    'every field must be a column' check (Vega will validate on render)."""
    s = json.dumps(spec)
    return ('"calculate"' in s or '"aggregate"' in s or '"fold"' in s
            or '"window"' in s or '"joinaggregate"' in s)


def tool_data_render_chart(args: dict) -> str:
    """Render a Vega-Lite chart bound to a workbench table → PNG.

    args: {spec: dict, table: str, scale?: float}
    returns: {ok, table, spec, png_b64, n_rows}
    """
    import brain
    spec = args.get("spec")
    table = (args.get("table") or "").strip()
    if not isinstance(spec, dict):
        return _err("data_render_chart: `spec` must be a Vega-Lite spec object")
    if not table or not _IDENT_RE.match(table):
        return _err("data_render_chart: `table` is required (name of a DuckDB table in this workbench)")
    db_path = data_db_path(create_dir=False)
    if not db_path or not os.path.exists(db_path):
        return _err("data_render_chart: no tables in this workbench yet")

    try:
        import duckdb
        import vl_convert as vlc
    except ImportError as e:
        return _err(f"data_render_chart: required package missing ({e})")

    con = None
    try:
        con = duckdb.connect(db_path, read_only=True)
        all_tables = {r[0] for r in con.execute("SHOW TABLES").fetchall()}
        if table not in all_tables:
            return _err(f"data_render_chart: table {table!r} not found (have: {sorted(all_tables)})")
        col_names = [c[0] for c in con.execute(f'DESCRIBE "{table}"').fetchall()]
        # Strict field check unless the spec synthesises fields via transforms.
        if not _spec_uses_calculate_or_aggregate(spec):
            used = _collect_encoded_fields(spec)
            missing = [f for f in used if f not in col_names]
            if missing:
                return _err(f"data_render_chart: spec references column(s) {missing} not in table {table!r}. "
                            f"Available columns: {col_names}")
        df = con.execute(f'SELECT * FROM "{table}" LIMIT {_CHART_SAMPLE_ROWS}').df()
        n_total = con.execute(f'SELECT count(*) FROM "{table}"').fetchone()[0]
    except Exception as e:
        return _err(f"data_render_chart: {type(e).__name__}: {e}")
    finally:
        if con is not None:
            try:
                con.close()
            except Exception:
                pass

    # Bind the table's sample as inline data (vl-convert can't query DuckDB).
    bound = dict(spec)
    # Convert the DataFrame to JSON-safe records.
    records = json.loads(df.to_json(orient="records", date_format="iso"))
    bound["data"] = {"values": records}
    bound.setdefault("$schema", "https://vega.github.io/schema/vega-lite/v5.json")

    try:
        scale = float(args.get("scale") or 2.0)
    except (TypeError, ValueError):
        scale = 2.0
    try:
        png_bytes = vlc.vegalite_to_png(json.dumps(bound), scale=scale)
    except Exception as e:
        return _err(f"data_render_chart: invalid Vega-Lite spec — {type(e).__name__}: {str(e)[:400]}")

    import base64
    png_b64 = base64.b64encode(png_bytes).decode("ascii")

    # Persist the PNG as an artifact so it shows up in the Artifacts panel.
    artifact_name = None
    try:
        sess_dir = data_session_dir()
        if sess_dir:
            i = 1
            while os.path.exists(os.path.join(sess_dir, f"chart_{i}.png")):
                i += 1
            artifact_name = f"chart_{i}.png"
            p = os.path.join(sess_dir, artifact_name)
            with open(p, "wb") as f:
                f.write(png_bytes)
            brain._after_file_write(p, "created", "main")
    except Exception:
        artifact_name = None

    # Emit a data_chart SSE event if anyone's listening (the chat UI).
    try:
        cb = getattr(brain._thread_local, "event_callback", None)
        if callable(cb):
            cb("data_chart", {"table": table, "spec": spec, "png_b64": png_b64,
                              "artifact": artifact_name, "n_rows": len(records)})
    except Exception:
        pass

    _audit("data_render_chart", session_id=getattr(brain._thread_local, "current_session_id", "") or "",
           args_summary=f"table={table}",
           result_summary=f"{len(records)} rows rendered, artifact={artifact_name}")

    out = {"ok": True, "table": table, "spec": spec, "png_b64": png_b64, "n_rows": len(records)}
    if n_total > len(records):
        out["note"] = f"chart rendered from a {len(records)}-row sample of {n_total} total rows"
    if artifact_name:
        out["artifact"] = artifact_name
    return _ok(out)
