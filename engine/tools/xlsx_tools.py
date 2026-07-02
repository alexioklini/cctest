"""Deterministic XLSX toolset — xlsx_inspect / xlsx_query / xlsx_create / xlsx_edit.

WHY (chats 2cb94154 / 98cceac2): spreadsheet quality used to depend entirely on
the chat model writing pandas/openpyxl code via python_exec — strong models
managed, weak ones churned through script iterations and delivered CSV instead
of styled workbooks. Like the docx/html pipeline (model writes markdown, code
renders the file), these tools make the model supply only INTENT (a SQL SELECT,
a small JSON spec) while the server moves the data deterministically:

  xlsx_inspect  — workbook profile (structure, column types, join-key candidates)
  xlsx_query    — ONE read-only SELECT over sheets loaded into in-memory SQLite
  xlsx_create   — declarative spec → styled workbook (house style preset)
  xlsx_edit     — declarative ops on an existing workbook, format-preserving

Bulk data never flows through the model: query results are capped for display
(full result → CSV artifact), create/edit pull rows server-side via
`source: {file, sheet?, sql?}`.

Wired per the 4-site rule (TOOL_DEFINITIONS / TOOL_GROUPS / impl here /
TOOL_DISPATCH). Reaches brain runtime via lazy `import brain as _brain`.
"""

from __future__ import annotations

import csv
import datetime
import json
import os
import re
import shutil
import sqlite3

from engine.context import get_request_context
from engine.tool_exec import _ok, _err
from engine import doc_convert

_XLSX_EXTS = {".xlsx", ".xlsm"}
_CSV_EXTS = {".csv", ".tsv"}
_LEGACY_EXTS = {".xls", ".ods"}  # converted to .xlsx via LibreOffice (v3)

# Caps: keep tool results small (the model only ever needs a preview — full
# results go to CSV artifacts) and refuse workbooks too big for an in-memory
# SQLite session.
QUERY_DISPLAY_ROWS = 50
QUERY_MAX_TOTAL_ROWS = 200_000
# xlsx_create sources may exceed the query cap — the streaming writer (v4)
# handles the output side, and the read side streams via read_only.
CREATE_MAX_SOURCE_ROWS = 750_000
QUERY_MAX_FILE_MB = 30
CREATE_INLINE_MAX_CELLS = 5_000
INSPECT_SAMPLE_VALUES = 3
INSPECT_DISTINCT_SAMPLE = 5_000
JOIN_KEY_OVERLAP_SAMPLE = 200
JOIN_KEY_MIN_OVERLAP = 0.5

_NUMBER_FORMATS = {
    "text": "@",
    "int": "#,##0",
    "number": "#,##0.00",
    "eur": '#,##0.00 "€"',
    "percent": "0.0%",
    "date": "DD.MM.YYYY",
}

# Excel's default 3-color scale + traffic-light fills for CellIs rules.
_COND_FILLS = {"red": "F8696B", "green": "63BE7B", "yellow": "FFEB84"}

_UMLAUT_MAP = str.maketrans(
    {"ä": "ae", "ö": "oe", "ü": "ue", "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
     "ß": "ss"})


# ---------------------------------------------------------------------------
# LibreOffice bridge (v3): legacy-format conversion + formula recalc
# ---------------------------------------------------------------------------
# openpyxl can neither read .xls/.ods nor COMPUTE formulas (it writes them;
# Excel calculates on open). A headless LibreOffice round-trip covers both:
# convert-to-xlsx re-saves with freshly calculated cached values. Config
# override: config.json → xlsx.soffice_path; else auto-detect.

_SOFFICE_CANDIDATES = (
    "/opt/homebrew/bin/soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
)


def _find_soffice() -> str | None:
    import brain as _brain
    try:
        cfg = (_brain._server_config().get("xlsx") or {}).get("soffice_path")
        if cfg and os.path.isfile(cfg):
            return cfg
    except Exception:
        pass
    for c in _SOFFICE_CANDIDATES:
        if os.path.isfile(c):
            return c
    return shutil.which("soffice")


def _soffice_convert(src: str, out_dir: str, timeout: int = 120) -> str:
    """`soffice --headless --convert-to xlsx` into out_dir; returns the
    produced path. LibreOffice recalculates formulas on load/save, so the
    output carries fresh cached values (the recalc mechanism)."""
    import subprocess
    soffice = _find_soffice()
    if not soffice:
        raise ValueError(
            "LibreOffice (soffice) not found — needed for .xls/.ods reading "
            "and recalc. Install it or set config.json → xlsx.soffice_path.")
    os.makedirs(out_dir, exist_ok=True)
    proc = subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", out_dir,
         src],
        capture_output=True, timeout=timeout)
    produced = os.path.join(
        out_dir, os.path.splitext(os.path.basename(src))[0] + ".xlsx")
    if proc.returncode != 0 or not os.path.isfile(produced):
        tail = (proc.stderr or proc.stdout or b"")[-300:].decode(
            "utf-8", "replace")
        raise ValueError(f"LibreOffice conversion failed: {tail}")
    return produced


def _legacy_to_xlsx(path: str) -> str:
    """Convert a .xls/.ods to .xlsx once per (path, mtime) in a tmp cache so
    repeated reads don't pay the soffice round-trip."""
    st = os.stat(path)
    cache_dir = os.path.join("/tmp", "brain-xlsx-convert")
    key = re.sub(r"[^0-9a-zA-Z]+", "_", os.path.abspath(path)) \
        + f"_{int(st.st_mtime)}"
    out_dir = os.path.join(cache_dir, key)
    produced = os.path.join(
        out_dir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
    if os.path.isfile(produced):
        return produced
    return _soffice_convert(path, out_dir)


def recalc_workbook(path: str) -> None:
    """Recalculate a workbook's formulas in place (LibreOffice round-trip) —
    used by xlsx_create/xlsx_edit `recalc: true` so a follow-up xlsx_query
    sees computed values instead of NULL formula cells."""
    import tempfile
    out_dir = tempfile.mkdtemp(prefix="xlsx-recalc-", dir="/tmp")
    produced = _soffice_convert(path, out_dir)
    shutil.move(produced, path)
    shutil.rmtree(out_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Grid loading (shared by inspect / query / create-source / edit-source)
# ---------------------------------------------------------------------------

def _cell_str(v) -> str:
    """Emptiness-preserving normalizer for the placeholder trim (matches the
    semantics of doc_convert._extract_xlsx's `_cell` — caps/pipe-escaping don't
    affect emptiness, so plain str+strip is equivalent for trimming)."""
    return "" if v is None else str(v).strip()


def _coerce_csv_value(s):
    """CSV cells arrive as strings; coerce int/float so SQL aggregates work.
    Comma-decimal locals stay text (ambiguous with thousands separators)."""
    if s is None:
        return None
    t = s.strip()
    if t == "":
        return None
    try:
        return int(t)
    except ValueError:
        pass
    try:
        return float(t)
    except ValueError:
        return s


def _is_headerish(row) -> bool:
    """Does this row look like a header (≥2 labels, mostly strings)?"""
    vals = [v for v in row if v is not None and str(v).strip() != ""]
    if len(vals) < 2:
        return False
    return sum(isinstance(v, str) for v in vals) / len(vals) >= 0.6


def _split_table_blocks(raw_rows: list) -> list[tuple[int, list]]:
    """Split a sheet into table BLOCKS (v2 multi-table detection). Real-world
    report sheets often stack several tables with titles/blank gaps between
    them; flattening them into one grid mis-detects the header and garbles
    types. Split rule (conservative — a legit sparse table must NOT split):
    ≥2 consecutive fully-blank rows AND the next non-blank row looks like a
    header. Returns [(start_row_idx, rows), ...] (start index is absolute in
    the sheet, needed to map merged-cell ranges)."""
    blocks: list[tuple[int, list]] = []
    cur_start, cur, blank_run = 0, [], 0
    for i, r in enumerate(raw_rows):
        if all(_cell_str(c) == "" for c in r):
            blank_run += 1
            cur.append(r)
            continue
        if (blank_run >= 2 and _is_headerish(r)
                and any(any(_cell_str(c) for c in rr) for rr in cur)):
            blocks.append((cur_start, cur))
            cur_start, cur = i, []
        blank_run = 0
        cur.append(r)
    if any(any(_cell_str(c) for c in rr) for rr in cur):
        blocks.append((cur_start, cur))
    return blocks or [(0, raw_rows)]


def _read_merges(path: str) -> dict:
    """{sheet_name: [(r1, c1, r2, c2), ...]} straight from the xlsx zip.
    read_only workbooks don't expose merged_cells and a full openpyxl load is
    too slow for the read path, so parse the <mergeCell> refs out of each
    sheet's XML directly. Best-effort — {} on any surprise."""
    import zipfile
    from xml.etree import ElementTree as ET
    from openpyxl.utils import range_boundaries
    out: dict = {}
    try:
        with zipfile.ZipFile(path) as z:
            wbxml = ET.fromstring(z.read("xl/workbook.xml"))
            rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            rid_target = {rel.get("Id"): rel.get("Target") for rel in rels}
            main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            rns = ("{http://schemas.openxmlformats.org/officeDocument/2006/"
                   "relationships}")
            for sh in wbxml.iter(f"{main}sheet"):
                name = sh.get("name")
                target = rid_target.get(sh.get(f"{rns}id")) or ""
                if target.startswith("/"):
                    target = target[1:]
                elif not target.startswith("xl/"):
                    target = "xl/" + target
                try:
                    sxml = z.read(target)
                except KeyError:
                    continue
                merges = []
                for m in re.finditer(rb'<mergeCell ref="([A-Z0-9:$]+)"', sxml):
                    c1, r1, c2, r2 = range_boundaries(
                        m.group(1).decode().replace("$", ""))
                    merges.append((r1, c1, r2, c2))
                if merges:
                    out[name] = merges
    except Exception:
        return {}
    return out


def _detect_header_row(rows, scan: int = 10) -> int:
    """Index of the header row within the first `scan` rows: the first row
    whose non-empty cells are ≥60% strings (labels, not data). Falls back to
    the first non-empty row. Deterministic — no LLM."""
    fallback = None
    for i, r in enumerate(rows[:scan]):
        vals = [v for v in r if v is not None and str(v).strip() != ""]
        if not vals:
            continue
        if fallback is None:
            fallback = i
        str_ratio = sum(isinstance(v, str) for v in vals) / len(vals)
        if str_ratio >= 0.6:
            return i
    return fallback if fallback is not None else 0


def _resolve_input_path(raw: str) -> str:
    """Input files (attachments, artifacts) may be given relative — resolve
    against the session artifact folder first, then cwd. Reads are not
    restricted to the artifact folder (mirrors read_document)."""
    from engine.tools.file_tools import _resolve_artifact_dir
    p = os.path.expanduser(raw or "")
    if os.path.isabs(p):
        return p
    art_dir, _ = _resolve_artifact_dir()
    if art_dir:
        cand = os.path.join(art_dir, p)
        if os.path.exists(cand):
            return cand
    return os.path.abspath(p)


def _load_grids(path: str, sheet: str | None = None,
                formulas: bool = False) -> list[dict]:
    """Load a workbook/CSV into plain grids:
    [{name, header:[str], rows:[[...]], header_row_idx}].
    Header row is detected, placeholder columns trimmed (the v9.261.0 logic,
    shared with doc_convert), empty header cells named col<N>. v2: a sheet
    with several table blocks yields several grids (`<sheet>`, `<sheet>_2`,
    …), and a merged two-row header composes into "Top / Sub" column names."""
    ext = os.path.splitext(path)[1].lower()
    if ext in _LEGACY_EXTS:
        # .xls/.ods → cached LibreOffice conversion, then the normal path.
        path = _legacy_to_xlsx(path)
        ext = ".xlsx"
    grids = []
    if ext in _CSV_EXTS:
        delim = "\t" if ext == ".tsv" else None
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            sample = f.read(8192)
            f.seek(0)
            if delim is None:
                try:
                    delim = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
                except csv.Error:
                    delim = ";" if sample.count(";") > sample.count(",") else ","
            raw_rows = [[_coerce_csv_value(c) for c in row]
                        for row in csv.reader(f, delimiter=delim)]
        grids.extend(_grids_from_rows(
            os.path.splitext(os.path.basename(path))[0], raw_rows, []))
        return grids

    import openpyxl
    merges = _read_merges(path)
    # formulas=True (diff compare='formulas'): formula CELLS yield their
    # formula string instead of the cached value.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=not formulas)
    try:
        for ws in wb.worksheets:
            if sheet and ws.title != sheet:
                continue
            raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            grids.extend(_grids_from_rows(ws.title, raw_rows,
                                          merges.get(ws.title) or []))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if sheet and not grids:
        raise ValueError(f"sheet '{sheet}' not found in {os.path.basename(path)}")
    return grids


def _grids_from_rows(name: str, raw_rows: list, merges: list) -> list[dict]:
    out = []
    for bi, (start, rows) in enumerate(_split_table_blocks(raw_rows)):
        g = _finish_grid(name if bi == 0 else f"{name}_{bi + 1}", rows,
                         merges=merges, row_offset=start)
        if g is not None:
            g["sheet_title"] = name  # real worksheet name (block grids differ)
            out.append(g)
    return out


def _compose_merged_header(raw_rows, hdr_idx, merges, row_offset):
    """v2: two-row hierarchical headers. When the detected header row carries
    multi-column MERGES (e.g. a "Q1" band over Umsatz|Marge) and the next row
    is also label-like, compose per-column names "Q1 / Umsatz" and start data
    one row later. Returns (header, data_start_idx) or None (normal path)."""
    abs_hdr_row = row_offset + hdr_idx + 1  # merges are 1-based absolute
    row_merges = [m for m in merges
                  if m[0] == abs_hdr_row and m[2] == abs_hdr_row
                  and m[3] > m[1]]
    if not row_merges:
        return None
    if hdr_idx + 2 >= len(raw_rows):
        return None  # no room for sub-header + data
    sub_row = raw_rows[hdr_idx + 1]
    if not _is_headerish(sub_row):
        return None
    top_row = raw_rows[hdr_idx]
    ncols = max(len(top_row), len(sub_row))
    header = []
    for ci in range(ncols):
        col1 = ci + 1
        top = None
        for (r1, c1, r2, c2) in row_merges:
            if c1 <= col1 <= c2:
                top = top_row[c1 - 1] if c1 - 1 < len(top_row) else None
                break
        if top is None:
            top = top_row[ci] if ci < len(top_row) else None
        sub = sub_row[ci] if ci < len(sub_row) else None
        ts, ss = _cell_str(top), _cell_str(sub)
        header.append(f"{ts} / {ss}" if ts and ss else (ss or ts))
    return header, hdr_idx + 2


def _finish_grid(name: str, raw_rows: list, merges: list | None = None,
                 row_offset: int = 0) -> dict | None:
    if not any(any(_cell_str(c) for c in r) for r in raw_rows):
        return None
    hdr_idx = _detect_header_row(raw_rows)
    composed = _compose_merged_header(raw_rows, hdr_idx, merges or [],
                                      row_offset)
    if composed:
        header, data_start = composed
    else:
        header, data_start = list(raw_rows[hdr_idx]), hdr_idx + 1
    data_rows = raw_rows[data_start:]
    header, data_rows = doc_convert._trim_placeholder_columns(
        header, data_rows, _cell_str)
    names = []
    for i, h in enumerate(header):
        hs = _cell_str(h)
        names.append(hs if hs else f"col{i + 1}")
    n = len(names)
    norm_rows = []
    row_nums = []  # absolute 1-based sheet row per kept data row (grid edit)
    for ri, r in enumerate(data_rows):
        row = list(r[:n]) + [None] * (n - len(r))
        if all(_cell_str(c) == "" for c in row):
            continue  # skip blank rows
        norm_rows.append(row)
        row_nums.append(row_offset + data_start + ri + 1)
    return {"name": name, "header": names, "rows": norm_rows,
            "header_row_idx": hdr_idx, "row_nums": row_nums}


# ---------------------------------------------------------------------------
# Named query results (v2 pipeline handles)
# ---------------------------------------------------------------------------
# xlsx_query save_as='x' stores its full result per SESSION; later calls
# reference it as path/file "result:x" (query FROM it, create a sheet from it)
# — multi-step pipelines without re-querying or re-reading the source files.
# In-memory only (dies with the server — handles are cheap to regenerate).

_RESULT_HANDLES: dict[str, dict] = {}
HANDLE_MAX_PER_SESSION = 8
HANDLE_MAX_CELLS = 500_000
_HANDLE_PREFIX = "result:"


def _handles_bucket() -> dict:
    sid = get_request_context().current_session_id or "global"
    return _RESULT_HANDLES.setdefault(sid, {})


def _store_handle(name: str, headers: list, rows: list) -> str:
    name = _sanitize_name(name, set())
    n_cells = len(headers) * max(1, len(rows))
    if n_cells > HANDLE_MAX_CELLS:
        raise ValueError(
            f"result too large to store ({n_cells:,} cells > "
            f"{HANDLE_MAX_CELLS:,}) — narrow the SELECT or use out='name.csv'")
    bucket = _handles_bucket()
    bucket[name] = {"header": headers, "rows": rows}
    while len(bucket) > HANDLE_MAX_PER_SESSION:
        bucket.pop(next(iter(bucket)))  # evict oldest
    return name


def _grid_from_handle(token: str) -> dict:
    name = token[len(_HANDLE_PREFIX):].strip()
    h = _handles_bucket().get(_sanitize_name(name, set()))
    if h is None:
        stored = ", ".join(_handles_bucket().keys()) or "(none)"
        raise FileNotFoundError(
            f"no stored result named '{name}' in this session — stored: "
            f"{stored}. Save one first with xlsx_query save_as='{name}'.")
    return {"name": _sanitize_name(name, set()), "header": h["header"],
            "rows": h["rows"], "header_row_idx": 0}


def _is_handle(p) -> bool:
    return isinstance(p, str) and p.startswith(_HANDLE_PREFIX)


# ---------------------------------------------------------------------------
# Name sanitization + SQLite session
# ---------------------------------------------------------------------------

def _sanitize_name(s: str, used: set) -> str:
    t = str(s).strip().translate(_UMLAUT_MAP)
    t = re.sub(r"[^0-9a-zA-Z_]+", "_", t).strip("_").lower() or "t"
    if t[0].isdigit():
        t = "t_" + t
    base, n = t, 2
    while t in used:
        t = f"{base}_{n}"
        n += 1
    used.add(t)
    return t


def _to_sql_value(v):
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, datetime.datetime):
        # Midnight timestamps are Excel "dates" — keep them date-shaped so
        # WHERE datum = '2026-01-02' works the way the model expects.
        if v.hour == v.minute == v.second == 0 and v.microsecond == 0:
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, (datetime.date, datetime.time)):
        return v.isoformat()
    return v


def _build_sqlite(paths: list[str], sheet: str | None = None,
                  max_rows: int = QUERY_MAX_TOTAL_ROWS):
    """Load every sheet of every file into ONE in-memory SQLite DB.
    Returns (conn, tables) — tables = [{table, file, sheet, columns:
    [(sql_name, orig_name)], rows}]. With >1 file, table names are prefixed
    with the file stem so same-named sheets don't collide (and cross-file
    joins read naturally: orders_alt vs orders_neu)."""
    total_rows = 0
    loaded = []  # (file_or_handle, grid)
    for p in paths:
        if _is_handle(p):
            g = _grid_from_handle(p)
            total_rows += len(g["rows"])
            loaded.append((p, g))
            continue
        rp = _resolve_input_path(p)
        if not os.path.isfile(rp):
            raise FileNotFoundError(f"File not found: {p}")
        size_mb = os.path.getsize(rp) / (1024 * 1024)
        # v3: big files ARE loadable when scoped to one sheet (read_only
        # streaming + the row cap bound the memory); unscoped stays refused.
        if size_mb > QUERY_MAX_FILE_MB and not sheet:
            raise ValueError(
                f"{os.path.basename(rp)} is {size_mb:.0f} MB (> {QUERY_MAX_FILE_MB} MB). "
                f"Pass sheet='<name>' to load a single sheet, or split the file.")
        for g in _load_grids(rp, sheet=sheet):
            total_rows += len(g["rows"])
            if total_rows > max_rows:
                raise ValueError(
                    f"More than {max_rows:,} rows across the loaded "
                    f"sheets. Pass sheet='<name>' to load a single sheet.")
            loaded.append((rp, g))

    conn = sqlite3.connect(":memory:")
    used_tables: set = set()
    multi = len(paths) > 1
    tables = []
    from engine.context import report_tool_progress
    for ti, (rp, g) in enumerate(loaded):
        if total_rows > HUGE_SHEET_ROWS:
            report_tool_progress(phase="xlsx_query:load", current=ti,
                                 total=len(loaded), note=g["name"])
        if _is_handle(rp):
            # A stored result IS its own identity — never file-stem-prefixed.
            raw_name = g["name"]
        else:
            stem = os.path.splitext(os.path.basename(rp))[0]
            raw_name = f"{stem}_{g['name']}" if multi else g["name"]
        tname = _sanitize_name(raw_name, used_tables)
        used_cols: set = set()
        cols = [(_sanitize_name(h, used_cols), h) for h in g["header"]]
        col_sql = ", ".join(f'"{c}"' for c, _ in cols)
        conn.execute(f'CREATE TABLE "{tname}" ({col_sql})')
        ph = ", ".join(["?"] * len(cols))
        conn.executemany(
            f'INSERT INTO "{tname}" VALUES ({ph})',
            ([_to_sql_value(v) for v in row] for row in g["rows"]))
        tables.append({"table": tname, "file": os.path.basename(rp),
                       "sheet": g["name"], "columns": cols,
                       "rows": len(g["rows"])})
    conn.commit()
    # Read-only from here: PRAGMA first (the authorizer would deny it), then
    # deny everything except reading — belt and braces on top of the
    # SELECT/WITH prefix check in tool_xlsx_query.
    conn.execute("PRAGMA query_only=ON")
    allowed = {sqlite3.SQLITE_SELECT, sqlite3.SQLITE_READ,
               sqlite3.SQLITE_FUNCTION, sqlite3.SQLITE_RECURSIVE}

    def _authorizer(action, *_):
        return sqlite3.SQLITE_OK if action in allowed else sqlite3.SQLITE_DENY

    conn.set_authorizer(_authorizer)
    return conn, tables


def _schema_echo(tables: list[dict]) -> str:
    lines = ["Tables for xlsx_query:"]
    for t in tables:
        cols = ", ".join(c for c, _ in t["columns"])
        lines.append(f"- {t['table']}({cols}) — {t['rows']} rows "
                     f"[{t['file']} / {t['sheet']}]")
    return "\n".join(lines)


def _check_select_only(sql: str) -> str | None:
    """SELECT/WITH single statements only. Returns an error message or None."""
    body = (sql or "").strip().rstrip(";").strip()
    if not body:
        return "empty sql"
    if not re.match(r"(?is)^(select|with)\b", body):
        return "only a single SELECT statement is allowed (start with SELECT or WITH)"
    if ";" in body:
        return ("only ONE statement is allowed — remove the ';' and send a "
                "single SELECT")
    return None


def _markdown_table(headers: list[str], rows: list) -> str:
    def esc(v):
        return "" if v is None else str(v).replace("|", "\\|").replace("\n", " ")
    out = ["| " + " | ".join(esc(h) for h in headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        out.append("| " + " | ".join(esc(c) for c in r) + " |")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# xlsx_inspect
# ---------------------------------------------------------------------------

def _infer_col_type(values) -> str:
    kinds = set()
    for v in values:
        if isinstance(v, bool):
            kinds.add("bool")
        elif isinstance(v, int):
            kinds.add("int")
        elif isinstance(v, float):
            kinds.add("float")
        elif isinstance(v, (datetime.datetime, datetime.date)):
            kinds.add("date")
        else:
            kinds.add("text")
    if not kinds:
        return "empty"
    if kinds <= {"int", "bool"}:
        return "int"
    if kinds <= {"int", "float", "bool"}:
        return "number"
    if kinds == {"date"}:
        return "date"
    return "text" if len(kinds) > 1 or "text" in kinds else kinds.pop()


def _fmt_sample(v) -> str:
    s = _cell_str(_to_sql_value(v))
    return s[:40] + "…" if len(s) > 40 else s


def _workbook_extras(path: str, deep: bool = False) -> dict:
    """Merged ranges, named ranges and formula count need extra passes that
    the values-grid can't provide (data_only=True hides formulas; read_only
    hides merges). `deep` also collects the formula STRINGS (capped) for the
    v2 pattern/dependency analysis. Best-effort — inspection must not fail on
    exotic files."""
    ext = os.path.splitext(path)[1].lower()
    out = {"merged": {}, "named_ranges": [], "formulas": {},
           "formula_list": {}}
    if ext not in _XLSX_EXTS:
        return out
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        for ws in wb.worksheets:
            n = 0
            flist = []
            for row in ws.iter_rows(values_only=True):
                for c in row:
                    if isinstance(c, str) and c.startswith("="):
                        n += 1
                        if deep and len(flist) < 2000:
                            flist.append(c)
            out["formulas"][ws.title] = n
            if deep:
                out["formula_list"][ws.title] = flist
        try:
            out["named_ranges"] = list(wb.defined_names.keys())
        except Exception:
            pass
        wb.close()
    except Exception:
        pass
    out["merged"] = {s: len(v) for s, v in _read_merges(path).items()}
    return out


def _deep_grid_checks(g: dict) -> list[str]:
    """v2 deep-mode data-quality findings for one grid: fully-duplicated
    rows + numeric outliers (3×IQR fence). Deterministic, no LLM."""
    lines = []
    rows = g["rows"]
    if not rows:
        return lines
    keys = [tuple(_cell_str(_to_sql_value(v)) for v in r) for r in rows]
    n_dup = len(keys) - len(set(keys))
    if n_dup:
        lines.append(f"- {n_dup} fully duplicated row(s)")
    for ci, name in enumerate(g["header"]):
        vals = sorted(v for r in rows if ci < len(r)
                      for v in [r[ci]]
                      if isinstance(v, (int, float)) and not isinstance(v, bool))
        if len(vals) < 8:
            continue
        q1 = vals[len(vals) // 4]
        q3 = vals[(3 * len(vals)) // 4]
        iqr = q3 - q1
        if iqr <= 0:
            continue
        lo, hi = q1 - 3 * iqr, q3 + 3 * iqr
        outliers = [v for v in vals if v < lo or v > hi]
        if outliers:
            ex = ", ".join(str(v) for v in outliers[:3])
            lines.append(f"- `{name}`: {len(outliers)} outlier value(s) "
                         f"beyond 3×IQR (e.g. {ex})")
    return lines


def _deep_orphan_checks(sheet_infos: list[dict]) -> list[str]:
    """v2 deep-mode referential check on join-key candidates: keys present on
    one side but not the other (the master-without-details / orphan-detail
    finding — 'Summe stimmt nicht' usually starts here)."""
    lines = []
    for i in range(len(sheet_infos)):
        for j in range(i + 1, len(sheet_infos)):
            a, b = sheet_infos[i], sheet_infos[j]
            cols_b = {c["sql"]: c for c in b["cols"]}
            for ca in a["cols"]:
                cb = cols_b.get(ca["sql"])
                if cb is None:
                    continue
                va = ca.get("value_full") or ca["value_sample"]
                vb = cb.get("value_full") or cb["value_sample"]
                if not va or not vb:
                    continue
                overlap = len(va & vb) / min(len(va), len(vb))
                if overlap < JOIN_KEY_MIN_OVERLAP:
                    continue
                only_a, only_b = va - vb, vb - va
                if only_a:
                    ex = ", ".join(str(x) for x in list(only_a)[:3])
                    lines.append(
                        f"- `{ca['name']}`: {len(only_a)} value(s) in "
                        f"{a['table']} missing from {b['table']} (e.g. {ex})")
                if only_b:
                    ex = ", ".join(str(x) for x in list(only_b)[:3])
                    lines.append(
                        f"- `{ca['name']}`: {len(only_b)} value(s) in "
                        f"{b['table']} missing from {a['table']} (e.g. {ex})")
    return lines


def _formula_analysis(formula_list: dict) -> list[str]:
    """v2 formula view: per sheet the dominant formula patterns (digits
    normalised to N) + which OTHER sheets its formulas reference — a compact
    dependency map of what the workbook computes."""
    from collections import Counter
    lines = []
    for sheet, flist in formula_list.items():
        if not flist:
            continue
        pats = Counter(re.sub(r"\d+", "N", f)[:80] for f in flist)
        top = "; ".join(f"`{p}` ×{c}" for p, c in pats.most_common(3))
        lines.append(f"- {sheet}: {len(flist)} formulas — top patterns: {top}")
        refs: Counter = Counter()
        for f in flist:
            for m in re.finditer(r"(?:'([^']+)'|([A-Za-z_][\w .]*))!", f):
                tgt = (m.group(1) or m.group(2)).strip()
                if tgt and tgt != sheet:
                    refs[tgt] += 1
        for tgt, c in refs.most_common(5):
            lines.append(f"  - references sheet `{tgt}` in {c} formula(s)")
    return lines


def _join_key_candidates(sheet_infos: list[dict]) -> list[str]:
    """Same-named columns across sheets whose value sets overlap — the model
    gets the join key handed to it instead of guessing (the marktorder case:
    MARKTORDERNUMMER appears in both sheets with ~full overlap)."""
    hits = []
    for i in range(len(sheet_infos)):
        for j in range(i + 1, len(sheet_infos)):
            a, b = sheet_infos[i], sheet_infos[j]
            cols_b = {c["sql"]: c for c in b["cols"]}
            for ca in a["cols"]:
                cb = cols_b.get(ca["sql"])
                if cb is None:
                    continue
                va, vb = ca["value_sample"], cb["value_sample"]
                if not va or not vb:
                    continue
                overlap = len(va & vb) / min(len(va), len(vb))
                if overlap >= JOIN_KEY_MIN_OVERLAP:
                    hits.append(
                        f"- `{ca['name']}`: {a['table']} ↔ {b['table']} "
                        f"(value overlap {overlap:.0%}) — likely JOIN key")
    return hits


def _inspect_report(paths: list, sheet: str | None = None,
                    deep: bool = False) -> str:
    """The xlsx_inspect report body (raw, no GDPR wrap / envelope) — shared by
    tool_xlsx_inspect and the project-sync profile miner (v9.264.0), which
    files one profile per project spreadsheet into the MemPalace wing so
    structure questions answer from retrieval without a live inspect call."""
    if True:  # keep the original body's indentation
        parts = []
        sheet_infos = []
        used_tables: set = set()
        multi = len(paths) > 1
        formula_lists: dict = {}
        for p in paths:
            if _is_handle(p):
                rp = p
                extras = {"merged": {}, "named_ranges": [], "formulas": {},
                          "formula_list": {}}
                grids = [_grid_from_handle(p)]
            else:
                rp = _resolve_input_path(p)
                if not os.path.isfile(rp):
                    raise FileNotFoundError(f"File not found: {p}")
                extras = _workbook_extras(rp, deep=deep)
                grids = _load_grids(rp, sheet=sheet)
            if deep:
                for sname, flist in (extras.get("formula_list") or {}).items():
                    formula_lists[f"{os.path.basename(rp)} / {sname}"] = flist
            if not grids:
                parts.append(f"# {os.path.basename(rp)}\n\n(no data found)")
                continue
            parts.append(f"# {os.path.basename(rp)}")
            for g in grids:
                stem = os.path.splitext(os.path.basename(rp))[0]
                raw_name = f"{stem}_{g['name']}" if multi else g["name"]
                tname = _sanitize_name(raw_name, used_tables)
                used_cols: set = set()
                n_rows = len(g["rows"])
                parts.append(
                    f"\n## Sheet: {g['name']} — {n_rows} data rows × "
                    f"{len(g['header'])} columns (header row "
                    f"{g['header_row_idx'] + 1})")
                extra_bits = []
                if extras["merged"].get(g["name"]):
                    extra_bits.append(f"{extras['merged'][g['name']]} merged ranges")
                if extras["formulas"].get(g["name"]):
                    extra_bits.append(f"{extras['formulas'][g['name']]} formula cells")
                if extra_bits:
                    parts.append("_" + ", ".join(extra_bits) + "_")
                from openpyxl.utils import get_column_letter
                col_rows = []
                cols_info = []
                for ci, h in enumerate(g["header"]):
                    values = [r[ci] for r in g["rows"] if ci < len(r)]
                    non_null = [v for v in values if _cell_str(v) != ""]
                    ctype = _infer_col_type(non_null[:INSPECT_DISTINCT_SAMPLE])
                    distinct_sample = set()
                    for v in non_null[:INSPECT_DISTINCT_SAMPLE]:
                        distinct_sample.add(_to_sql_value(v))
                    n_distinct = len(distinct_sample)
                    distinct_str = (f"{n_distinct}"
                                    if len(non_null) <= INSPECT_DISTINCT_SAMPLE
                                    else f"{n_distinct}+")
                    minmax = ""
                    if ctype in ("int", "number", "date") and non_null:
                        try:
                            lo = min(_to_sql_value(v) for v in non_null)
                            hi = max(_to_sql_value(v) for v in non_null)
                            minmax = f"{_fmt_sample(lo)} – {_fmt_sample(hi)}"
                        except TypeError:
                            pass
                    samples = []
                    for v in non_null:
                        s = _fmt_sample(v)
                        if s not in samples:
                            samples.append(s)
                        if len(samples) >= INSPECT_SAMPLE_VALUES:
                            break
                    sql_name = _sanitize_name(h, used_cols)
                    col_rows.append([
                        get_column_letter(ci + 1), h, sql_name, ctype,
                        len(values) - len(non_null), distinct_str, minmax,
                        ", ".join(samples)])
                    ci_info = {
                        "name": h, "sql": sql_name,
                        "value_sample": set(
                            _to_sql_value(v) for v in
                            non_null[:JOIN_KEY_OVERLAP_SAMPLE])}
                    if deep:
                        ci_info["value_full"] = set(
                            _to_sql_value(v) for v in non_null)
                    cols_info.append(ci_info)
                parts.append(_markdown_table(
                    ["Col", "Name", "SQL name", "Type", "Nulls", "Distinct",
                     "Min–Max", "Samples"], col_rows))
                if deep:
                    checks = _deep_grid_checks(g)
                    if checks:
                        parts.append("\n**Data quality:**\n" + "\n".join(checks))
                sheet_infos.append({"table": tname, "cols": cols_info,
                                    "file": os.path.basename(rp),
                                    "sheet": g["name"], "rows": n_rows,
                                    "columns": [(c["sql"], c["name"])
                                                for c in cols_info]})
            if extras["named_ranges"]:
                parts.append("\n_Named ranges: "
                             + ", ".join(extras["named_ranges"][:20]) + "_")
        joins = _join_key_candidates(sheet_infos)
        if joins:
            parts.append("\n## Join-key candidates\n" + "\n".join(joins))
        if deep:
            orphans = _deep_orphan_checks(sheet_infos)
            if orphans:
                parts.append("\n## Referential findings (deep)\n"
                             + "\n".join(orphans))
            formulas = _formula_analysis(formula_lists)
            if formulas:
                parts.append("\n## Formula map (deep)\n" + "\n".join(formulas))
        parts.append("\n## " + _schema_echo(sheet_infos))
        return "\n".join(parts)


def tool_xlsx_inspect(args: dict) -> str:
    import brain as _brain
    paths = args.get("paths") or ([args["path"]] if args.get("path") else [])
    if not paths:
        return _err("xlsx_inspect: 'path' (or 'paths') is required")
    try:
        report = _inspect_report(paths, sheet=args.get("sheet"),
                                 deep=bool(args.get("deep")))
        src = "xlsx:" + ",".join(os.path.basename(_resolve_input_path(p))
                                 for p in paths)
        report = _brain._gdpr_anon_tool_text(report, src)
        return _ok({"paths": paths, "report": report})
    except FileNotFoundError as e:
        return _err(str(e))
    except Exception as e:
        return _err(f"xlsx_inspect: {type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# xlsx_query
# ---------------------------------------------------------------------------

def tool_xlsx_query(args: dict) -> str:
    import brain as _brain
    paths = args.get("paths") or ([args["path"]] if args.get("path") else [])
    if not paths:
        return _err("xlsx_query: 'path' (or 'paths') is required")
    sql = args.get("sql") or ""
    sel_err = _check_select_only(sql)
    if sel_err:
        return _err(f"xlsx_query: {sel_err}")
    sql = sql.strip().rstrip(";")
    try:
        conn, tables = _build_sqlite(paths, sheet=args.get("sheet"))
    except (FileNotFoundError, ValueError) as e:
        return _err(f"xlsx_query: {e}")
    except Exception as e:
        return _err(f"xlsx_query: {type(e).__name__}: {e}")
    try:
        try:
            cur = conn.execute(sql)
        except sqlite3.Error as e:
            # The self-correction loop: echo the real schema so the model can
            # fix identifiers in one round instead of guessing again.
            return _err(f"xlsx_query: SQL error: {e}\n\n{_schema_echo(tables)}")
        headers = [d[0] for d in cur.description or []]
        rows = cur.fetchall()
        row_count = len(rows)
        display = rows[:QUERY_DISPLAY_ROWS]
        md = _markdown_table(headers, display)
        if row_count > QUERY_DISPLAY_ROWS:
            md += (f"\n\n_({row_count:,} rows total, showing first "
                   f"{QUERY_DISPLAY_ROWS} — pass out='name.csv' to save the "
                   f"full result)_")
        out_info = None
        out_name = (args.get("out") or "").strip()
        if out_name:
            from engine.tools.file_tools import _enforce_artifact_path
            if not out_name.lower().endswith(".csv"):
                out_name += ".csv"
            out_path, perr = _enforce_artifact_path(out_name, "xlsx_query")
            if perr:
                return perr
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                w.writerows(rows)
            agent = get_request_context().current_agent
            _brain._after_file_write(
                out_path, "created", agent.agent_id if agent else "main")
            out_info = {"path": out_path, "rows": row_count}
        src = "xlsx_query:" + ",".join(os.path.basename(p) for p in paths)
        md = _brain._gdpr_anon_tool_text(md, src)
        res = {"row_count": row_count, "result": md}
        if out_info:
            res["saved"] = out_info
        save_as = (args.get("save_as") or "").strip()
        if save_as:
            hname = _store_handle(save_as, headers, [list(r) for r in rows])
            res["saved_as"] = hname
            res["note"] = (f"full result stored for this session — reference "
                           f"it as 'result:{hname}' in xlsx_query paths or an "
                           f"xlsx_create/xlsx_edit source.file")
        return _ok(res)
    except Exception as e:
        return _err(f"xlsx_query: {type(e).__name__}: {e}")
    finally:
        try:
            conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# xlsx_create — declarative spec → styled workbook
# ---------------------------------------------------------------------------

def _hex(color: str, fallback: str) -> str:
    c = (color or fallback or "").lstrip("#").upper()
    return c if re.fullmatch(r"[0-9A-F]{6}", c) else fallback.lstrip("#").upper()


def _style_kit(style: dict) -> dict:
    """Derive the small set of openpyxl style objects from a doc-style preset
    (the same YAML presets docx/html use — one house look everywhere)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    colors = style.get("colors") or {}
    docx = style.get("docx") or {}
    header_bg = _hex(colors.get("table_header_bg"), "#1F3864")
    header_fg = _hex(colors.get("table_header_text"), "#FFFFFF")
    zebra = _hex(docx.get("zebra_fill"), "#EDF1F8")
    thin = Side(style="thin", color="B4B4B4")
    return {
        "header_font": Font(bold=True, color=header_fg,
                            name=(style.get("fonts") or {}).get("body", "Calibri")),
        "header_fill": PatternFill("solid", start_color=header_bg,
                                   end_color=header_bg),
        "header_align": Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
        "zebra_fill": PatternFill("solid", start_color=zebra, end_color=zebra),
        "master_fill": PatternFill("solid", start_color=zebra, end_color=zebra),
        "detail_fill": PatternFill("solid", start_color="F7F7F7",
                                   end_color="F7F7F7"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _clean_sheet_name(name: str, used: set) -> str:
    t = re.sub(r"[\[\]:*?/\\]", "", str(name or "Sheet")).strip() or "Sheet"
    t = t[:31]
    base, n = t, 2
    while t in used:
        suffix = f"_{n}"
        t = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(t)
    return t


def _resolve_source(source: dict) -> tuple[list[str], list[list]]:
    """A create/edit data source: {file, sheet?, sql?} → (header, rows).
    Data flows file→file server-side; the model never transcribes rows."""
    if not isinstance(source, dict) or not source.get("file"):
        raise ValueError("source needs {file, sheet?, sql?}")
    fpath = source["file"]
    sql = (source.get("sql") or "").strip()
    if sql:
        sel_err = _check_select_only(sql)
        if sel_err:
            raise ValueError(f"source.sql: {sel_err}")
        conn, tables = _build_sqlite([fpath], sheet=source.get("sheet"),
                                     max_rows=CREATE_MAX_SOURCE_ROWS)
        try:
            try:
                cur = conn.execute(sql.rstrip(";"))
            except sqlite3.Error as e:
                raise ValueError(f"source.sql error: {e}\n{_schema_echo(tables)}")
            headers = [d[0] for d in cur.description or []]
            return headers, [list(r) for r in cur.fetchall()]
        finally:
            conn.close()
    if _is_handle(fpath):
        g = _grid_from_handle(fpath)
        return g["header"], g["rows"]
    rp = _resolve_input_path(fpath)
    if not os.path.isfile(rp):
        raise FileNotFoundError(f"File not found: {fpath}")
    grids = _load_grids(rp, sheet=source.get("sheet"))
    if not grids:
        raise ValueError(f"no data found in {os.path.basename(rp)}")
    g = grids[0]
    return g["header"], g["rows"]


def _sheet_data(sheet_spec: dict) -> tuple[list[str], list[list]]:
    """Resolve the (exactly one) data source of a create-spec sheet."""
    cols = sheet_spec.get("columns") or []
    col_names = [c.get("name") for c in cols if isinstance(c, dict)]
    if sheet_spec.get("rows") is not None:
        rows = sheet_spec["rows"]
        n_cells = sum(len(r) for r in rows if isinstance(r, (list, tuple)))
        if n_cells > CREATE_INLINE_MAX_CELLS:
            raise ValueError(
                f"inline rows too large ({n_cells} cells > "
                f"{CREATE_INLINE_MAX_CELLS}). Do NOT copy data into the spec — "
                f"use source: {{file, sheet?, sql?}} so the server moves the "
                f"data itself.")
        if col_names and all(col_names):
            return list(col_names), [list(r) for r in rows]
        if not rows:
            # An intentionally empty sheet (write_document markdown sections
            # without a table produce this) — render as a bare sheet.
            return [], []
        return [str(c) for c in rows[0]], [list(r) for r in rows[1:]]
    if sheet_spec.get("source"):
        header, rows = _resolve_source(sheet_spec["source"])
        if col_names and all(col_names):
            idx = _column_indices(header, col_names)
            rows = [[r[i] for i in idx] for r in rows]
            header = list(col_names)
        return header, rows
    raise ValueError("sheet needs exactly one of: rows, source, master_detail")


def _column_indices(header: list[str], wanted: list[str]) -> list[int]:
    lower = {str(h).strip().lower(): i for i, h in enumerate(header)}
    idx = []
    for w in wanted:
        i = lower.get(str(w).strip().lower())
        if i is None:
            raise ValueError(
                f"column '{w}' not found — available: {', '.join(map(str, header))}")
        idx.append(i)
    return idx


def _apply_header_row(ws, header, kit, row=1):
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=row, column=ci, value=str(h))
        c.font = kit["header_font"]
        c.fill = kit["header_fill"]
        c.alignment = kit["header_align"]
        c.border = kit["border"]


def _auto_widths(ws, header, rows, col_specs=None):
    from openpyxl.utils import get_column_letter
    widths = {}
    for ci, h in enumerate(header):
        w = len(str(h))
        for r in rows[:200]:
            if ci < len(r) and r[ci] is not None:
                w = max(w, len(str(r[ci])))
        widths[ci] = min(60, w + 2)
    if col_specs:
        for ci, cs in enumerate(col_specs):
            if isinstance(cs, dict) and cs.get("width"):
                widths[ci] = cs["width"]
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(8, w)


def _apply_number_formats(ws, header, col_specs, first_row, last_row):
    if not col_specs:
        return
    by_name = {str(c.get("name", "")).strip().lower(): c
               for c in col_specs if isinstance(c, dict)}
    for ci, h in enumerate(header, 1):
        cs = by_name.get(str(h).strip().lower())
        fmt = _NUMBER_FORMATS.get((cs or {}).get("format") or "")
        if not fmt:
            continue
        for r in range(first_row, last_row + 1):
            ws.cell(row=r, column=ci).number_format = fmt


def _render_charts(ws, sheet_spec, header, n_data_rows, header_row=1):
    """Charts v1: bar|line|pie. v2 (v9.264.0): + scatter (labels column = X
    values) and area; `stacked: true` (bar/area/line); `secondary: [col]`
    puts those series on a right-hand Y axis (combo chart via the openpyxl
    second-axis recipe: axId 200 + crosses='max' + chart addition)."""
    charts = sheet_spec.get("charts") or []
    if not charts or n_data_rows == 0:
        return
    from openpyxl.chart import (BarChart, LineChart, PieChart, AreaChart,
                                ScatterChart, Reference, Series)
    from openpyxl.utils import get_column_letter
    kinds = {"bar": BarChart, "line": LineChart, "pie": PieChart,
             "area": AreaChart, "scatter": ScatterChart}
    anchor_col = len(header) + 2
    anchor_row = 2
    for ch in charts:
        if not isinstance(ch, dict):
            continue
        ctype = (ch.get("type") or "bar").lower()
        cls = kinds.get(ctype)
        if cls is None:
            raise ValueError(
                f"chart type '{ch.get('type')}' — use bar|line|pie|area|scatter")
        series_names = ch.get("series") or []
        label_name = ch.get("labels")
        if not series_names:
            raise ValueError("chart needs series: ['<column>', …]")
        secondary = {str(s).strip().lower() for s in (ch.get("secondary") or [])}
        prim = [s for s in series_names
                if str(s).strip().lower() not in secondary]
        sec = [s for s in series_names if str(s).strip().lower() in secondary]
        first, last = header_row + 1, header_row + n_data_rows
        chart = cls()
        chart.title = ch.get("title") or None
        if ctype == "scatter":
            if not label_name:
                raise ValueError("scatter needs labels: '<x column>'")
            xi = _column_indices(header, [label_name])[0]
            xref = Reference(ws, min_col=xi + 1, min_row=first, max_row=last)
            for sname in (prim or series_names):
                si = _column_indices(header, [sname])[0]
                yref = Reference(ws, min_col=si + 1, min_row=header_row,
                                 max_row=last)
                chart.series.append(Series(yref, xref, title_from_data=True))
        else:
            for sname in (prim or series_names):
                si = _column_indices(header, [sname])[0]
                chart.add_data(Reference(ws, min_col=si + 1,
                                         min_row=header_row, max_row=last),
                               titles_from_data=True)
            if label_name:
                li = _column_indices(header, [label_name])[0]
                chart.set_categories(
                    Reference(ws, min_col=li + 1, min_row=first, max_row=last))
            if ch.get("stacked") and ctype in ("bar", "area", "line"):
                chart.grouping = "stacked"
                if ctype == "bar":
                    chart.overlap = 100
            if sec and ctype != "pie":
                c2 = LineChart()
                for sname in sec:
                    si = _column_indices(header, [sname])[0]
                    c2.add_data(Reference(ws, min_col=si + 1,
                                          min_row=header_row, max_row=last),
                                titles_from_data=True)
                c2.y_axis.axId = 200
                chart.y_axis.crosses = "max"
                chart += c2
        pos = ch.get("position") or f"{get_column_letter(anchor_col)}{anchor_row}"
        ws.add_chart(chart, pos)
        anchor_row += 16  # stack subsequent charts below each other


def _render_conditional(ws, sheet_spec, header, n_data_rows, header_row=1):
    rules = sheet_spec.get("conditional") or []
    if not rules or n_data_rows == 0:
        return
    from openpyxl.formatting.rule import (ColorScaleRule, DataBarRule,
                                          CellIsRule)
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    first, last = header_row + 1, header_row + n_data_rows
    for rule in rules:
        if not isinstance(rule, dict):
            continue
        cols = rule.get("columns") or []
        spec = rule.get("rule")
        for name in cols:
            ci = _column_indices(header, [name])[0]
            letter = get_column_letter(ci + 1)
            rng = f"{letter}{first}:{letter}{last}"
            if spec == "color_scale":
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B"))
            elif spec == "data_bars":
                ws.conditional_formatting.add(rng, DataBarRule(
                    start_type="min", end_type="max", color="638EC6",
                    showValue=True))
            elif isinstance(spec, dict):
                op_map = {"lt": "lessThan", "gt": "greaterThan", "eq": "equal"}
                op = next((k for k in op_map if k in spec), None)
                if op is None:
                    raise ValueError(
                        "conditional rule dict needs lt|gt|eq, e.g. "
                        '{"lt": 0, "fill": "red"}')
                fill_hex = _COND_FILLS.get(
                    (spec.get("fill") or "red").lower(), _COND_FILLS["red"])
                val = spec[op]
                formula = [f'"{val}"' if isinstance(val, str) else str(val)]
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator=op_map[op], formula=formula,
                    fill=PatternFill("solid", start_color=fill_hex,
                                     end_color=fill_hex)))
            else:
                raise ValueError(
                    f"conditional rule '{spec}' — use \"color_scale\", "
                    f"\"data_bars\" or {{\"lt\"|\"gt\"|\"eq\": value, "
                    f"\"fill\": \"red|green|yellow\"}}")


def _apply_data_validation(ws, header, col_specs, first_row, last_row):
    """v2: column spec `choices: [...]` → an Excel list dropdown on that
    column's data range."""
    if not col_specs or last_row < first_row:
        return
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    by_name = {str(c.get("name", "")).strip().lower(): c
               for c in col_specs if isinstance(c, dict)}
    for ci, h in enumerate(header, 1):
        cs = by_name.get(str(h).strip().lower())
        choices = (cs or {}).get("choices")
        if not choices:
            continue
        formula = '"' + ",".join(str(c) for c in choices) + '"'
        if len(formula) > 255:
            raise ValueError(
                f"choices for '{h}' exceed Excel's 255-char list limit — "
                f"use fewer/shorter options")
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(ci)
        dv.add(f"{letter}{first_row}:{letter}{last_row}")


def _apply_print_setup(ws, sheet_spec):
    """v2: sheet spec `print: {orientation?, fit_width?, repeat_header?}`."""
    p = sheet_spec.get("print")
    if not isinstance(p, dict):
        return
    orient = (p.get("orientation") or "").lower()
    if orient in ("landscape", "portrait"):
        ws.page_setup.orientation = orient
    if p.get("fit_width"):
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    if p.get("repeat_header"):
        ws.print_title_rows = "1:1"


HUGE_SHEET_ROWS = 50_000
# v4: past this row count the WHOLE workbook is written in openpyxl's
# write_only streaming mode (constant memory) instead of the in-memory model.
WRITE_ONLY_ROWS = 100_000


def _render_spec_write_only(path: str, resolved: list, kit) -> dict:
    """v4 streaming writer for 100k+-row exports: openpyxl write_only keeps
    memory flat by serializing rows as they're appended (the regular model
    holds every cell object — GBs at 500k rows). Trade (documented in the
    result): styled header/freeze/column widths/totals survive; banded rows,
    per-cell number formats, charts, conditional formatting and validation
    don't exist in this mode. Only plain table sheets stream — master_detail/
    pivot are structural layouts and are rejected upstream."""
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter
    from engine.context import report_tool_progress
    wb = openpyxl.Workbook(write_only=True)
    used_names: set = set()
    out_sheets = []
    for sheet_spec, header, rows in resolved:
        name = _clean_sheet_name(sheet_spec.get("name"), used_names)
        ws = wb.create_sheet(name)
        # Widths + freeze must be set BEFORE the first append in write_only.
        widths = {}
        for ci, h in enumerate(header):
            w = len(str(h))
            for r in rows[:200]:
                if ci < len(r) and r[ci] is not None:
                    w = max(w, len(str(r[ci])))
            widths[ci] = min(60, max(8, w + 2))
        for cs_i, cs in enumerate(sheet_spec.get("columns") or []):
            if isinstance(cs, dict) and cs.get("width"):
                widths[cs_i] = cs["width"]
        for ci, w in widths.items():
            ws.column_dimensions[get_column_letter(ci + 1)].width = w
        if sheet_spec.get("freeze_header", True):
            ws.freeze_panes = "A2"
        hdr_cells = []
        for h in header:
            c = WriteOnlyCell(ws, value=str(h))
            c.font = kit["header_font"]
            c.fill = kit["header_fill"]
            c.alignment = kit["header_align"]
            hdr_cells.append(c)
        ws.append(hdr_cells)
        n_cols = len(header)
        for ri, row in enumerate(rows):
            if ri % 50_000 == 0:
                report_tool_progress(phase="xlsx_create:stream", current=ri,
                                     total=len(rows), note=name)
            ws.append([row[ci] if ci < len(row) else None
                       for ci in range(n_cols)])
        totals = sheet_spec.get("totals") or []
        if totals and rows:
            t_idx = _column_indices(header, totals)
            trow = [None] * n_cols
            if 0 not in t_idx:
                trow[0] = "Summe"
            for ci in t_idx:
                letter = get_column_letter(ci + 1)
                trow[ci] = f"=SUM({letter}2:{letter}{len(rows) + 1})"
            ws.append(trow)
        out_sheets.append({"name": name, "rows": len(rows)})
    wb.save(path)
    return {"sheets": out_sheets, "mode": "streaming",
            "note": ("streaming mode (>100k rows): styled header/freeze/"
                     "widths/totals kept; banded rows, number formats, "
                     "charts, conditional formatting skipped")}


def _render_table_sheet(ws, sheet_spec, kit, data=None) -> int:
    header, rows = data if data is not None else _sheet_data(sheet_spec)
    if not header:
        return 0  # intentionally empty sheet
    col_specs = sheet_spec.get("columns") or []
    _apply_header_row(ws, header, kit)
    banded = sheet_spec.get("banded", True)
    # v3 huge-sheet fast path: per-cell border/fill objects are the render
    # cost. Past the threshold, write bare values (styled header/freeze/
    # widths stay) — a 500k-row export must not take minutes.
    huge = len(rows) > HUGE_SHEET_ROWS
    from engine.context import report_tool_progress
    for ri, row in enumerate(rows):
        if huge and ri % 50_000 == 0:
            report_tool_progress(phase="xlsx_create", current=ri,
                                 total=len(rows), note=ws.title)
        for ci in range(len(header)):
            v = row[ci] if ci < len(row) else None
            c = ws.cell(row=ri + 2, column=ci + 1, value=v)
            if not huge:
                c.border = kit["border"]
                if banded and ri % 2 == 1:
                    c.fill = kit["zebra_fill"]
    n = len(rows)
    if not huge:  # per-cell number formats are part of the fast-path trade
        _apply_number_formats(ws, header, col_specs, 2, n + 1)
    totals = sheet_spec.get("totals") or []
    if totals and n:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        t_idx = _column_indices(header, totals)
        trow = n + 2
        first_cell = ws.cell(row=trow, column=1)
        if 0 not in t_idx:
            first_cell.value = "Summe"
        for ci in t_idx:
            letter = get_column_letter(ci + 1)
            ws.cell(row=trow, column=ci + 1,
                    value=f"=SUM({letter}2:{letter}{n + 1})")
        for ci in range(len(header)):
            c = ws.cell(row=trow, column=ci + 1)
            c.font = Font(bold=True)
            c.border = kit["border"]
        _apply_number_formats(ws, header, col_specs, trow, trow)
    if sheet_spec.get("freeze_header", True):
        ws.freeze_panes = "A2"
    if sheet_spec.get("autofilter") and n:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{n + 1}"
    _auto_widths(ws, header, rows, col_specs)
    _apply_data_validation(ws, header, col_specs, 2, n + 1)
    _apply_print_setup(ws, sheet_spec)
    _render_charts(ws, sheet_spec, header, n)
    _render_conditional(ws, sheet_spec, header, n)
    return n + (1 if totals and n else 0)


_PIVOT_AGGS = {"sum", "count", "avg", "min", "max"}


def _render_pivot_sheet(ws, sheet_spec, kit) -> int:
    """v3 pivot layout: a deterministic cross-tab — distinct `rows` values
    down, distinct `cols` values across, `values` aggregated per cell.
    pivot: {rows, cols?, values, agg?: sum|count|avg|min|max, source|rows}.
    Without `cols` it degrades to a grouped summary (one value column)."""
    pv = sheet_spec["pivot"]
    agg = (pv.get("agg") or "sum").lower()
    if agg not in _PIVOT_AGGS:
        raise ValueError(f"pivot.agg must be one of {sorted(_PIVOT_AGGS)}")
    if not pv.get("rows") or not pv.get("values"):
        raise ValueError("pivot needs rows: '<col>' and values: '<col>'")
    header, data = _sheet_data({"rows": pv.get("data"),
                                "source": pv.get("source")}
                               if pv.get("source") or pv.get("data") is not None
                               else sheet_spec)
    ri_ = _column_indices(header, [pv["rows"]])[0]
    vi_ = _column_indices(header, [pv["values"]])[0]
    ci_ = _column_indices(header, [pv["cols"]])[0] if pv.get("cols") else None

    def _key(v):
        return _cell_str(_to_sql_value(v)) or "(leer)"

    buckets: dict = {}
    col_keys: list = []
    row_keys: list = []
    for r in data:
        rk = _key(r[ri_] if ri_ < len(r) else None)
        ck = _key(r[ci_] if ci_ is not None and ci_ < len(r) else None) \
            if ci_ is not None else pv["values"]
        v = r[vi_] if vi_ < len(r) else None
        if rk not in row_keys:
            row_keys.append(rk)
        if ck not in col_keys:
            col_keys.append(ck)
        buckets.setdefault((rk, ck), []).append(v)
    row_keys.sort()
    col_keys.sort()

    def _aggregate(vals):
        nums = [v for v in vals
                if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if agg == "count":
            return len([v for v in vals if _cell_str(v) != ""])
        if not nums:
            return None
        if agg == "sum":
            return sum(nums)
        if agg == "avg":
            return round(sum(nums) / len(nums), 6)
        return min(nums) if agg == "min" else max(nums)

    out_header = [pv["rows"]] + col_keys
    _apply_header_row(ws, out_header, kit)
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for r_i, rk in enumerate(row_keys):
        c = ws.cell(row=r_i + 2, column=1, value=rk)
        c.border = kit["border"]
        c.font = bold
        for c_i, ck in enumerate(col_keys):
            cell = ws.cell(row=r_i + 2, column=c_i + 2,
                           value=_aggregate(buckets.get((rk, ck), [])))
            cell.border = kit["border"]
    n = len(row_keys)
    if pv.get("totals", True) and n and agg in ("sum", "count"):
        from openpyxl.utils import get_column_letter
        trow = n + 2
        ws.cell(row=trow, column=1, value="Gesamt").font = bold
        for c_i in range(len(col_keys)):
            letter = get_column_letter(c_i + 2)
            tc = ws.cell(row=trow, column=c_i + 2,
                         value=f"=SUM({letter}2:{letter}{n + 1})")
            tc.font = bold
        for c_i in range(len(out_header)):
            ws.cell(row=trow, column=c_i + 1).border = kit["border"]
    if sheet_spec.get("freeze_header", True):
        ws.freeze_panes = "B2"
    _auto_widths(ws, out_header, [[rk] for rk in row_keys])
    _apply_print_setup(ws, sheet_spec)
    return n


def _render_master_detail_sheet(ws, sheet_spec, kit) -> int:
    """The marktorder shape: one master row (tinted, bold key), its detail
    rows grouped beneath (outline level 1, collapsible), detail columns offset
    to the right of the master columns, one frozen combined header."""
    md = sheet_spec["master_detail"]
    key = md.get("key")
    if not key:
        raise ValueError("master_detail needs key: '<column name>'")
    m_header, m_rows = _resolve_source((md.get("master") or {}).get("source")
                                       or md.get("master"))
    d_header, d_rows = _resolve_source((md.get("detail") or {}).get("source")
                                       or md.get("detail"))
    m_cols = (md.get("master") or {}).get("columns") or m_header
    d_cols = ((md.get("detail") or {}).get("columns")
              or [h for h in d_header
                  if str(h).strip().lower() != str(key).strip().lower()])
    m_idx = _column_indices(m_header, m_cols)
    d_idx = _column_indices(d_header, d_cols)
    m_key = _column_indices(m_header, [key])[0]
    d_key = _column_indices(d_header, [key])[0]

    detail_by_key: dict = {}
    for r in d_rows:
        detail_by_key.setdefault(_cell_str(_to_sql_value(r[d_key])), []).append(r)

    combined = [m_header[i] for i in m_idx] + [d_header[i] for i in d_idx]
    _apply_header_row(ws, combined, kit)
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    bold = Font(bold=True)
    n_m = len(m_idx)
    row_no = 2
    key_out_idx = m_idx.index(m_key) if m_key in m_idx else None
    # v2 group subtotals: detail column names → output column positions.
    sub_out_cols = []
    for sname in md.get("subtotals") or []:
        di = _column_indices([d_header[i] for i in d_idx], [sname])[0]
        sub_out_cols.append(n_m + di + 1)
    for mr in m_rows:
        for ci, si in enumerate(m_idx):
            c = ws.cell(row=row_no, column=ci + 1,
                        value=mr[si] if si < len(mr) else None)
            c.fill = kit["master_fill"]
            c.border = kit["border"]
            if key_out_idx is not None and ci == key_out_idx:
                c.font = bold
        for ci in range(n_m, n_m + len(d_idx)):
            c = ws.cell(row=row_no, column=ci + 1)
            c.fill = kit["master_fill"]
            c.border = kit["border"]
        row_no += 1
        detail_start = row_no
        for dr in detail_by_key.get(_cell_str(_to_sql_value(mr[m_key])), []):
            for ci, si in enumerate(d_idx):
                c = ws.cell(row=row_no, column=n_m + ci + 1,
                            value=dr[si] if si < len(dr) else None)
                c.fill = kit["detail_fill"]
                c.border = kit["border"]
            ws.row_dimensions[row_no].outline_level = 1
            row_no += 1
        if sub_out_cols and row_no > detail_start:
            label = ws.cell(row=row_no, column=1, value="Zwischensumme")
            label.font = bold
            for col in sub_out_cols:
                letter = get_column_letter(col)
                c = ws.cell(row=row_no, column=col,
                            value=f"=SUM({letter}{detail_start}:"
                                  f"{letter}{row_no - 1})")
                c.font = bold
            for ci in range(len(combined)):
                cc = ws.cell(row=row_no, column=ci + 1)
                cc.fill = kit["master_fill"]
                cc.border = kit["border"]
            ws.row_dimensions[row_no].outline_level = 1
            row_no += 1
    if sheet_spec.get("freeze_header", True):
        ws.freeze_panes = "A2"
    _apply_print_setup(ws, sheet_spec)
    sample_rows = [[None] * len(combined)]
    _auto_widths(ws, combined, sample_rows)
    n_data = row_no - 2
    _apply_number_formats(ws, combined, sheet_spec.get("columns") or [],
                          2, row_no - 1)
    _render_charts(ws, sheet_spec, combined, n_data)
    _render_conditional(ws, sheet_spec, combined, n_data)
    return n_data


def _fill_template(out_path: str, spec: dict) -> dict:
    """v2 template fill: copy an existing (corporate) workbook and write DATA
    into it — values only, the template's styling/formulas/charts stay
    untouched. Analog of write_document's style='reference' for docx. Per
    sheet: target an existing sheet by name, place rows at `anchor` (default
    A1) or at a workbook `named_range`."""
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple
    tpl = spec["template"]
    if isinstance(tpl, str):
        tpl = {"file": tpl}
    src = _resolve_input_path(tpl.get("file") or "")
    if not os.path.isfile(src):
        raise FileNotFoundError(f"template file not found: {tpl.get('file')}")
    shutil.copy(src, out_path)
    wb = openpyxl.load_workbook(
        out_path, keep_vba=out_path.lower().endswith(".xlsm"))
    out_sheets = []
    for sh in spec.get("sheets") or []:
        target = sh.get("name")
        if target not in wb.sheetnames:
            raise ValueError(
                f"sheet '{target}' not in template — sheets: {wb.sheetnames}")
        ws = wb[target]
        # Inline rows in template mode are pure DATA (placed at the anchor —
        # the template already has its own headers); source keeps its header
        # for optional include_header.
        if sh.get("rows") is not None:
            header, rows = [], [list(r) for r in sh["rows"]]
        elif sh.get("source"):
            header, rows = _resolve_source(sh["source"])
        else:
            raise ValueError("template sheet needs rows or source")
        anchor = sh.get("anchor") or "A1"
        nr = sh.get("named_range")
        if nr:
            try:
                dests = list(wb.defined_names[nr].destinations)
                sheet_name, ref = dests[0]
                anchor = ref.replace("$", "").split(":")[0]
                ws = wb[sheet_name]
            except Exception:
                raise ValueError(
                    f"named_range '{nr}' not found — available: "
                    f"{list(wb.defined_names.keys())}")
        r0, c0 = coordinate_to_tuple(anchor)
        write_rows = ([header] + rows) if (sh.get("include_header")
                                           and header) else rows
        for ri, row in enumerate(write_rows):
            for ci, v in enumerate(row):
                ws.cell(row=r0 + ri, column=c0 + ci, value=v)
        out_sheets.append({"name": ws.title, "rows": len(write_rows)})
    wb.save(out_path)
    return {"sheets": out_sheets, "template": os.path.basename(src)}


def render_spec(path: str, spec: dict, style_name: str = "",
                style: dict | None = None) -> dict:
    """Render a declarative workbook spec to `path`. Shared by xlsx_create and
    the write_document .xlsx branch (which builds a spec from markdown tables
    and passes its already-loaded `style` dict). Raises ValueError/
    FileNotFoundError with model-actionable messages."""
    from engine.tools.file_tools import _resolve_default_style, _load_doc_style
    import openpyxl
    if not isinstance(spec, dict) or not isinstance(spec.get("sheets"), list) \
            or not spec["sheets"]:
        raise ValueError('spec needs {"sheets": [{...}]}')
    if style is None:
        style = _load_doc_style(_resolve_default_style(
            style_name or spec.get("style") or ""))
    kit = _style_kit(style)
    # v4: pre-resolve plain-table data so we can pick the engine. Any sheet
    # past WRITE_ONLY_ROWS switches the WHOLE workbook to the streaming
    # writer (openpyxl can't mix write_only and regular sheets).
    prepared = []  # (sheet_spec, kind, data)
    huge = False
    for sheet_spec in spec["sheets"]:
        if not isinstance(sheet_spec, dict):
            raise ValueError("each sheet must be an object")
        if sheet_spec.get("master_detail") or sheet_spec.get("pivot"):
            prepared.append((sheet_spec, "complex", None))
            continue
        header, rows = _sheet_data(sheet_spec)
        if len(rows) > WRITE_ONLY_ROWS:
            huge = True
        prepared.append((sheet_spec, "table", (header, rows)))
    if huge:
        if any(kind == "complex" for _, kind, _ in prepared):
            raise ValueError(
                f"a sheet exceeds {WRITE_ONLY_ROWS:,} rows (streaming mode), "
                f"but master_detail/pivot sheets can't stream — put the huge "
                f"table in its own file, or aggregate it first via source.sql")
        return _render_spec_write_only(
            path, [(s, d[0], d[1]) for s, _, d in prepared], kit)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set = set()
    out_sheets = []
    for sheet_spec, kind, data in prepared:
        name = _clean_sheet_name(sheet_spec.get("name"), used_names)
        ws = wb.create_sheet(name)
        if sheet_spec.get("master_detail"):
            n = _render_master_detail_sheet(ws, sheet_spec, kit)
        elif sheet_spec.get("pivot"):
            n = _render_pivot_sheet(ws, sheet_spec, kit)
        else:
            n = _render_table_sheet(ws, sheet_spec, kit, data=data)
        out_sheets.append({"name": name, "rows": n})
    wb.save(path)
    return {"sheets": out_sheets}


def tool_xlsx_create(args: dict) -> str:
    import brain as _brain
    from engine.tools.file_tools import _enforce_artifact_path
    raw_path = (args.get("path") or "").strip()
    if not raw_path:
        return _err("xlsx_create: 'path' is required (e.g. 'report.xlsx')")
    if not raw_path.lower().endswith(".xlsx"):
        raw_path += ".xlsx"
    spec = args.get("spec")
    if isinstance(spec, str):
        try:
            spec = json.loads(spec)
        except ValueError as e:
            return _err(f"xlsx_create: spec is not valid JSON: {e}")
    path, perr = _enforce_artifact_path(raw_path, "xlsx_create")
    if perr:
        return perr
    try:
        if isinstance(spec, dict) and spec.get("template"):
            info = _fill_template(path, spec)
        else:
            info = render_spec(path, spec or {})
        if args.get("recalc") or (isinstance(spec, dict) and spec.get("recalc")):
            recalc_workbook(path)
    except (ValueError, FileNotFoundError) as e:
        return _err(f"xlsx_create: {e}")
    except Exception as e:
        return _err(f"xlsx_create: {type(e).__name__}: {e}")
    agent = get_request_context().current_agent
    _brain._after_file_write(path, "created",
                             agent.agent_id if agent else "main")
    size = os.path.getsize(path)
    res = {"path": path, "size": size, "sheets": info["sheets"],
           "status": "written"}
    notes = []
    if info.get("mode"):
        res["mode"] = info["mode"]
        notes.append(info.get("note") or "")
    if size > 5 * 1024 * 1024:
        notes.append("file exceeds the 5MB artifact-version snapshot cap")
    if any(notes):
        res["note"] = "; ".join(n for n in notes if n)
    return _ok(res)


# ---------------------------------------------------------------------------
# xlsx_diff — deterministic workbook comparison (v2)
# ---------------------------------------------------------------------------

DIFF_DETAIL_CAP = 50


def _format_signature(cell) -> str:
    """Compact, human-readable per-cell format signature for
    compare='formats': number format · bold/italic/underline · font colour ·
    solid fill colour. Deterministic — two cells with the same look yield the
    same string."""
    try:
        parts = [cell.number_format or "General"]
        f = cell.font
        flags = ""
        if f is not None:
            if f.bold:
                flags += "fett"
            if f.italic:
                flags += ("+" if flags else "") + "kursiv"
            if f.underline and f.underline != "none":
                flags += ("+" if flags else "") + "unterstrichen"
            col = getattr(getattr(f, "color", None), "rgb", None)
            if isinstance(col, str) and col not in ("FF000000",):
                parts.append(f"schrift:{col[-6:]}")
        if flags:
            parts.insert(1, flags)
        fill = cell.fill
        if fill is not None and fill.fill_type == "solid":
            fc = getattr(getattr(fill, "start_color", None), "rgb", None)
            if isinstance(fc, str) and fc not in ("00000000", "FFFFFFFF"):
                parts.append(f"füllung:{fc[-6:]}")
        return " · ".join(parts)
    except Exception:
        return "?"


def _sig_matrices(path: str, sheet: str | None = None) -> dict:
    """{sheet_title: [[signature, ...], ...]} — the full per-cell format map
    of a workbook (read_only pass; styles are parsed lazily). Aligned back to
    value grids via grid.row_nums (absolute rows) + 1:1 column mapping."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    try:
        for ws in wb.worksheets:
            if sheet and ws.title != sheet:
                continue
            out[ws.title] = [[_format_signature(c) for c in row]
                             for row in ws.iter_rows()]
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _fmt_rows_for_grid(g: dict, sigs: dict) -> list | None:
    """Format rows aligned to a value grid: same row order (row_nums) and the
    same trimmed column count. None when the grid has no row mapping."""
    mat = sigs.get(g.get("sheet_title") or g["name"])
    row_nums = g.get("row_nums") or []
    if mat is None or not row_nums:
        return None
    n = len(g["header"])
    out = []
    for rn in row_nums:
        src = mat[rn - 1] if 0 <= rn - 1 < len(mat) else []
        out.append(list(src[:n]) + [""] * (n - len(src)))
    return out


def _diff_grids(ga: dict, gb: dict, keys: list | None,
                label_a: str, label_b: str,
                fmt_a: list | None = None, fmt_b: list | None = None):
    """Compare two grids. With `keys` (v3: composite keys supported): keyed
    row diff (added/removed/changed). Without: positional. Returns
    (summary_lines, detail_rows, marks) — detail_rows = [(key, column,
    value_a, value_b), ...] for the CSV export; marks = the cell/row
    coordinates on the B side for the highlighted-xlsx export:
    {changed: {(b_row_idx, b_col_idx): old_value}, added: {b_row_idx},
    removed_rows: [A-row values in B column order]}."""
    lines = []
    details = []
    marks = {"changed": {}, "added": set(), "removed_rows": []}
    cols_a = [str(h) for h in ga["header"]]
    cols_b = [str(h) for h in gb["header"]]
    la = {c.strip().lower(): i for i, c in enumerate(cols_a)}
    lb = {c.strip().lower(): i for i, c in enumerate(cols_b)}
    only_a = [c for c in cols_a if c.strip().lower() not in lb]
    only_b = [c for c in cols_b if c.strip().lower() not in la]
    if only_a:
        lines.append(f"- columns only in {label_a}: {', '.join(only_a)}")
    if only_b:
        lines.append(f"- columns only in {label_b}: {', '.join(only_b)}")
    common = [c for c in cols_a if c.strip().lower() in lb]
    lines.append(f"- rows: {label_a} {len(ga['rows'])} / "
                 f"{label_b} {len(gb['rows'])}")

    def norm(v):
        return _cell_str(_to_sql_value(v))

    def _a_row_in_b_order(ra):
        out = [None] * len(cols_b)
        for c in common:
            cl = c.strip().lower()
            if la[cl] < len(ra):
                out[lb[cl]] = ra[la[cl]]
        return out

    if keys:
        kls = [str(k).strip().lower() for k in keys]
        for kl, korig in zip(kls, keys):
            if kl not in la or kl not in lb:
                raise ValueError(
                    f"key '{korig}' must exist on both sides — {label_a}: "
                    f"{cols_a}; {label_b}: {cols_b}")
        ias = [la[kl] for kl in kls]
        ibs = [lb[kl] for kl in kls]

        def _ka(r):
            return " | ".join(norm(r[i]) if i < len(r) else "" for i in ias)

        def _kb(r):
            return " | ".join(norm(r[i]) if i < len(r) else "" for i in ibs)

        rows_a = {_ka(r): (i, r) for i, r in enumerate(ga["rows"])}
        rows_b = {_kb(r): (i, r) for i, r in enumerate(gb["rows"])}
        added = [k for k in rows_b if k not in rows_a]
        removed = [k for k in rows_a if k not in rows_b]
        changed = []
        for k, (ai, ra) in rows_a.items():
            hit = rows_b.get(k)
            if hit is None:
                continue
            bi, rb = hit
            diff_cols = []
            for c in common:
                cl = c.strip().lower()
                # compare='formats': the compared payload per cell is the
                # FORMAT signature; row matching stays on the value key.
                if fmt_a is not None and fmt_b is not None:
                    va = fmt_a[ai][la[cl]] if la[cl] < len(fmt_a[ai]) else ""
                    vb = fmt_b[bi][lb[cl]] if lb[cl] < len(fmt_b[bi]) else ""
                else:
                    va = ra[la[cl]] if la[cl] < len(ra) else None
                    vb = rb[lb[cl]] if lb[cl] < len(rb) else None
                if norm(va) != norm(vb):
                    diff_cols.append((c, va, vb))
                    details.append((k, c, va, vb))
                    marks["changed"][(bi, lb[cl])] = va
            if diff_cols:
                changed.append((k, diff_cols))
        for k in added:
            details.append((k, "(row added)", "", "present"))
            marks["added"].add(rows_b[k][0])
        for k in removed:
            details.append((k, "(row removed)", "present", ""))
            marks["removed_rows"].append(_a_row_in_b_order(rows_a[k][1]))
        key_label = " + ".join(str(k) for k in keys)
        lines.append(f"- keyed on `{key_label}`: {len(added)} added, "
                     f"{len(removed)} removed, {len(changed)} changed row(s)")
        if len(rows_a) < len(ga["rows"]) or len(rows_b) < len(gb["rows"]):
            lines.append("- note: duplicate key values collapsed "
                         "(last occurrence wins) — key is not unique")
        shown = 0
        for k, diff_cols in changed:
            if shown >= DIFF_DETAIL_CAP:
                lines.append(f"- … {len(changed) - shown} more changed rows "
                             f"(pass out='diff.csv' for the full list)")
                break
            cells = "; ".join(f"{c}: {norm(va)!r} → {norm(vb)!r}"
                              for c, va, vb in diff_cols[:6])
            lines.append(f"  - `{k}`: {cells}")
            shown += 1
        if added[:10]:
            lines.append("  - added keys: "
                         + ", ".join(added[:10])
                         + (" …" if len(added) > 10 else ""))
        if removed[:10]:
            lines.append("  - removed keys: "
                         + ", ".join(removed[:10])
                         + (" …" if len(removed) > 10 else ""))
    else:
        n = max(len(ga["rows"]), len(gb["rows"]))
        shown = 0
        for i in range(n):
            ra = ga["rows"][i] if i < len(ga["rows"]) else None
            rb = gb["rows"][i] if i < len(gb["rows"]) else None
            if ra is None:
                details.append((f"row {i + 2}", "(row added)", "", ""))
                marks["added"].add(i)
                continue
            if rb is None:
                details.append((f"row {i + 2}", "(row removed)", "", ""))
                marks["removed_rows"].append(_a_row_in_b_order(ra))
                continue
            for c in common:
                cl = c.strip().lower()
                if fmt_a is not None and fmt_b is not None:
                    va = fmt_a[i][la[cl]] if la[cl] < len(fmt_a[i]) else ""
                    vb = fmt_b[i][lb[cl]] if lb[cl] < len(fmt_b[i]) else ""
                else:
                    va = ra[la[cl]] if la[cl] < len(ra) else None
                    vb = rb[lb[cl]] if lb[cl] < len(rb) else None
                if norm(va) != norm(vb):
                    details.append((f"row {i + 2}", c, va, vb))
                    marks["changed"][(i, lb[cl])] = va
                    if shown < DIFF_DETAIL_CAP:
                        lines.append(f"  - row {i + 2} `{c}`: "
                                     f"{norm(va)!r} → {norm(vb)!r}")
                        shown += 1
        if len(details) > shown:
            lines.append(f"- … {len(details) - shown} more cell differences "
                         f"(pass out='diff.csv' for the full list)")
        lines.append(f"- positional compare (no key given): "
                     f"{len(details)} differing cell(s)/row(s) — pass "
                     f"key='<column>' for a keyed compare")
    return lines, details, marks


def _write_diff_xlsx(out_path: str, per_sheet: list, kit):
    """v3: the diff as a HIGHLIGHTED workbook — per compared sheet the B side
    rendered with changed cells yellow (+ comment 'vorher: <alt>'), added rows
    green, and the removed A rows appended red under a marker row."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    yellow = PatternFill("solid", start_color="FFF2AB", end_color="FFF2AB")
    green = PatternFill("solid", start_color="C9EFC9", end_color="C9EFC9")
    red = PatternFill("solid", start_color="F5C1C1", end_color="F5C1C1")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set = set()
    for sheet_name, gb, marks in per_sheet:
        ws = wb.create_sheet(_clean_sheet_name(sheet_name, used))
        _apply_header_row(ws, gb["header"], kit)
        for ri, row in enumerate(gb["rows"]):
            for ci in range(len(gb["header"])):
                v = row[ci] if ci < len(row) else None
                c = ws.cell(row=ri + 2, column=ci + 1, value=v)
                c.border = kit["border"]
                if (ri, ci) in marks["changed"]:
                    c.fill = yellow
                    old = _cell_str(_to_sql_value(marks["changed"][(ri, ci)]))
                    c.comment = Comment(f"vorher: {old}", "xlsx_diff")
                elif ri in marks["added"]:
                    c.fill = green
        row_no = len(gb["rows"]) + 2
        if marks["removed_rows"]:
            mark = ws.cell(row=row_no, column=1,
                           value="— Entfernte Zeilen (nur in der alten Datei) —")
            mark.font = Font(bold=True, italic=True)
            row_no += 1
            for rr in marks["removed_rows"]:
                for ci in range(len(gb["header"])):
                    v = rr[ci] if ci < len(rr) else None
                    c = ws.cell(row=row_no, column=ci + 1, value=v)
                    c.fill = red
                    c.border = kit["border"]
                row_no += 1
        ws.freeze_panes = "A2"
        _auto_widths(ws, gb["header"], gb["rows"])
    wb.save(out_path)


def tool_xlsx_diff(args: dict) -> str:
    import brain as _brain
    pa, pb = args.get("path_a") or "", args.get("path_b") or ""
    if not pa or not pb:
        return _err("xlsx_diff: 'path_a' and 'path_b' are required")
    raw_key = args.get("key")
    if isinstance(raw_key, (list, tuple)):
        keys = [str(k).strip() for k in raw_key if str(k).strip()] or None
    else:
        keys = [k.strip() for k in str(raw_key or "").split(",")
                if k.strip()] or None
    sheet = args.get("sheet")
    # compare='formulas' (v3): diff the formula STRINGS instead of values.
    # compare='formats' (v4): diff the per-cell FORMAT signatures (number
    # format, bold/italic/underline, font/fill colour) — rows still matched
    # by their VALUE key, so a re-coloured cell surfaces even when values
    # are identical.
    mode = (args.get("compare") or "").strip().lower()
    formulas = mode == "formulas"
    formats = mode == "formats"
    try:
        def _load_side(p):
            if _is_handle(p):
                if formulas or formats:
                    raise ValueError(
                        f"compare='{mode}' works on files, not stored results")
                return [_grid_from_handle(p)]
            return _load_grids(_resolve_input_path(p), sheet=sheet,
                               formulas=formulas)

        grids_a = _load_side(pa)
        grids_b = _load_side(pb)
        fmt_sigs_a = fmt_sigs_b = None
        if formats:
            fmt_sigs_a = _sig_matrices(_resolve_input_path(pa), sheet=sheet)
            fmt_sigs_b = _sig_matrices(_resolve_input_path(pb), sheet=sheet)
        la, lb = os.path.basename(pa), os.path.basename(pb)
        by_name_a = {g["name"]: g for g in grids_a}
        by_name_b = {g["name"]: g for g in grids_b}
        parts = [f"# Diff: {la} ↔ {lb}"
                 + (" (Formel-Vergleich)" if formulas else "")
                 + (" (Formatierungs-Vergleich)" if formats else "")]
        all_details = []
        per_sheet = []
        only_a = [n for n in by_name_a if n not in by_name_b]
        only_b = [n for n in by_name_b if n not in by_name_a]
        if only_a:
            parts.append(f"- sheets only in {la}: {', '.join(only_a)}")
        if only_b:
            parts.append(f"- sheets only in {lb}: {', '.join(only_b)}")
        common = [n for n in by_name_a if n in by_name_b]
        def _fmt_pair(gpa, gpb):
            if not formats:
                return None, None
            return (_fmt_rows_for_grid(gpa, fmt_sigs_a),
                    _fmt_rows_for_grid(gpb, fmt_sigs_b))

        if not common and len(grids_a) == 1 and len(grids_b) == 1:
            # single-sheet workbooks with different sheet names still compare
            parts.append(f"\n## {grids_a[0]['name']} ↔ {grids_b[0]['name']}")
            fa, fb = _fmt_pair(grids_a[0], grids_b[0])
            lines, details, marks = _diff_grids(grids_a[0], grids_b[0],
                                                keys, la, lb,
                                                fmt_a=fa, fmt_b=fb)
            parts.extend(lines)
            all_details.extend(details)
            per_sheet.append((grids_b[0]["name"], grids_b[0], marks))
        else:
            for n in common:
                parts.append(f"\n## Sheet: {n}")
                fa, fb = _fmt_pair(by_name_a[n], by_name_b[n])
                lines, details, marks = _diff_grids(by_name_a[n], by_name_b[n],
                                                    keys, la, lb,
                                                    fmt_a=fa, fmt_b=fb)
                parts.extend(lines)
                all_details.extend(details)
                per_sheet.append((n, by_name_b[n], marks))
        out_info = None
        out_name = (args.get("out") or "").strip()
        if out_name and all_details:
            from engine.tools.file_tools import (_enforce_artifact_path,
                                                 _resolve_default_style,
                                                 _load_doc_style)
            if not out_name.lower().endswith((".csv", ".xlsx")):
                out_name += ".xlsx"
            out_path, perr = _enforce_artifact_path(out_name, "xlsx_diff")
            if perr:
                return perr
            if out_path.lower().endswith(".xlsx"):
                kit = _style_kit(_load_doc_style(_resolve_default_style("")))
                _write_diff_xlsx(out_path, per_sheet, kit)
            else:
                with open(out_path, "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(["key", "column", "value_a", "value_b"])
                    for k, c, va, vb in all_details:
                        w.writerow([k, c, _cell_str(_to_sql_value(va)),
                                    _cell_str(_to_sql_value(vb))])
            agent = get_request_context().current_agent
            _brain._after_file_write(
                out_path, "created", agent.agent_id if agent else "main")
            out_info = {"path": out_path, "rows": len(all_details)}
        report = "\n".join(parts)
        report = _brain._gdpr_anon_tool_text(report, f"xlsx_diff:{la},{lb}")
        res = {"differences": len(all_details), "report": report}
        if out_info:
            res["saved"] = out_info
        return _ok(res)
    except (ValueError, FileNotFoundError) as e:
        return _err(f"xlsx_diff: {e}")
    except Exception as e:
        return _err(f"xlsx_diff: {type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# xlsx_edit — declarative ops on an existing workbook, format-preserving
# ---------------------------------------------------------------------------

def _copy_row_styles(ws, src_row: int, dst_row: int, n_cols: int):
    from copy import copy
    for ci in range(1, n_cols + 1):
        s, d = ws.cell(row=src_row, column=ci), ws.cell(row=dst_row, column=ci)
        if s.has_style:
            d._style = copy(s._style)


def _edit_header_map(ws) -> dict:
    """Column name → index from row 1 (xlsx_edit assumes header in row 1 —
    stated in the tool description)."""
    return {str(c.value).strip().lower(): c.column
            for c in ws[1] if c.value is not None}


def _op_rows(op: dict) -> list[list]:
    if op.get("rows") is not None:
        return [list(r) for r in op["rows"]]
    if op.get("source"):
        _, rows = _resolve_source(op["source"])
        return rows
    raise ValueError(f"{op.get('op')} needs rows or source")


def _apply_edit_ops(wb, ops: list, kit) -> list[dict]:
    applied = []
    used_names = {ws.title for ws in wb.worksheets}
    for op in ops:
        if not isinstance(op, dict) or not op.get("op"):
            raise ValueError('each op needs {"op": "..."}')
        kind = op["op"]
        if kind == "add_sheet":
            name = _clean_sheet_name(op.get("name"), used_names)
            ws = wb.create_sheet(name)
            if op.get("master_detail"):
                n = _render_master_detail_sheet(ws, op, kit)
            elif op.get("pivot"):
                n = _render_pivot_sheet(ws, op, kit)
            else:
                n = _render_table_sheet(ws, op, kit)
            applied.append({"op": kind, "sheet": name, "rows_affected": n})
            continue
        if kind == "rename_sheet":
            src = op.get("from") or op.get("sheet")
            if src not in wb.sheetnames:
                raise ValueError(
                    f"sheet '{src}' not found — sheets: {wb.sheetnames}")
            ws = wb[src]
            ws.title = str(op.get("to") or "")[:31]
            used_names.discard(src)
            used_names.add(ws.title)
            applied.append({"op": kind, "sheet": ws.title, "rows_affected": 0})
            continue
        if kind == "delete_sheet":
            name = op.get("name") or op.get("sheet")
            if name not in wb.sheetnames:
                raise ValueError(
                    f"sheet '{name}' not found — sheets: {wb.sheetnames}")
            if len(wb.sheetnames) == 1:
                raise ValueError("cannot delete the only sheet")
            wb.remove(wb[name])
            applied.append({"op": kind, "sheet": name, "rows_affected": 0})
            continue

        sheet_name = op.get("sheet")
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"sheet '{sheet_name}' not found — sheets: {wb.sheetnames}")
        ws = wb[sheet_name]
        hdr = _edit_header_map(ws)

        if kind == "append_rows":
            rows = _op_rows(op)
            template_row = ws.max_row if ws.max_row >= 2 else None
            n_cols = ws.max_column
            start = ws.max_row + 1
            for ri, row in enumerate(rows):
                for ci, v in enumerate(row[:n_cols], 1):
                    ws.cell(row=start + ri, column=ci, value=v)
                if template_row:
                    _copy_row_styles(ws, template_row, start + ri, n_cols)
            applied.append({"op": kind, "sheet": sheet_name,
                            "rows_affected": len(rows)})
        elif kind == "add_column":
            name = op.get("name")
            if not name:
                raise ValueError("add_column needs name")
            ci = ws.max_column + 1
            from copy import copy
            hcell = ws.cell(row=1, column=ci, value=name)
            neighbor = ws.cell(row=1, column=ci - 1)
            if neighbor.has_style:
                hcell._style = copy(neighbor._style)
            values = op.get("values")
            formula = op.get("formula")
            n = 0
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                if formula:
                    cell.value = formula.replace("{row}", str(r))
                elif values is not None:
                    if r - 2 < len(values):
                        cell.value = values[r - 2]
                body_neighbor = ws.cell(row=r, column=ci - 1)
                if body_neighbor.has_style:
                    cell._style = copy(body_neighbor._style)
                fmt = _NUMBER_FORMATS.get(op.get("format") or "")
                if fmt:
                    cell.number_format = fmt
                n += 1
            applied.append({"op": kind, "sheet": sheet_name,
                            "rows_affected": n})
        elif kind == "update_cells":
            where = op.get("where") or {}
            setter = op.get("set") or {}
            wcol = hdr.get(str(where.get("column", "")).strip().lower())
            if wcol is None:
                raise ValueError(
                    f"where.column '{where.get('column')}' not found — "
                    f"columns: {sorted(hdr)}")
            target_cols = {}
            for k, v in setter.items():
                tc = hdr.get(str(k).strip().lower())
                if tc is None:
                    raise ValueError(
                        f"set column '{k}' not found — columns: {sorted(hdr)}")
                target_cols[tc] = v
            if not target_cols:
                raise ValueError("update_cells needs set: {column: value}")
            cond_op = next((k for k in ("equals", "contains", "lt", "gt")
                            if k in where), None)
            if cond_op is None:
                raise ValueError(
                    "where needs one of equals|contains|lt|gt")
            ref = where[cond_op]
            n = 0
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=wcol).value
                hit = False
                try:
                    if cond_op == "equals":
                        hit = (v == ref or _cell_str(v) == _cell_str(ref))
                    elif cond_op == "contains":
                        hit = str(ref).lower() in _cell_str(v).lower()
                    elif cond_op == "lt":
                        hit = v is not None and v < ref
                    elif cond_op == "gt":
                        hit = v is not None and v > ref
                except TypeError:
                    hit = False
                if hit:
                    for tc, val in target_cols.items():
                        ws.cell(row=r, column=tc, value=val)
                    n += 1
            applied.append({"op": kind, "sheet": sheet_name,
                            "rows_affected": n})
        elif kind == "set_format":
            fmt = _NUMBER_FORMATS.get(op.get("format") or "")
            if not fmt:
                raise ValueError(
                    f"format must be one of {sorted(_NUMBER_FORMATS)}")
            n = 0
            for cname in op.get("columns") or []:
                ci = hdr.get(str(cname).strip().lower())
                if ci is None:
                    raise ValueError(
                        f"column '{cname}' not found — columns: {sorted(hdr)}")
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=ci).number_format = fmt
                    n += 1
            applied.append({"op": kind, "sheet": sheet_name,
                            "rows_affected": n})
        else:
            raise ValueError(
                f"unknown op '{kind}' — use append_rows|add_column|"
                f"update_cells|add_sheet|rename_sheet|delete_sheet|set_format")
    return applied


def tool_xlsx_edit(args: dict) -> str:
    import brain as _brain
    import openpyxl
    from engine.tools.file_tools import (_enforce_artifact_path,
                                         _resolve_default_style,
                                         _load_doc_style)
    raw_path = (args.get("path") or "").strip()
    if not raw_path:
        return _err("xlsx_edit: 'path' is required")
    spec = args.get("spec")
    if isinstance(spec, str):
        try:
            spec = json.loads(spec)
        except ValueError as e:
            return _err(f"xlsx_edit: spec is not valid JSON: {e}")
    ops = (spec or {}).get("ops") if isinstance(spec, dict) else None
    if not ops:
        return _err('xlsx_edit: spec needs {"ops": [{"op": "...", ...}]}')
    # Edits write the file back — same artifact-folder rule as every write
    # tool. Attachments under /tmp/brain-attachments are read-only inputs; to
    # modify one, the model should xlsx_create a new file from it instead.
    path, perr = _enforce_artifact_path(raw_path, "xlsx_edit")
    if perr:
        return perr
    if not os.path.isfile(path):
        return _err(f"xlsx_edit: File not found: {raw_path} — xlsx_edit "
                    f"changes an EXISTING workbook in your artifact folder; "
                    f"use xlsx_create for new files.")
    try:
        # data_only=False preserves formulas on save (openpyxl drops cached
        # values; Excel recalculates on open — the standard trade).
        wb = openpyxl.load_workbook(
            path, keep_vba=path.lower().endswith(".xlsm"))
        kit = _style_kit(_load_doc_style(_resolve_default_style("")))
        applied = _apply_edit_ops(wb, ops, kit)
        wb.save(path)
        if args.get("recalc") or (spec or {}).get("recalc"):
            # LibreOffice round-trip: computed formula values land in the
            # file, so a follow-up xlsx_query sees numbers, not NULLs.
            recalc_workbook(path)
    except (ValueError, FileNotFoundError) as e:
        return _err(f"xlsx_edit: {e}")
    except Exception as e:
        return _err(f"xlsx_edit: {type(e).__name__}: {e}")
    agent = get_request_context().current_agent
    _brain._after_file_write(path, "modified",
                             agent.agent_id if agent else "main")
    return _ok({"path": path, "ops_applied": applied, "status": "edited"})
