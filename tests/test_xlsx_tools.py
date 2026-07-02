"""Tests for the deterministic XLSX toolset (xlsx_inspect / xlsx_query /
xlsx_create / xlsx_edit, engine/tools/xlsx_tools.py) + a byte-stability pin
on doc_convert._extract_xlsx.

WHY the pin comes first: xlsx_tools reuses the v9.261.0 placeholder-column
trim, which gets factored OUT of _extract_xlsx into
doc_convert._trim_placeholder_columns. The mining daemon's companion-`.md`
output must stay byte-stable across that refactor (changing it would re-embed
every project doc) — so the expected strings below were captured by PROBING
the PRE-refactor _extract_xlsx (2026-07-02) and the refactor must reproduce
them exactly.

The synthetic fixture mirrors the production case that motivated the toolset
(chats 2cb94154 / 98cceac2): a 2-sheet workbook, market orders in sheet 1,
partial executions in sheet 2, joined via MARKTORDERNUMMER.

Run: python3 -m unittest tests.test_xlsx_tools -v
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
import uuid  # noqa: F401  (used by the tool fixtures below)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from engine import doc_convert  # noqa: E402


def _build_orders_workbook(path):
    """The marktorder shape: orders + partial executions keyed by
    MARKTORDERNUMMER (MO-1 has 2 executions, MO-2 has 1)."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["MARKTORDERNUMMER", "NAME", "STUECK", "LIMIT"])
    ws1.append(["MO-1", "Alice", 100, 25.5])
    ws1.append(["MO-2", "Bob", 50, None])
    ws2 = wb.create_sheet("Ausfuehrungen")
    ws2.append(["MARKTORDERNUMMER", "DATUM", "STUECK", "KURS"])
    ws2.append(["MO-1", "2026-01-02", 60, 25.4])
    ws2.append(["MO-1", "2026-01-03", 40, 25.6])
    ws2.append(["MO-2", "2026-01-04", 50, 11.0])
    wb.save(path)


def _build_placeholder_workbook(path):
    """The v9.261.0 failure shape: real columns A/B/C plus an auto-named
    'Spalte<N>' placeholder tail that carries headers but no data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["A", "B", "C"] + [f"Spalte{i}" for i in range(4, 40)])
    ws.append([1, 2, 3] + [None] * 36)
    ws.append([4, None, 6] + [None] * 36)
    wb.save(path)


class _XlsxFixture(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.mkdtemp(prefix="xlsxtools_test_")
        self.orders_path = os.path.join(self._tmp, "orders.xlsx")
        self.placeholder_path = os.path.join(self._tmp, "placeholder.xlsx")
        _build_orders_workbook(self.orders_path)
        _build_placeholder_workbook(self.placeholder_path)


class TestExtractXlsxBytePin(_XlsxFixture):
    """Byte-stability pin for _extract_xlsx across the trim refactor.
    Expected values probed against the PRE-refactor code — do not 'fix' them
    to match new behavior; if they fail, the refactor changed mining output."""

    ORDERS_EXPECTED = (
        "# orders\n\n\n"
        "## Sheet: Orders\n\n"
        "| MARKTORDERNUMMER | NAME | STUECK | LIMIT |\n"
        "|---|---|---|---|\n"
        "| MO-1 | Alice | 100 | 25.5 |\n"
        "| MO-2 | Bob | 50 |  |\n\n"
        "## Sheet: Ausfuehrungen\n\n"
        "| MARKTORDERNUMMER | DATUM | STUECK | KURS |\n"
        "|---|---|---|---|\n"
        "| MO-1 | 2026-01-02 | 60 | 25.4 |\n"
        "| MO-1 | 2026-01-03 | 40 | 25.6 |\n"
        "| MO-2 | 2026-01-04 | 50 | 11 |\n"
    )

    PLACEHOLDER_EXPECTED = (
        "# placeholder\n\n\n"
        "## Sheet: Daten\n\n"
        "| A | B | C |\n"
        "|---|---|---|\n"
        "| 1 | 2 | 3 |\n"
        "| 4 |  | 6 |\n"
    )

    def test_orders_caps_false_and_true_identical_to_pin(self):
        for caps in (False, True):
            text, err = doc_convert._extract_xlsx(self.orders_path, caps=caps)
            self.assertIsNone(err)
            self.assertEqual(text, self.ORDERS_EXPECTED)

    def test_placeholder_tail_trimmed_identical_to_pin(self):
        for caps in (False, True):
            text, err = doc_convert._extract_xlsx(
                self.placeholder_path, caps=caps)
            self.assertIsNone(err)
            self.assertEqual(text, self.PLACEHOLDER_EXPECTED)


import brain  # noqa: E402  (loads TOOL_DISPATCH etc.; also warms lazy imports)
from engine.context import request_context  # noqa: E402
from engine.tools import xlsx_tools  # noqa: E402


class _FakeAgent:
    agent_id = "main"


class _ToolFixture(_XlsxFixture):
    """Adds a request context (unique session id → fresh artifact folder, same
    pattern as test_file_tools_characterization) on top of the workbook tmpdir."""

    def setUp(self):
        super().setUp()
        self._prev_cwd = os.getcwd()
        os.chdir(self._tmp)
        self._sid = "xlsxtools-test-" + uuid.uuid4().hex[:8]
        self.enterContext(request_context(
            current_session_id=self._sid, current_agent=_FakeAgent()))

    def tearDown(self):
        os.chdir(self._prev_cwd)
        # Don't leave per-test artifact folders behind in the repo (the sid is
        # unique per test, so this only ever removes what THIS test created).
        try:
            import shutil
            from engine.tools.file_tools import _resolve_artifact_dir
            with request_context(current_session_id=self._sid,
                                 current_agent=_FakeAgent()):
                art_dir, _ = _resolve_artifact_dir()
            if art_dir and self._sid in art_dir and os.path.isdir(art_dir):
                shutil.rmtree(art_dir)
        except Exception:
            pass


class TestXlsxInspect(_ToolFixture):
    def test_profile_and_join_key(self):
        out = json.loads(xlsx_tools.tool_xlsx_inspect(
            {"path": self.orders_path}))
        rep = out["report"]
        self.assertIn("Sheet: Orders", rep)
        self.assertIn("Sheet: Ausfuehrungen", rep)
        self.assertIn("2 data rows", rep)          # Orders
        self.assertIn("3 data rows", rep)          # Ausfuehrungen
        # join-key candidate detected with full overlap
        self.assertIn("MARKTORDERNUMMER", rep)
        self.assertIn("likely JOIN key", rep)
        # copy-paste schema block with sanitized names
        self.assertIn("Tables for xlsx_query:", rep)
        self.assertIn("orders(marktordernummer", rep)

    def test_placeholder_columns_absent(self):
        out = json.loads(xlsx_tools.tool_xlsx_inspect(
            {"path": self.placeholder_path}))
        rep = out["report"]
        self.assertIn("3 columns", rep)
        self.assertNotIn("Spalte4", rep)

    def test_missing_file_is_error(self):
        out = json.loads(xlsx_tools.tool_xlsx_inspect({"path": "nope.xlsx"}))
        self.assertIn("File not found", out["error"])


class TestXlsxQuery(_ToolFixture):
    def test_join_across_sheets(self):
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": self.orders_path,
            "sql": ("SELECT o.marktordernummer, o.name, a.datum, a.kurs "
                    "FROM orders o JOIN ausfuehrungen a "
                    "ON a.marktordernummer = o.marktordernummer "
                    "ORDER BY o.marktordernummer, a.datum")}))
        self.assertEqual(out["row_count"], 3)
        self.assertIn("| MO-1 | Alice | 2026-01-02 | 25.4 |", out["result"])
        self.assertIn("| MO-2 | Bob | 2026-01-04 |", out["result"])

    def test_aggregate(self):
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": self.orders_path,
            "sql": ("SELECT marktordernummer, SUM(stueck) AS s FROM "
                    "ausfuehrungen GROUP BY marktordernummer "
                    "ORDER BY marktordernummer")}))
        self.assertEqual(out["row_count"], 2)
        self.assertIn("| MO-1 | 100 |", out["result"])

    def test_select_only_rejections(self):
        for bad in ("INSERT INTO orders VALUES (1)",
                    "UPDATE orders SET name='x'",
                    "DROP TABLE orders",
                    "PRAGMA table_info(orders)",
                    "ATTACH DATABASE 'x' AS y",
                    "SELECT 1; DROP TABLE orders"):
            out = json.loads(xlsx_tools.tool_xlsx_query(
                {"path": self.orders_path, "sql": bad}))
            self.assertIn("error", out, bad)

    def test_cte_allowed(self):
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": self.orders_path,
            "sql": "WITH x AS (SELECT * FROM orders) SELECT COUNT(*) FROM x"}))
        self.assertEqual(out["row_count"], 1)

    def test_sql_error_echoes_schema(self):
        out = json.loads(xlsx_tools.tool_xlsx_query(
            {"path": self.orders_path, "sql": "SELECT nope FROM orders"}))
        self.assertIn("error", out)
        self.assertIn("Tables for xlsx_query:", out["error"])
        self.assertIn("orders(", out["error"])

    def test_out_writes_full_csv(self):
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": self.orders_path,
            "sql": "SELECT * FROM ausfuehrungen",
            "out": "result.csv"}))
        saved = out["saved"]
        self.assertEqual(saved["rows"], 3)
        with open(saved["path"], encoding="utf-8") as f:
            lines = f.read().strip().splitlines()
        self.assertEqual(len(lines), 4)  # header + 3 rows

    def test_multi_file_join_with_prefixes(self):
        second = os.path.join(self._tmp, "orders_neu.xlsx")
        _build_orders_workbook(second)
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "paths": [self.orders_path, second],
            "sql": ("SELECT COUNT(*) FROM orders_orders a "
                    "JOIN orders_neu_orders b "
                    "ON a.marktordernummer = b.marktordernummer")}))
        self.assertEqual(out["row_count"], 1)
        self.assertIn("| 2 |", out["result"])


class TestXlsxCreate(_ToolFixture):
    def test_inline_table_roundtrip(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "report.xlsx",
            "spec": {"sheets": [{
                "name": "Umsatz",
                "columns": [{"name": "Monat"},
                            {"name": "Betrag", "format": "eur"}],
                "rows": [["Jan", 100.5], ["Feb", 200.25]],
                "totals": ["Betrag"],
            }]}}))
        self.assertEqual(out["status"], "written")
        wb = openpyxl.load_workbook(out["path"])
        ws = wb["Umsatz"]
        self.assertEqual(ws["A1"].value, "Monat")
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws["A1"].fill.fill_type, "solid")
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIn("€", ws["B2"].number_format)
        self.assertEqual(ws["B4"].value, "=SUM(B2:B3)")
        self.assertGreaterEqual(
            ws.column_dimensions["A"].width, 8)

    def test_source_moves_data_server_side(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "kombiniert.xlsx",
            "spec": {"sheets": [{
                "name": "Alle",
                "source": {"file": self.orders_path,
                           "sql": ("SELECT o.marktordernummer, o.name, a.kurs "
                                   "FROM orders o JOIN ausfuehrungen a ON "
                                   "a.marktordernummer = o.marktordernummer")},
            }]}}))
        self.assertEqual(out["sheets"][0]["rows"], 3)
        wb = openpyxl.load_workbook(out["path"])
        ws = wb["Alle"]
        self.assertEqual(ws["A1"].value, "marktordernummer")
        self.assertEqual(ws.max_row, 4)  # header + 3

    def test_master_detail_layout(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "md.xlsx",
            "spec": {"sheets": [{
                "name": "MD",
                "master_detail": {
                    "key": "MARKTORDERNUMMER",
                    "master": {"source": {"file": self.orders_path,
                                          "sheet": "Orders"}},
                    "detail": {"source": {"file": self.orders_path,
                                          "sheet": "Ausfuehrungen"}},
                },
            }]}}))
        wb = openpyxl.load_workbook(out["path"])
        ws = wb["MD"]
        # combined header: 4 master cols + 3 detail cols (key deduped)
        self.assertEqual(ws.max_column, 7)
        # row 2 = MO-1 master (tinted, bold key), rows 3-4 its details
        self.assertEqual(ws["A2"].value, "MO-1")
        self.assertTrue(ws["A2"].font.bold)
        self.assertEqual(ws["A3"].value, None)          # detail row: master empty
        self.assertEqual(ws["E3"].value, "2026-01-02")  # first detail col
        self.assertEqual(ws.row_dimensions[3].outline_level, 1)
        # MO-2 master follows its details
        self.assertEqual(ws["A5"].value, "MO-2")
        self.assertEqual(ws["E6"].value, "2026-01-04")

    def test_inline_cap_steers_to_source(self):
        big = [["x"] * 10] * 600  # 6000 cells > cap
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "big.xlsx",
            "spec": {"sheets": [{"name": "S", "rows": big}]}}))
        self.assertIn("error", out)
        self.assertIn("source", out["error"])

    def test_charts_and_conditional(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "chart.xlsx",
            "spec": {"sheets": [{
                "name": "KPI",
                "columns": [{"name": "Monat"}, {"name": "Wert"}],
                "rows": [["Jan", 5], ["Feb", -3], ["Mrz", 8]],
                "charts": [{"type": "bar", "labels": "Monat",
                            "series": ["Wert"], "title": "Verlauf"}],
                "conditional": [{"columns": ["Wert"],
                                 "rule": {"lt": 0, "fill": "red"}}],
            }]}}))
        wb = openpyxl.load_workbook(out["path"])
        ws = wb["KPI"]
        self.assertEqual(len(ws._charts), 1)
        self.assertEqual(ws._charts[0].title.tx.rich.p[0].r[0].t, "Verlauf")
        ranges = list(ws.conditional_formatting)
        self.assertEqual(len(ranges), 1)
        self.assertIn("B2:B4", str(ranges[0].sqref))

    def test_create_inspect_query_roundtrip(self):
        created = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "loop.xlsx",
            "spec": {"sheets": [{
                "name": "Daten",
                "columns": [{"name": "K"}, {"name": "V", "format": "int"}],
                "rows": [["a", 1], ["b", 2]],
            }]}}))
        rep = json.loads(xlsx_tools.tool_xlsx_inspect(
            {"path": created["path"]}))["report"]
        self.assertIn("Sheet: Daten", rep)
        q = json.loads(xlsx_tools.tool_xlsx_query({
            "path": created["path"],
            "sql": "SELECT SUM(v) FROM daten"}))
        self.assertIn("| 3 |", q["result"])


class TestWriteDocumentXlsxRouting(_ToolFixture):
    """write_document .xlsx keeps its markdown contract (## sections → sheets,
    pipe tables, numeric coercion) but now renders through xlsx_tools —
    headers styled, freeze panes, no data change."""

    def test_two_sheet_markdown(self):
        content = (
            "## Umsatz\n\n"
            "| Monat | Betrag |\n|---|---|\n| Jan | 10.5 |\n| Feb | 20 |\n\n"
            "## Notizen\n\nKein Tabelleninhalt hier.\n")
        out = json.loads(brain.tool_write_document(
            {"path": "wd.xlsx", "content": content}))
        wb = openpyxl.load_workbook(out["path"])
        self.assertEqual(wb.sheetnames, ["Umsatz", "Notizen"])
        ws = wb["Umsatz"]
        self.assertEqual(ws["A1"].value, "Monat")
        self.assertEqual(ws["B2"].value, 10.5)   # float coercion intact
        self.assertEqual(ws["B3"].value, 20)     # int coercion intact
        self.assertTrue(ws["A1"].font.bold)      # NEW: styled header
        self.assertEqual(ws.freeze_panes, "A2")

    def test_single_table_no_sections(self):
        out = json.loads(brain.tool_write_document(
            {"path": "one.xlsx",
             "content": "| A | B |\n|---|---|\n| 1 | 2 |\n"}))
        wb = openpyxl.load_workbook(out["path"])
        self.assertEqual(wb.sheetnames, ["Sheet1"])
        self.assertEqual(wb["Sheet1"]["A2"].value, 1)


class TestXlsxEdit(_ToolFixture):
    def _create_base(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "base.xlsx",
            "spec": {"sheets": [{
                "name": "T",
                "columns": [{"name": "Name"}, {"name": "Stueck"},
                            {"name": "Kurs"}],
                "rows": [["Alice", 10, 2.5], ["Bob", 20, 3.0]],
            }]}}))
        return out["path"]

    def test_append_rows_inherits_style(self):
        path = self._create_base()
        out = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [{"op": "append_rows", "sheet": "T",
                              "rows": [["Carol", 30, 4.0]]}]}}))
        self.assertEqual(out["ops_applied"][0]["rows_affected"], 1)
        wb = openpyxl.load_workbook(path)
        ws = wb["T"]
        self.assertEqual(ws["A4"].value, "Carol")
        # style inherited from the last data row (border applied by create)
        self.assertEqual(ws["A4"].border.left.style, "thin")
        # untouched cells keep their formatting
        self.assertTrue(ws["A1"].font.bold)

    def test_add_column_formula_fills_rows(self):
        path = self._create_base()
        out = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [{"op": "add_column", "sheet": "T",
                              "name": "Wert", "formula": "=B{row}*C{row}",
                              "format": "number"}]}}))
        self.assertEqual(out["ops_applied"][0]["rows_affected"], 2)
        wb = openpyxl.load_workbook(path)
        ws = wb["T"]
        self.assertEqual(ws["D1"].value, "Wert")
        self.assertEqual(ws["D2"].value, "=B2*C2")
        self.assertEqual(ws["D3"].value, "=B3*C3")

    def test_update_cells_where(self):
        path = self._create_base()
        out = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [{"op": "update_cells", "sheet": "T",
                              "where": {"column": "Name", "equals": "Bob"},
                              "set": {"Stueck": 99}}]}}))
        self.assertEqual(out["ops_applied"][0]["rows_affected"], 1)
        wb = openpyxl.load_workbook(path)
        ws = wb["T"]
        self.assertEqual(ws["B3"].value, 99)
        self.assertEqual(ws["B2"].value, 10)  # non-matching row untouched

    def test_sheet_ops(self):
        path = self._create_base()
        out = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [
                {"op": "add_sheet", "name": "Neu",
                 "rows": [["A", "B"], [1, 2]]},
                {"op": "rename_sheet", "from": "T", "to": "Alt"},
            ]}}))
        self.assertEqual(len(out["ops_applied"]), 2)
        wb = openpyxl.load_workbook(path)
        self.assertEqual(set(wb.sheetnames), {"Alt", "Neu"})
        out2 = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [{"op": "delete_sheet", "name": "Neu"}]}}))
        self.assertEqual(out2["status"], "edited")
        wb2 = openpyxl.load_workbook(path)
        self.assertEqual(wb2.sheetnames, ["Alt"])

    def test_unknown_sheet_is_error(self):
        path = self._create_base()
        out = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": path,
            "spec": {"ops": [{"op": "append_rows", "sheet": "Nope",
                              "rows": [[1]]}]}}))
        self.assertIn("error", out)
        self.assertIn("not found", out["error"])


# ---------------------------------------------------------------------------
# v2 features
# ---------------------------------------------------------------------------

def _build_multitable_workbook(path):
    """Two stacked tables on ONE sheet, separated by ≥2 blank rows — the
    report-sheet shape that used to flatten into one mis-typed grid."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["Monat", "Umsatz"])
    ws.append(["Jan", 100])
    ws.append(["Feb", 200])
    ws.append([])
    ws.append([])
    ws.append(["Region", "Anteil"])
    ws.append(["Nord", 0.6])
    ws.append(["Sued", 0.4])
    wb.save(path)


def _build_merged_header_workbook(path):
    """Two-row hierarchical header: merged 'Q1'/'Q2' bands over Umsatz|Marge."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quartale"
    ws["A1"] = "Land"
    ws["B1"] = "Q1"
    ws.merge_cells("B1:C1")
    ws["D1"] = "Q2"
    ws.merge_cells("D1:E1")
    ws["A2"] = "Land"
    ws["B2"] = "Umsatz"
    ws["C2"] = "Marge"
    ws["D2"] = "Umsatz"
    ws["E2"] = "Marge"
    ws.append(["AT", 10, 1, 20, 2])
    ws.append(["DE", 30, 3, 40, 4])
    wb.save(path)


class TestV2GridReading(_ToolFixture):
    def test_multi_table_split(self):
        p = os.path.join(self._tmp, "multi.xlsx")
        _build_multitable_workbook(p)
        out = json.loads(xlsx_tools.tool_xlsx_inspect({"path": p}))
        rep = out["report"]
        self.assertIn("Sheet: Report", rep)
        self.assertIn("Sheet: Report_2", rep)   # second block = own table
        self.assertIn("report_2(region, anteil)", rep)
        q = json.loads(xlsx_tools.tool_xlsx_query(
            {"path": p, "sql": "SELECT SUM(umsatz) FROM report"}))
        self.assertIn("| 300 |", q["result"])

    def test_merged_two_row_header_composes(self):
        p = os.path.join(self._tmp, "merged.xlsx")
        _build_merged_header_workbook(p)
        out = json.loads(xlsx_tools.tool_xlsx_inspect({"path": p}))
        rep = out["report"]
        self.assertIn("Q1 / Umsatz", rep)
        self.assertIn("Q2 / Marge", rep)
        q = json.loads(xlsx_tools.tool_xlsx_query(
            {"path": p, "sql": "SELECT SUM(q1_umsatz) FROM quartale"}))
        self.assertIn("| 40 |", q["result"])


class TestV2DeepInspect(_ToolFixture):
    def test_deep_orphans_duplicates_outliers(self):
        p = os.path.join(self._tmp, "deep.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Master"
        ws1.append(["KEY", "WERT"])
        for k, v in [("A", 1), ("B", 2), ("C", 3), ("D", 4), ("E", 5),
                     ("F", 6), ("G", 7), ("H", 8), ("I", 9), ("X", 9999)]:
            ws1.append([k, v])
        ws1.append(["A", 1])  # duplicate row
        ws2 = wb.create_sheet("Detail")
        ws2.append(["KEY", "MENGE"])
        for k in ["A", "B", "C", "D", "E", "F", "G", "H", "ORPHAN"]:
            ws2.append([k, 1])
        wb.save(p)
        out = json.loads(xlsx_tools.tool_xlsx_inspect({"path": p, "deep": True}))
        rep = out["report"]
        self.assertIn("Data quality:", rep)
        self.assertIn("fully duplicated row", rep)
        self.assertIn("outlier", rep)              # 9999 beyond 3×IQR
        self.assertIn("Referential findings", rep)
        self.assertIn("ORPHAN", rep)               # orphan detail key surfaced

    def test_deep_formula_map(self):
        p = os.path.join(self._tmp, "form.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calc"
        ws.append(["A", "B"])
        ws.append([1, "=Daten!A1*2"])
        ws.append([2, "=Daten!A2*2"])
        wb.create_sheet("Daten").append([10])
        wb.save(p)
        out = json.loads(xlsx_tools.tool_xlsx_inspect({"path": p, "deep": True}))
        rep = out["report"]
        self.assertIn("Formula map (deep)", rep)
        self.assertIn("references sheet `Daten` in 2 formula(s)", rep)


class TestV2QueryHandles(_ToolFixture):
    def test_save_as_and_reuse(self):
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": self.orders_path,
            "sql": ("SELECT marktordernummer, SUM(stueck) AS summe FROM "
                    "ausfuehrungen GROUP BY marktordernummer"),
            "save_as": "summen"}))
        self.assertEqual(out["saved_as"], "summen")
        # query FROM the handle
        q2 = json.loads(xlsx_tools.tool_xlsx_query({
            "path": "result:summen",
            "sql": "SELECT COUNT(*) FROM summen"}))
        self.assertIn("| 2 |", q2["result"])
        # create a sheet FROM the handle
        c = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "aus_handle.xlsx",
            "spec": {"sheets": [{"name": "S",
                                 "source": {"file": "result:summen"}}]}}))
        self.assertEqual(c["sheets"][0]["rows"], 2)

    def test_missing_handle_is_error(self):
        out = json.loads(xlsx_tools.tool_xlsx_query(
            {"path": "result:nope", "sql": "SELECT 1"}))
        self.assertIn("no stored result named", out["error"])


class TestV2CreateExtensions(_ToolFixture):
    def test_template_fill(self):
        tpl = os.path.join(self._tmp, "template.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bericht"
        ws["A1"] = "Firmenbericht"          # template content must survive
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=16)
        wb.save(tpl)
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "gefuellt.xlsx",
            "spec": {"template": {"file": tpl},
                     "sheets": [{"name": "Bericht", "anchor": "A3",
                                 "rows": [["x", 1], ["y", 2]]}]}}))
        self.assertEqual(out["sheets"][0]["rows"], 2)
        wb2 = openpyxl.load_workbook(out["path"])
        ws2 = wb2["Bericht"]
        self.assertEqual(ws2["A1"].value, "Firmenbericht")
        self.assertTrue(ws2["A1"].font.bold)        # template style intact
        self.assertEqual(ws2["A3"].value, "x")
        self.assertEqual(ws2["B4"].value, 2)

    def test_subtotals_autofilter_validation_print(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "extras.xlsx",
            "spec": {"sheets": [
                {
                    "name": "MD",
                    "master_detail": {
                        "key": "MARKTORDERNUMMER",
                        "master": {"source": {"file": self.orders_path,
                                              "sheet": "Orders"}},
                        "detail": {"source": {"file": self.orders_path,
                                              "sheet": "Ausfuehrungen"}},
                        "subtotals": ["STUECK"],
                    },
                },
                {
                    "name": "Flach",
                    "columns": [{"name": "Status",
                                 "choices": ["offen", "erledigt"]},
                                {"name": "Wert"}],
                    "rows": [["offen", 1], ["erledigt", 2]],
                    "autofilter": True,
                    "print": {"orientation": "landscape", "fit_width": True,
                              "repeat_header": True},
                },
            ]}}))
        wb = openpyxl.load_workbook(out["path"])
        md = wb["MD"]
        formulas = [c.value for row in md.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=SUM(")]
        self.assertEqual(len(formulas), 2)          # one subtotal per MO group
        labels = [c.value for row in md.iter_rows() for c in row
                  if c.value == "Zwischensumme"]
        self.assertEqual(len(labels), 2)
        flach = wb["Flach"]
        self.assertEqual(str(flach.auto_filter.ref), "A1:B3")
        self.assertEqual(len(flach.data_validations.dataValidation), 1)
        self.assertEqual(flach.page_setup.orientation, "landscape")
        self.assertEqual(flach.print_title_rows, "$1:$1")


class TestV2Diff(_ToolFixture):
    def test_keyed_diff(self):
        pa = os.path.join(self._tmp, "alt.xlsx")
        pb = os.path.join(self._tmp, "neu.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daten"
        ws.append(["ID", "Betrag"])
        ws.append(["k1", 10])
        ws.append(["k2", 20])
        ws.append(["k3", 30])
        wb.save(pa)
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Daten"
        ws2.append(["ID", "Betrag"])
        ws2.append(["k1", 10])       # unchanged
        ws2.append(["k2", 99])       # changed
        ws2.append(["k4", 40])       # added (k3 removed)
        wb2.save(pb)
        out = json.loads(xlsx_tools.tool_xlsx_diff({
            "path_a": pa, "path_b": pb, "key": "ID", "out": "diff.csv"}))
        rep = out["report"]
        self.assertIn("1 added, 1 removed, 1 changed", rep)
        self.assertIn("k2", rep)
        self.assertIn("'20' → '99'", rep)
        self.assertEqual(out["differences"], 3)
        with open(out["saved"]["path"], encoding="utf-8") as f:
            csv_text = f.read()
        self.assertIn("k2;Betrag;20;99", csv_text)

    def test_missing_key_is_error(self):
        out = json.loads(xlsx_tools.tool_xlsx_diff({
            "path_a": self.orders_path, "path_b": self.orders_path,
            "key": "GIBTSNICHT"}))
        self.assertIn("error", out)
        self.assertIn("must exist on both sides", out["error"])


# ---------------------------------------------------------------------------
# v3 features
# ---------------------------------------------------------------------------

class TestV3Pivot(_ToolFixture):
    def test_cross_tab_sum(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "pivot.xlsx",
            "spec": {"sheets": [{
                "name": "Pivot",
                "rows": [["Region", "Monat", "Umsatz"],
                         ["Nord", "Jan", 10], ["Nord", "Feb", 20],
                         ["Sued", "Jan", 5], ["Sued", "Jan", 7]],
                "pivot": {"rows": "Region", "cols": "Monat",
                          "values": "Umsatz", "agg": "sum"},
            }]}}))
        wb = openpyxl.load_workbook(out["path"])
        ws = wb["Pivot"]
        self.assertEqual(ws["A1"].value, "Region")
        self.assertEqual([ws.cell(1, c).value for c in (2, 3)], ["Feb", "Jan"])
        # Nord: Feb 20, Jan 10 · Sued: Jan 5+7=12
        self.assertEqual(ws["B2"].value, 20)
        self.assertEqual(ws["C2"].value, 10)
        self.assertEqual(ws["C3"].value, 12)
        self.assertEqual(ws["A4"].value, "Gesamt")
        self.assertEqual(ws["B4"].value, "=SUM(B2:B3)")

    def test_bad_agg_is_error(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "p2.xlsx",
            "spec": {"sheets": [{
                "name": "P", "rows": [["A", "B"], ["x", 1]],
                "pivot": {"rows": "A", "values": "B", "agg": "median"}}]}}))
        self.assertIn("error", out)


class TestV3ChartsV2(_ToolFixture):
    def test_scatter_stacked_secondary(self):
        out = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "charts2.xlsx",
            "spec": {"sheets": [{
                "name": "C",
                "columns": [{"name": "X"}, {"name": "A"}, {"name": "B"},
                            {"name": "Quote"}],
                "rows": [[1, 5, 2, 0.5], [2, 6, 3, 0.6], [3, 7, 4, 0.7]],
                "charts": [
                    {"type": "scatter", "labels": "X", "series": ["A"]},
                    {"type": "bar", "labels": "X", "series": ["A", "B"],
                     "stacked": True},
                    {"type": "bar", "labels": "X",
                     "series": ["A", "Quote"], "secondary": ["Quote"]},
                ],
            }]}}))
        wb = openpyxl.load_workbook(out["path"])
        charts = wb["C"]._charts
        self.assertEqual(len(charts), 3)
        types = {type(c).__name__ for c in charts}
        self.assertIn("ScatterChart", types)
        stacked = [c for c in charts
                   if getattr(c, "grouping", "") == "stacked"]
        self.assertEqual(len(stacked), 1)
        self.assertEqual(stacked[0].overlap, 100)


class TestV3DiffV2(_ToolFixture):
    def _two_files(self):
        pa = os.path.join(self._tmp, "a.xlsx")
        pb = os.path.join(self._tmp, "b.xlsx")
        for p, rows in ((pa, [["k1", "x", 10], ["k1", "y", 20]]),
                        (pb, [["k1", "x", 10], ["k1", "y", 99]])):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "D"
            ws.append(["KUNDE", "TYP", "WERT"])
            for r in rows:
                ws.append(r)
            wb.save(p)
        return pa, pb

    def test_composite_key(self):
        pa, pb = self._two_files()
        out = json.loads(xlsx_tools.tool_xlsx_diff({
            "path_a": pa, "path_b": pb, "key": "KUNDE,TYP"}))
        self.assertIn("keyed on `KUNDE + TYP`", out["report"])
        self.assertIn("1 changed", out["report"])
        self.assertEqual(out["differences"], 1)

    def test_highlighted_xlsx_output(self):
        pa, pb = self._two_files()
        out = json.loads(xlsx_tools.tool_xlsx_diff({
            "path_a": pa, "path_b": pb, "key": "KUNDE,TYP",
            "out": "diff.xlsx"}))
        wb = openpyxl.load_workbook(out["saved"]["path"])
        ws = wb.active
        # changed cell C3 (WERT of k1|y) is yellow + carries the old value
        cell = ws["C3"]
        self.assertEqual(cell.value, 99)
        self.assertEqual(cell.fill.start_color.rgb[-6:], "FFF2AB")
        self.assertIn("vorher: 20", cell.comment.text)

    def test_formula_compare(self):
        pa = os.path.join(self._tmp, "fa.xlsx")
        pb = os.path.join(self._tmp, "fb.xlsx")
        for p, formula in ((pa, "=A2*2"), (pb, "=A2*3")):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "F"
            ws.append(["ID", "CALC"])
            ws.append([1, formula])
            wb.save(p)
        out = json.loads(xlsx_tools.tool_xlsx_diff({
            "path_a": pa, "path_b": pb, "key": "ID",
            "compare": "formulas"}))
        self.assertIn("'=A2*2' → '=A2*3'", out["report"])


class TestV3LegacyAndRecalc(_ToolFixture):
    def test_recalc_computes_formula_values(self):
        if xlsx_tools._find_soffice() is None:
            self.skipTest("soffice not installed")
        created = json.loads(xlsx_tools.tool_xlsx_create({
            "path": "calc.xlsx",
            "spec": {"sheets": [{
                "name": "T",
                "columns": [{"name": "A"}, {"name": "B"}],
                "rows": [[2, 3], [4, 5]],
            }]}}))
        edited = json.loads(xlsx_tools.tool_xlsx_edit({
            "path": created["path"],
            "recalc": True,
            "spec": {"ops": [{"op": "add_column", "sheet": "T",
                              "name": "Summe", "formula": "=A{row}+B{row}"}]}}))
        self.assertEqual(edited["status"], "edited")
        # after recalc the formula VALUES are queryable
        q = json.loads(xlsx_tools.tool_xlsx_query({
            "path": created["path"], "sql": "SELECT SUM(summe) FROM t"}))
        self.assertIn("| 14 |", q["result"])

    def test_ods_readable(self):
        if xlsx_tools._find_soffice() is None:
            self.skipTest("soffice not installed")
        # build an ods by converting an xlsx via soffice (round-trip test)
        import subprocess
        soffice = xlsx_tools._find_soffice()
        subprocess.run([soffice, "--headless", "--convert-to", "ods",
                        "--outdir", self._tmp, self.orders_path],
                       capture_output=True, timeout=120, check=True)
        ods = os.path.join(self._tmp, "orders.ods")
        self.assertTrue(os.path.isfile(ods))
        out = json.loads(xlsx_tools.tool_xlsx_query({
            "path": ods, "sql": "SELECT COUNT(*) FROM orders"}))
        self.assertIn("| 2 |", out["result"])


if __name__ == "__main__":
    unittest.main()
