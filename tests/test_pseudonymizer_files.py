"""Unit tests for the file walker layer of the pseudonymizer.

Strategy: build a tiny but real file in each supported format (containing
known PII), pseudonymise it, read the bytes back to verify the original
value is gone, then deanonymise into a third file and verify the roundtrip
preserves every original value. Mapping is fresh per test so cross-test
collisions don't happen.

Run with: python -m unittest tests.test_pseudonymizer_files -v
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pseudonymizer as ps  # noqa: E402

# Known PII strings. IBAN + credit card chosen because both have checksum-
# strict rules and shape-preserving fakes — exercises both the scanner and
# the format-preserving generators on the file path.
TEST_IBAN = "DE89370400440532013000"
TEST_CC = "4111-1111-1111-1111"   # Visa test number, Luhn-valid
TEST_PHONE = "+49 30 12345678"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_docx(path: str, body: str) -> None:
    """Write a minimal but valid .docx with one paragraph containing `body`."""
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body><w:p><w:r><w:t xml:space="preserve">' + body +
        '</w:t></w:r></w:p></w:body></w:document>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)


def _make_pptx(path: str, body: str) -> None:
    """Minimal .pptx with one slide carrying `body`."""
    slide_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        '<p:cSld><p:spTree>'
        '<p:sp><p:txBody><a:p><a:r><a:t>' + body + '</a:t></a:r></a:p></p:txBody></p:sp>'
        '</p:spTree></p:cSld></p:sld>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/ppt/slides/slide1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("ppt/slides/slide1.xml", slide_xml)


def _read_office_text(path: str, member: str) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        return zf.read(member).decode("utf-8")


def _make_xlsx(path: str, *, body_iban: str, formula: str = "=A1*2") -> None:
    """Write a real xlsx via openpyxl with one PII string + one formula."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = body_iban
    ws["A2"] = formula           # formula — walker must skip
    ws["A3"] = "plain text"
    ws["B1"] = 12345             # numeric — walker must skip
    wb.save(path)


def _read_xlsx_a1(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return ws["A1"].value or ""


# ---------------------------------------------------------------------------
# Plain text
# ---------------------------------------------------------------------------


class TestPlainText(unittest.TestCase):

    def _roundtrip(self, ext: str):
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "input" + ext)
            anon = os.path.join(td, "anon" + ext)
            restored = os.path.join(td, "restored" + ext)
            body = f"Bitte überweise auf {TEST_IBAN} — danke."
            with open(src, "w", encoding="utf-8") as f:
                f.write(body)

            mapping = ps.new_mapping()
            try:
                added = ps.pseudonymize_file(src, anon, mapping=mapping)
                self.assertGreaterEqual(added, 1,
                                        "expected at least one new mapping entry")
                with open(anon, encoding="utf-8") as f:
                    anon_text = f.read()
                self.assertNotIn(TEST_IBAN, anon_text,
                                 "original IBAN must be absent from anonymised file")

                restored_count = ps.deanonymize_file(anon, restored, mapping=mapping)
                self.assertGreaterEqual(restored_count, 1)
                with open(restored, encoding="utf-8") as f:
                    self.assertEqual(f.read(), body)
            finally:
                ps.close_mapping(mapping.mapping_id)

    def test_txt_roundtrip(self):
        self._roundtrip(".txt")

    def test_md_roundtrip(self):
        self._roundtrip(".md")

    def test_log_roundtrip(self):
        self._roundtrip(".log")


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class TestCsv(unittest.TestCase):

    def test_csv_roundtrip(self):
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "input.csv")
            anon = os.path.join(td, "anon.csv")
            restored = os.path.join(td, "restored.csv")
            # Cell 0 = PII, cell 1 = innocent. CSV writer preserves layout.
            with open(src, "w", encoding="utf-8", newline="") as f:
                f.write(f"name,iban\nAlice,{TEST_IBAN}\nBob,DE12500105170648489890\n")

            mapping = ps.new_mapping()
            try:
                ps.pseudonymize_file(src, anon, mapping=mapping)
                with open(anon, encoding="utf-8") as f:
                    anon_text = f.read()
                self.assertNotIn(TEST_IBAN, anon_text)
                self.assertIn("Alice", anon_text)   # innocent cell preserved

                ps.deanonymize_file(anon, restored, mapping=mapping)
                with open(src, encoding="utf-8") as f:
                    src_text = f.read()
                with open(restored, encoding="utf-8") as f:
                    restored_text = f.read()
                self.assertIn(TEST_IBAN, restored_text)
                self.assertIn("Alice", restored_text)
                src_cells = sorted(src_text.replace("\r", "").strip().split())
                restored_cells = sorted(
                    restored_text.replace("\r", "").strip().split())
                self.assertEqual(src_cells, restored_cells)
            finally:
                ps.close_mapping(mapping.mapping_id)


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------


class TestDocx(unittest.TestCase):

    def test_docx_iban_roundtrip(self):
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "input.docx")
            anon = os.path.join(td, "anon.docx")
            restored = os.path.join(td, "restored.docx")
            body = f"Konto: {TEST_IBAN} bitte überweisen."
            _make_docx(src, body)

            mapping = ps.new_mapping()
            try:
                added = ps.pseudonymize_file(src, anon, mapping=mapping)
                self.assertGreaterEqual(added, 1)
                anon_doc = _read_office_text(anon, "word/document.xml")
                self.assertNotIn(TEST_IBAN, anon_doc,
                                 "original IBAN leaked into anonymised docx")
                # Token presence — shape-fake for IBAN preserves country prefix
                # but body must differ. Verify by checking the digits.
                # (We can't grep for "<IBAN_" because IBAN is shape-preserving.)
                self.assertIn("DE", anon_doc)  # country preserved

                ps.deanonymize_file(anon, restored, mapping=mapping)
                restored_doc = _read_office_text(restored, "word/document.xml")
                self.assertIn(TEST_IBAN, restored_doc,
                              "original IBAN didn't return after deanonymise")
            finally:
                ps.close_mapping(mapping.mapping_id)


# ---------------------------------------------------------------------------
# PPTX
# ---------------------------------------------------------------------------


class TestPptx(unittest.TestCase):

    def test_pptx_iban_roundtrip(self):
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "input.pptx")
            anon = os.path.join(td, "anon.pptx")
            restored = os.path.join(td, "restored.pptx")
            body = f"Sensitive {TEST_IBAN} on this slide."
            _make_pptx(src, body)

            mapping = ps.new_mapping()
            try:
                added = ps.pseudonymize_file(src, anon, mapping=mapping)
                self.assertGreaterEqual(added, 1)
                anon_slide = _read_office_text(anon, "ppt/slides/slide1.xml")
                self.assertNotIn(TEST_IBAN, anon_slide)

                ps.deanonymize_file(anon, restored, mapping=mapping)
                restored_slide = _read_office_text(restored, "ppt/slides/slide1.xml")
                self.assertIn(TEST_IBAN, restored_slide)
            finally:
                ps.close_mapping(mapping.mapping_id)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


class TestXlsx(unittest.TestCase):

    def test_xlsx_iban_roundtrip_with_formula_preserved(self):
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "input.xlsx")
            anon = os.path.join(td, "anon.xlsx")
            restored = os.path.join(td, "restored.xlsx")
            _make_xlsx(src, body_iban=TEST_IBAN, formula="=A1&B1")

            mapping = ps.new_mapping()
            try:
                added = ps.pseudonymize_file(src, anon, mapping=mapping)
                self.assertGreaterEqual(added, 1)

                import openpyxl
                wb = openpyxl.load_workbook(anon)
                ws = wb.active
                self.assertNotEqual(ws["A1"].value, TEST_IBAN,
                                    "A1 still contains original IBAN")
                # Formula must be untouched (string starting with '=').
                self.assertEqual(ws["A2"].value, "=A1&B1")
                self.assertEqual(ws["A3"].value, "plain text")
                self.assertEqual(ws["B1"].value, 12345)

                ps.deanonymize_file(anon, restored, mapping=mapping)
                wb2 = openpyxl.load_workbook(restored)
                self.assertEqual(wb2.active["A1"].value, TEST_IBAN)
            finally:
                ps.close_mapping(mapping.mapping_id)


# ---------------------------------------------------------------------------
# Dispatch / error paths
# ---------------------------------------------------------------------------


class TestDispatch(unittest.TestCase):

    def test_unsupported_extension_raises(self):
        from engine.file_pseudonymize import FilePseudonymizeError
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "x.bin")
            with open(src, "wb") as f:
                f.write(b"\x00\x01\x02")
            mapping = ps.new_mapping()
            try:
                with self.assertRaises(FilePseudonymizeError):
                    ps.pseudonymize_file(src, os.path.join(td, "out.bin"),
                                         mapping=mapping)
            finally:
                ps.close_mapping(mapping.mapping_id)

    def test_deanonymize_unsupported_copies_through(self):
        # LLM may write a .png that we never pseudonymised — deanonymise must
        # not blow up, just copy through.
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "x.png")
            dst = os.path.join(td, "out.png")
            payload = b"\x89PNG\r\n\x1a\n" + b"\x00" * 50
            with open(src, "wb") as f:
                f.write(payload)
            mapping = ps.new_mapping()
            try:
                n = ps.deanonymize_file(src, dst, mapping=mapping)
                self.assertEqual(n, 0)
                with open(dst, "rb") as f:
                    self.assertEqual(f.read(), payload)
            finally:
                ps.close_mapping(mapping.mapping_id)

    def test_source_defaults_to_attachment_basename(self):
        # Verify the convenience default of `source=`.
        with tempfile.TemporaryDirectory() as td:
            src = os.path.join(td, "report.md")
            with open(src, "w") as f:
                f.write(f"Konto: {TEST_IBAN}")
            mapping = ps.new_mapping()
            try:
                ps.pseudonymize_file(src, os.path.join(td, "anon.md"),
                                     mapping=mapping)
                self.assertIn("attachment:report.md", mapping.sources)
            finally:
                ps.close_mapping(mapping.mapping_id)


# ---------------------------------------------------------------------------
# Stability across files in the same mapping
# ---------------------------------------------------------------------------


class TestCrossFileStability(unittest.TestCase):
    """Same original value across multiple files must produce the same
    token within one mapping — that's the invariant that lets a multi-file
    user submission share one anonymisation surface."""

    def test_same_iban_same_token_across_files(self):
        with tempfile.TemporaryDirectory() as td:
            src_a = os.path.join(td, "a.md")
            src_b = os.path.join(td, "b.md")
            anon_a = os.path.join(td, "anon_a.md")
            anon_b = os.path.join(td, "anon_b.md")
            body = f"IBAN: {TEST_IBAN}"
            for p in (src_a, src_b):
                with open(p, "w") as f:
                    f.write(body)

            mapping = ps.new_mapping()
            try:
                ps.pseudonymize_file(src_a, anon_a, mapping=mapping)
                ps.pseudonymize_file(src_b, anon_b, mapping=mapping)
                with open(anon_a) as f:
                    a_out = f.read()
                with open(anon_b) as f:
                    b_out = f.read()
                self.assertEqual(a_out, b_out,
                                 "same IBAN in two files must produce identical anon text")
                # Both source labels recorded on the mapping.
                self.assertEqual(
                    sorted(s for s in mapping.sources if s.startswith("attachment:")),
                    sorted(["attachment:a.md", "attachment:b.md"]))
            finally:
                ps.close_mapping(mapping.mapping_id)


if __name__ == "__main__":
    unittest.main()
