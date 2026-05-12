"""HTTP handlers for the Data Workbench (/v1/data/*).

A Data Workbench session is a *real* chat session under the hood — created
via the same `SessionManager` the rest of the app uses, just flagged
`is_data_workbench` so `_build_system_prompt` injects the DATA WORKBENCH
block and the `data_viz` tool group is enabled. The conversation itself
goes through the ordinary `/v1/chat` endpoint; these endpoints only handle
session creation, file upload (→ a per-session DuckDB file inside the
session's artifact folder), and listing tables.

Reuse, don't rebuild — the GDPR scanner, the artifact folder convention,
the auth/RBAC layer, the multipart parser are all called, not reimplemented.
"""
from __future__ import annotations

import json
import os
import re

# Per-upload cap. DuckDB compresses on disk; the raw bytes are what we guard.
MAX_UPLOAD_BYTES = 200 * 1024 * 1024
_TABLE_NAME_RE = re.compile(r"[^A-Za-z0-9_]+")
# Local-scan sample size: rows of the table we run the GDPR scanner over and
# show the model. The full column stays in DuckDB.
_SAMPLE_ROWS = 20


def _slug_table_name(filename: str, sheet: str | None = None) -> str:
    """Turn an upload's name (+ optional sheet) into a safe DuckDB identifier."""
    base = os.path.splitext(os.path.basename(filename))[0]
    name = _TABLE_NAME_RE.sub("_", base).strip("_") or "t"
    if name[0].isdigit():
        name = "t_" + name
    if sheet:
        s = _TABLE_NAME_RE.sub("_", sheet).strip("_")
        if s:
            name = f"{name}__{s}"
    return name[:120]


class DataVizHandlerMixin:
    """Mixin with /v1/data/* handlers."""

    # ─── Session lifecycle ──────────────────────────────────────────────────

    def _handle_data_create_session(self):
        """POST /v1/data/sessions — create a workbench session.

        Body (all optional): {model, title}. Returns {session_id, model, title}.
        """
        body = self._read_json() or {}
        model = (body.get("model") or "").strip() or server_config["default_model"]
        title = (body.get("title") or "").strip()
        user = getattr(self, "_auth_user", _auth_mod.SYNTHETIC_ADMIN)
        if not _auth_mod.can_access_model(user, model):
            self._send_json({"error": f"Access to model '{model}' not permitted"}, 403)
            return
        try:
            provider = self._resolve_provider(model)
        except Exception as e:
            self._send_json({"error": f"provider resolution failed: {e}"}, 400)
            return

        session = sessions.create(
            agent_id="main",
            model=model,
            api_key=provider["api_key"],
            base_url=provider["base_url"],
            max_context=engine.get_model_max_context(model),
        )
        session.is_data_workbench = True
        if title:
            session.title = title
            try:
                ChatDB.save_session(session.id, session.agent_id, session.model,
                                    session.title, session.status,
                                    session.created_at, session.last_active,
                                    session.project or "")
            except Exception:
                pass

        # Owner stamping (mirrors _handle_create_session).
        uid = ""
        if user and user.get("id"):
            if user["id"] != "__system__":
                uid = user["id"]
            else:
                try:
                    _users = _auth_mod.AuthDB.list_users()
                    if _users:
                        uid = _users[0]["id"]
                except Exception:
                    pass
        if uid:
            session.user_id = uid
            try:
                ChatDB.update_session_user(session.id, uid)
            except Exception:
                pass

        from server_lib.db import DataSessionDB
        DataSessionDB.create(sid=session.id, agent_id="main", user_id=uid, title=title)
        self._send_json({"session_id": session.id, "model": model, "title": title})

    def _handle_data_list_sessions(self):
        """GET /v1/data/sessions — workbench sessions visible to the caller."""
        from server_lib.db import DataSessionDB
        user = getattr(self, "_auth_user", _auth_mod.SYNTHETIC_ADMIN)
        is_admin = bool(user and user.get("role") == "admin")
        uid = (user or {}).get("id") or ""
        rows = DataSessionDB.list_all() if is_admin else DataSessionDB.list_for_user(uid)
        out = []
        for r in rows:
            try:
                tables = json.loads(r.get("tables_json") or "[]")
            except Exception:
                tables = []
            out.append({
                "session_id": r["sid"], "title": r.get("title") or "",
                "tables": tables, "created_at": r.get("created_at"),
                "updated_at": r.get("updated_at"),
            })
        self._send_json({"sessions": out})

    # ─── Upload + tables ────────────────────────────────────────────────────

    def _data_session_or_403(self, sid: str):
        """Return (session, db_dir) for a workbench session the caller owns, or None (after sending an error)."""
        from server_lib.db import DataSessionDB
        meta = DataSessionDB.get(sid)
        if not meta:
            self._send_json({"error": "workbench session not found"}, 404)
            return None
        user = getattr(self, "_auth_user", _auth_mod.SYNTHETIC_ADMIN)
        is_admin = bool(user and user.get("role") == "admin")
        uid = (user or {}).get("id") or ""
        owner = meta.get("user_id") or ""
        if owner and uid and owner != uid and not is_admin:
            self._send_json({"error": "not your workbench session"}, 403)
            return None
        session = sessions.get(sid)
        if session is None:
            self._send_json({"error": "session vanished"}, 404)
            return None
        # Resolve the artifact folder (same convention as python_exec).
        folder = engine._get_artifact_session_folder(sid)
        db_dir = os.path.join(engine.AGENTS_DIR, session.agent_id, "artifacts", folder)
        os.makedirs(db_dir, exist_ok=True)
        return session, db_dir

    def _handle_data_upload(self, sid: str):
        """POST /v1/data/sessions/<sid>/upload — multipart .xlsx/.csv → DuckDB table(s).

        Returns {tables: [...], scan: {dirty_columns: [...]}}.
        """
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, db_dir = ctx

        ctype = self.headers.get("Content-Type", "")
        if not ctype.startswith("multipart/form-data"):
            self._send_json({"error": "multipart/form-data required"}, 400)
            return
        boundary = None
        for part in ctype.split(";"):
            part = part.strip()
            if part.startswith("boundary="):
                boundary = part.split("=", 1)[1].strip('"')
                break
        if not boundary:
            self._send_json({"error": "missing boundary"}, 400)
            return
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0 or length > MAX_UPLOAD_BYTES + 64 * 1024:
            self._send_json({"error": "payload too large"}, 413)
            return
        raw = self.rfile.read(length)
        from handlers.translate import _parse_multipart
        _fields, file_name, file_bytes = _parse_multipart(raw, boundary)
        if not file_name or not file_bytes:
            self._send_json({"error": "missing file"}, 400)
            return
        if len(file_bytes) > MAX_UPLOAD_BYTES:
            self._send_json({"error": "file too large"}, 413)
            return
        ext = os.path.splitext(file_name)[1].lower()
        if ext not in (".csv", ".tsv", ".xlsx", ".xlsm"):
            self._send_json({"error": f"unsupported file type '{ext}' — Phase A supports .csv / .tsv / .xlsx"}, 400)
            return

        try:
            import duckdb
        except ImportError:
            self._send_json({"error": "duckdb is not installed on the server"}, 500)
            return

        db_path = os.path.join(db_dir, "_data.duckdb")
        # Persist the upload alongside the DuckDB so deanonymise / preserve-format
        # work in later PRs can find the original.
        src_path = os.path.join(db_dir, os.path.basename(file_name))
        try:
            with open(src_path, "wb") as f:
                f.write(file_bytes)
        except OSError as e:
            self._send_json({"error": f"save failed: {e}"}, 500)
            return

        con = None
        created = []
        try:
            con = duckdb.connect(db_path)
            existing = {r[0] for r in con.execute("SHOW TABLES").fetchall()}
            if ext in (".csv", ".tsv"):
                tname = _slug_table_name(file_name)
                t = tname; i = 2
                while t in existing:
                    t = f"{tname}_{i}"; i += 1
                con.execute(f'CREATE TABLE "{t}" AS SELECT * FROM read_csv_auto(?, sample_size=-1)', [src_path])
                created.append(t); existing.add(t)
            else:
                # xlsx — read each sheet with openpyxl, register as a DataFrame.
                import openpyxl
                wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    header = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(rows[0])]
                    # De-dupe / sanitise header names for SQL.
                    seen = {}
                    cols = []
                    for h in header:
                        h2 = _TABLE_NAME_RE.sub("_", h).strip("_") or "col"
                        if h2[0].isdigit():
                            h2 = "c_" + h2
                        if h2 in seen:
                            seen[h2] += 1; h2 = f"{h2}_{seen[h2]}"
                        else:
                            seen[h2] = 0
                        cols.append(h2)
                    data = [list(r) for r in rows[1:]]
                    import pandas as pd  # pandas ships alongside openpyxl in this env
                    df = pd.DataFrame(data, columns=cols)
                    tname = _slug_table_name(file_name, sheet if len(wb.sheetnames) > 1 else None)
                    t = tname; i = 2
                    while t in existing:
                        t = f"{tname}_{i}"; i += 1
                    con.register("_df_tmp", df)
                    con.execute(f'CREATE TABLE "{t}" AS SELECT * FROM _df_tmp')
                    con.unregister("_df_tmp")
                    created.append(t); existing.add(t)
                wb.close()

            # Build per-table summary + a scanned sample.
            summary = []
            dirty_cols = []
            for t in created:
                cols_info = con.execute(f'DESCRIBE "{t}"').fetchall()
                col_names = [c[0] for c in cols_info]
                col_types = {c[0]: c[1] for c in cols_info}
                n_rows = con.execute(f'SELECT count(*) FROM "{t}"').fetchone()[0]
                sample_rows = con.execute(f'SELECT * FROM "{t}" LIMIT {_SAMPLE_ROWS}').fetchall()
                # GDPR scan over the sample's cell values + per-column.
                masked_sample = []
                flagged_here = set()
                for row in sample_rows:
                    out_row = []
                    for cname, val in zip(col_names, row):
                        sval = "" if val is None else str(val)
                        try:
                            findings = engine._pii_scan_text(sval, max_findings=3) if sval else []
                        except Exception:
                            findings = []
                        if findings:
                            flagged_here.add(cname)
                            out_row.append("«PII»")
                        else:
                            out_row.append(val)
                    masked_sample.append(out_row)
                for cname in sorted(flagged_here):
                    dirty_cols.append({"table": t, "column": cname})
                summary.append({
                    "name": t, "n_rows": n_rows,
                    "columns": [{"name": c, "type": col_types.get(c, "")} for c in col_names],
                    "flagged_columns": sorted(flagged_here),
                    "sample": [[ (None if v is None else (v if not isinstance(v, (bytes, bytearray)) else "<binary>")) for v in r] for r in masked_sample],
                })

            # Update DataSessionDB index with the cumulative table list.
            all_tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
        except Exception as e:
            self._send_json({"error": f"ingest failed: {type(e).__name__}: {e}"}, 500)
            return
        finally:
            if con is not None:
                try:
                    con.close()
                except Exception:
                    pass

        try:
            from server_lib.db import DataSessionDB
            DataSessionDB.set_tables(sid, all_tables)
        except Exception:
            pass
        # Audit line.
        try:
            if engine._audit_log:
                engine._audit_log.log_action(
                    agent="main", action_type="data_upload", tool_name="data_upload",
                    args_summary=f"file={os.path.basename(file_name)} tables={created}",
                    result_summary=f"{len(created)} tables, {len(dirty_cols)} flagged columns",
                    session_id=sid, source="data_workbench")
        except Exception:
            pass

        self._send_json({"tables": summary, "scan": {"dirty_columns": dirty_cols},
                         "all_tables": all_tables})

    def _handle_data_tables(self, sid: str):
        """GET /v1/data/sessions/<sid>/tables — list tables, row counts, schemas, sample."""
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, db_dir = ctx
        db_path = os.path.join(db_dir, "_data.duckdb")
        if not os.path.exists(db_path):
            self._send_json({"tables": []})
            return
        try:
            import duckdb
        except ImportError:
            self._send_json({"error": "duckdb is not installed on the server"}, 500)
            return
        con = None
        try:
            con = duckdb.connect(db_path, read_only=True)
            out = []
            for (t,) in con.execute("SHOW TABLES").fetchall():
                cols_info = con.execute(f'DESCRIBE "{t}"').fetchall()
                col_names = [c[0] for c in cols_info]
                n_rows = con.execute(f'SELECT count(*) FROM "{t}"').fetchone()[0]
                sample = con.execute(f'SELECT * FROM "{t}" LIMIT {_SAMPLE_ROWS}').fetchall()
                masked = []
                flagged = set()
                for row in sample:
                    out_row = []
                    for cname, val in zip(col_names, row):
                        sval = "" if val is None else str(val)
                        try:
                            f = engine._pii_scan_text(sval, max_findings=2) if sval else []
                        except Exception:
                            f = []
                        if f:
                            flagged.add(cname); out_row.append("«PII»")
                        else:
                            out_row.append(val if not isinstance(val, (bytes, bytearray)) else "<binary>")
                    masked.append(out_row)
                out.append({
                    "name": t, "n_rows": n_rows,
                    "columns": [{"name": c, "type": d[1]} for c, d in zip(col_names, cols_info)],
                    "flagged_columns": sorted(flagged),
                    "sample": masked,
                })
            self._send_json({"tables": out})
        except Exception as e:
            self._send_json({"error": f"{type(e).__name__}: {e}"}, 500)
        finally:
            if con is not None:
                try:
                    con.close()
                except Exception:
                    pass

    # ─── Anonymise / deanonymise / mapping ──────────────────────────────────

    def _data_prime_threadlocals(self, sid: str, session):
        """Make engine._thread_local look like a workbench chat turn so the
        tool body resolves the artifact folder + _after_file_write fires."""
        engine._thread_local.current_session_id = sid
        try:
            engine._thread_local.current_agent = engine.AgentConfig(session.agent_id)
        except Exception:
            pass
        if not getattr(engine._thread_local, "event_callback", None):
            engine._thread_local.event_callback = lambda *_a, **_k: None
        engine._thread_local.data_workbench = True

    def _data_clear_threadlocals(self):
        try:
            engine._thread_local.current_session_id = None
            engine._thread_local.current_agent = None
            engine._thread_local.event_callback = None
            engine._thread_local.data_workbench = False
        except Exception:
            pass

    def _handle_data_anonymise(self, sid: str):
        """POST /v1/data/sessions/<sid>/anonymise — thin no-LLM call into tool_data_anonymise.

        Body: {table, columns:[{name,strategy,opts?}], output_format?, new_table?,
               mapping_out?, source_file?, embed_salt?, k?}
        """
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, _db_dir = ctx
        body = self._read_json() or {}
        body.pop("mode", None)  # this endpoint is anonymise-only
        self._data_prime_threadlocals(sid, session)
        try:
            result_str = engine.tool_data_anonymise({**body, "mode": "anonymise"})
        finally:
            self._data_clear_threadlocals()
        try:
            result = json.loads(result_str)
        except Exception:
            result = {"error": "tool returned non-JSON"}
        status = 400 if isinstance(result, dict) and result.get("error") else 200
        self._send_json(result, status)

    def _handle_data_deanonymise(self, sid: str):
        """POST /v1/data/sessions/<sid>/deanonymise — restore from an index file.

        Body: {source_file, index_file, output_format?}. RBAC: session owner / admin
        (already enforced by _data_session_or_403).
        """
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, _db_dir = ctx
        body = self._read_json() or {}
        self._data_prime_threadlocals(sid, session)
        try:
            result_str = engine.tool_data_anonymise({**body, "mode": "deanonymise"})
        finally:
            self._data_clear_threadlocals()
        try:
            result = json.loads(result_str)
        except Exception:
            result = {"error": "tool returned non-JSON"}
        status = 400 if isinstance(result, dict) and result.get("error") else 200
        self._send_json(result, status)

    def _handle_data_mapping(self, sid: str, fname: str):
        """GET /v1/data/sessions/<sid>/mapping/<file> — download an index/mapping
        workbook. The download is itself an audit event (mapping_downloaded) —
        the index is the re-identification key. RBAC: session owner / admin."""
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, db_dir = ctx
        # Defend against path traversal: basename only, must live in the session dir.
        safe = os.path.basename(fname)
        if not safe or safe != fname or "/" in fname or ".." in fname:
            self._send_json({"error": "bad filename"}, 400)
            return
        path = os.path.join(db_dir, safe)
        if not os.path.isfile(path):
            self._send_json({"error": "mapping file not found"}, 404)
            return
        try:
            with open(path, "rb") as f:
                data = f.read()
        except OSError as e:
            self._send_json({"error": f"read failed: {e}"}, 500)
            return
        # Audit the access — every fetch of a re-identification key is logged.
        try:
            if engine._audit_log:
                engine._audit_log.log_action(
                    agent="main", action_type="mapping_downloaded", tool_name="mapping_downloaded",
                    args_summary=f"session/{sid} file={safe}",
                    result_summary=f"{len(data)} bytes", session_id=sid, source="data_workbench")
        except Exception:
            pass
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                         if safe.lower().endswith(".xlsx") else "application/octet-stream")
        self.send_header("Content-Disposition", f'attachment; filename="{safe}"')
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        try:
            self.wfile.write(data)
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass

    # ─── File-level GDPR scan (§17) ─────────────────────────────────────────

    def _handle_data_scan(self, sid: str):
        """POST /v1/data/sessions/<sid>/scan — thin no-LLM call into tool_data_scan_files.

        Body (optional): {tables: [str]} — omit to scan every table in the workbench.
        """
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, _db_dir = ctx
        body = self._read_json() or {}
        self._data_prime_threadlocals(sid, session)
        try:
            result_str = engine.tool_data_scan_files({"tables": body.get("tables") or []})
        finally:
            self._data_clear_threadlocals()
        try:
            result = json.loads(result_str)
        except Exception:
            result = {"error": "tool returned non-JSON"}
        status = 400 if isinstance(result, dict) and result.get("error") else 200
        self._send_json(result, status)

    # ─── Chart re-render (no-LLM) ───────────────────────────────────────────

    def _handle_data_render(self, sid: str):
        """POST /v1/data/sessions/<sid>/render — re-render an edited Vega-Lite spec.

        Body: {spec: dict, table: str, scale?: float}. Same body the
        data_render_chart tool calls — no LLM in the loop.
        """
        ctx = self._data_session_or_403(sid)
        if ctx is None:
            return
        session, _db_dir = ctx
        body = self._read_json() or {}
        self._data_prime_threadlocals(sid, session)
        try:
            result_str = engine.tool_data_render_chart({
                "spec": body.get("spec"), "table": body.get("table"),
                "scale": body.get("scale"),
            })
        finally:
            self._data_clear_threadlocals()
        try:
            result = json.loads(result_str)
        except Exception:
            result = {"error": "tool returned non-JSON"}
        status = 400 if isinstance(result, dict) and result.get("error") else 200
        self._send_json(result, status)
