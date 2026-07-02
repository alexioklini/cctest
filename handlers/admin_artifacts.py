"""Artifacts, files, channels, nodes, sidecar, services, backup, worker, refine/soul, and notification handlers.

Sub-mixin of AdminHandlerMixin (handlers/admin.py module-split refactor). Holds
ONLY this area's `_handle_*` methods (+ area-only private helpers).
AdminHandlerMixin inherits this class, so the combined BrainAgentHandler MRO is
unchanged.

Like admin.py, this module references `engine`, `brain`, `client`, `_db_conn`,
`sqlite3`, `subprocess`, etc. as BARE MODULE GLOBALS injected at runtime by
server._inject_server_globals(). This module's name is in that function's
injection list. All other helpers (`_send_json`, `_read_json`,
`_parse_agent_from_path`, …) resolve via `self.` against the combined MRO.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import time
import threading
import urllib.request
import urllib.error
import uuid
from urllib.parse import unquote, urlencode


class AdminArtifactsHandlers:
    """Artifacts, files, channels, nodes, sidecar, services, backup, worker, refine/soul, and notification handlers."""

    def _handle_refine(self):
        """POST /v1/refine — refine text with LLM one-shot call.

        Accepts optional `purpose` to swap the system prompt for non-chat
        targets:
          - "" / "chat_prompt"      → rewrite as a clearer chat prompt (default)
          - "profile_field"         → polish a free-text profile entry
                                      (e.g. job_description, comm prefs)
        Anything else falls back to chat_prompt behavior."""
        body = self._read_json()
        text = body.get("text", "") or body.get("content", "")
        context = body.get("context", "general")
        purpose = (body.get("purpose") or "").strip().lower()
        field_label = (body.get("field_label") or "").strip()
        # Two-tier refine (REFINE_ENHANCEMENT_DESIGN.md): "polish" = the
        # conservative grammar/clarity cleaner (default, unchanged behaviour);
        # "engineer" = intent-extract + restructure + grounding. Engineer is a
        # no-op for profile_field (a bio has nothing to engineer) → falls back
        # to polish. The Engineer prompts here are the spec validated by
        # eval/refine_eval.py (build_new) — keep the two in sync.
        tier = (body.get("tier") or "polish").strip().lower()
        if tier not in ("polish", "engineer"):
            tier = "polish"
        if purpose == "profile_field":
            tier = "polish"  # engineering a profile field is meaningless
        # Optional caveman compression for the *refine LLM call itself*
        # (0 off / 1 lite / 2 full / 3 ultra). Compresses the polish system
        # prompt + appends the chat-style suffix so the refiner produces
        # tighter output without us touching its rules.
        try:
            caveman = int(body.get("caveman") or 0)
        except (TypeError, ValueError):
            caveman = 0
        if caveman not in (0, 1, 2, 3):
            caveman = 0
        if not text:
            self._send_json({"error": "No text provided"}, 400)
            return

        # Find model: request body > tools_config setting > auto-select
        refine_model = body.get("model")
        if not refine_model:
            tc = engine.get_tool_config()
            refine_model = tc.get("refinement", {}).get("model", "")
        if not refine_model or not engine._is_model_available(refine_model):
            refine_model = engine._background_model_default()
        if not refine_model:
            self._send_json({"error": "No model available for refinement"}, 503)
            return

        provider = self._resolve_provider(refine_model)

        # Build context from current session — chat-prompt mode only. Profile
        # polishing must NOT read chat history (privacy, and the polish prompt
        # doesn't need it anyway).
        session_id = body.get("session_id", "")
        agent_id = body.get("agent", "main")
        project = body.get("project", "")
        chat_context = ""

        if purpose not in ("profile_field", "soul"):
            # Get agent info
            try:
                agent_cfg = engine.AgentConfig(agent_id)
                soul_summary = (agent_cfg.soul or "")[:200]
                if soul_summary:
                    chat_context += f"Agent: {agent_id} — {soul_summary}\n"
            except Exception:
                pass

            # Get recent conversation for context (last 5 messages)
            if session_id:
                try:
                    s = sessions.get(session_id)
                    if s and s.messages:
                        recent = s.messages[-5:]
                        chat_context += "Recent conversation:\n"
                        for m in recent:
                            role = m.get("role", "?")
                            content = m.get("content", "")
                            if isinstance(content, str):
                                chat_context += f"  [{role}] {content[:150]}\n"
                        chat_context += "\n"
                except Exception:
                    pass

            if project:
                chat_context += f"Active project: {project}\n"

        context_block = ""
        if chat_context:
            context_block = (
                f"\nCONTEXT (use this to make the rewrite more specific and relevant):\n"
                f"{chat_context}\n"
            )

        # Engineer-tier grounding: model hint + active tool NAMES + project
        # instructions excerpt. Built ONLY for engineer (the polish one-click
        # path pays nothing extra). Tool names let the rewrite say e.g. "read
        # the file with read_document first"; the model hint steers
        # reasoning-native / local targets; project instructions keep the
        # rewrite inside project discipline.
        ground_block = ""
        if tier == "engineer" and purpose != "profile_field":
            grounds = []
            hint = self._refine_model_hint(refine_model)
            if hint:
                grounds.append(hint)
            try:
                _tools = engine.resolve_active_tools(
                    purpose="interactive", agent_id=agent_id)
                _names = [t.get("name", "") for t in _tools if t.get("name")]
                if _names:
                    grounds.append(
                        "Available tools: " + ", ".join(sorted(_names)[:40]) +
                        " — reference them by name when the task needs one.")
            except Exception:
                pass
            if project:
                try:
                    _proj = engine.get_project(agent_id, project) or {}
                    _instr = (_proj.get("instructions") or "").strip()
                    if _instr:
                        grounds.append(
                            f"Active project '{project}' instructions (respect them): "
                            + _instr[:400])
                except Exception:
                    pass
            if grounds:
                ground_block = "\nGROUNDING:\n" + "\n".join(grounds) + "\n"

        if purpose == "profile_field":
            # Polish a free-text profile entry. Different rules: the user
            # is describing themselves, not asking the AI a question, so
            # we keep first-person voice and don't re-frame as a request.
            label_hint = f" The field is: {field_label}." if field_label else ""
            instructions = (
                "You are a TEXT POLISHER for a user profile field." + label_hint + " "
                "The user is describing themselves or their preferences. "
                "Your job is to lightly polish what they wrote.\n"
                "CRITICAL RULES:\n"
                "- Output ONLY the polished text, nothing else.\n"
                "- Keep first-person voice (\"I am…\", \"I prefer…\") if present.\n"
                "- Do NOT add new facts, opinions, or content the user didn't write.\n"
                "- Do NOT answer or respond — just clean up what's there.\n"
                "- Fix grammar, spelling, punctuation, awkward phrasing.\n"
                "- Preserve line breaks and paragraph structure when present.\n"
                "- Keep the user's tone (formal/casual) and language.\n"
                "- If the input is already clean, return it unchanged.\n"
                "- No markdown headings, no bullet rewrites unless the input had them."
            )
            request_line = (
                "Polish this profile text (output ONLY the polished "
                "version, preserve line breaks):\n\n" + text
            )
        elif purpose == "soul" and tier == "engineer":
            # Engineer soul: structural improvement (tighten, dedupe, surface
            # implied guardrails) WITHOUT changing identity. Validated against
            # eval/refine_eval.py build_new soul branch — restraint is the
            # default; over-editing a tight soul is a failure.
            instructions = (
                "You are an EDITOR for an AI agent's soul.md (its system prompt, "
                "second person: 'You are ...', 'Your job is ...'). Improve it "
                "structurally without changing who the agent is.\n"
                "CRITICAL RULES:\n"
                "- Output ONLY the improved soul. No commentary.\n"
                "- Keep second-person voice. Keep the agent's name, role, and "
                "listed tools.\n"
                "- You MAY: tighten wording, remove redundancy, group related "
                "rules, surface a missing stop-condition/guardrail that the "
                "existing rules clearly imply.\n"
                "- You MUST NOT: invent new capabilities, tools, or behaviours "
                "the user didn't imply; remove an existing rule; change the "
                "tone; pad with ceremony.\n"
                "- Do NOT add Markdown you weren't given: no new bold/**emphasis**, "
                "no converting bullets to numbered lists, no extra nesting, and "
                "NEVER wrap the whole soul in a ```code fence```. Match the "
                "input's existing formatting exactly.\n"
                "- Preserve all Markdown structure and code/inline `code` exactly.\n"
                "- DEFAULT TO RETURNING IT UNCHANGED. Only edit if there is a real "
                "grammar error, true redundancy, or a clearly-implied missing "
                "guardrail. If the soul already reads cleanly, return it "
                "byte-for-byte. Restructuring a good soul is a failure."
            )
            request_line = (
                "Improve this soul.md (output ONLY the improved version):\n\n"
                + text
            )
        elif purpose == "soul":
            # Polish an agent's soul.md — its system prompt that defines
            # personality, role, and behavioural rules. Different rules
            # again: this is *imperative second-person* prose addressed to
            # the agent ("You are …", "Your job is …"). We must NOT flip
            # it into first or third person, must NOT change the agent's
            # name/role/tools, and must preserve any embedded code/command
            # examples and section structure.
            instructions = (
                "You are a TEXT POLISHER for an AI agent's soul.md "
                "(its system prompt). The soul defines the agent's "
                "identity, role, and behavioural rules — it is written in "
                "second person ('You are …', 'Your job is …'). Your job "
                "is to lightly polish what the user wrote without "
                "changing meaning.\n"
                "CRITICAL RULES:\n"
                "- Output ONLY the polished soul, nothing else.\n"
                "- Keep second-person voice ('You are …', 'Your job …'). "
                "Do NOT switch to first or third person.\n"
                "- Do NOT change the agent's name, role, or capabilities.\n"
                "- Do NOT add new behaviours, tools, or rules. Do NOT "
                "remove existing rules.\n"
                "- Do NOT answer or respond — just clean up what's there.\n"
                "- Fix grammar, spelling, punctuation, awkward phrasing, "
                "redundancy.\n"
                "- Preserve Markdown structure: headings (#, ##, ###), "
                "bullet lists, numbered lists, blockquotes, horizontal "
                "rules — keep them all.\n"
                "- Preserve code blocks and inline `code` exactly. Do not "
                "rewrite commands, paths, tool names, or examples.\n"
                "- Preserve line breaks and paragraph structure.\n"
                "- Keep the existing tone (terse / verbose / playful / "
                "formal) — do not normalise it.\n"
                "- If the input is already clean, return it unchanged."
            )
            request_line = (
                "Polish this soul.md (output ONLY the polished version, "
                "preserve all Markdown structure and code blocks):\n\n"
                + text
            )
        elif tier == "engineer":
            # Engineer chat / scheduled_task: intent-extract + restructure +
            # grounding + (for scheduled) unattended discipline. Validated
            # against eval/refine_eval.py build_new — restraint is default,
            # ask-back on hopelessly-vague drafts, no invented specifics.
            instructions = (
                "You are a PROMPT ENGINEER for an AI assistant. The user gives you a "
                "rough draft of what they want the assistant to do. Turn it into a "
                "noticeably STRONGER, more effective prompt that gets the right "
                "result on the first try. A good rewrite is clearly more capable "
                "than the draft — not a near-copy with the typos fixed.\n"
                "DO add (this is the value — apply whatever the task needs):\n"
                "- A precise task verb (replace 'fix/make/handle/do' with the exact "
                "operation).\n"
                "- The expected OUTPUT shape when implied (format, structure, length, "
                "language) — e.g. 'as a bulleted list', 'a single function', 'in 3 "
                "sentences'.\n"
                "- An explicit success criterion when the task has one ('Done when: "
                "...').\n"
                "- A role/expert framing when the task is specialized.\n"
                "- Structure (steps, sections) when the request is multi-part.\n"
                "THE ONE HARD LIMIT — do NOT INVENT FALSE FACTS the draft didn't give: "
                "no specific filenames, paths, URLs, numbers, API fields, library "
                "names, or pixel sizes the user never mentioned. Adding structure, "
                "format, and explicitness is REQUIRED; fabricating concrete details "
                "is FORBIDDEN. (Generic placeholders like '[the relevant file]' are "
                "fine; a made-up 'index.html' is not.)\n"
                "OTHER RULES:\n"
                "- PROPORTION: the rewrite should be at most ~2× the draft's length "
                "unless the draft is genuinely vague and needs real structure. A "
                "one-line, already-clear request should come back as a tightened one- "
                "or two-line prompt — never a multi-section spec. If you're adding "
                "more than the task needs, cut it.\n"
                "- Output ONLY the rewritten prompt. No commentary, no 'here is'.\n"
                "- Preserve the user's actual intent and language. Do NOT answer the "
                "request yourself.\n"
                "- If two unrelated tasks are mixed, keep the primary one and note the "
                "split in ONE trailing line '(Second task: ...)'.\n"
                "- CALIBRATE to the draft. If it is already strong — it already "
                "names a clear task AND its scope (and, for a recurring task, a stop "
                "condition) — then it does NOT need your scaffolding: do only light "
                "tightening and do NOT bolt on a role, a multi-section format, a data "
                "flow, or 'Done when' that it didn't ask for. Adding ceremony to an "
                "already-complete prompt is a FAILURE. Save the heavy structuring for "
                "drafts that are actually rough or vague.\n"
                "- If the draft is so under-specified that even adding structure would "
                "require GUESSING the actual goal (e.g. 'fix the bug' with no hint of "
                "which bug), do NOT invent it — instead return a short prompt that "
                "asks the user for the missing piece(s). One focused question beats a "
                "confident wrong guess.\n"
                "- DO NOT OVER-STRICTIFY a casual factual lookup. If the draft is a "
                "casual everyday question whose answer is a quick web lookup (weather, "
                "exchange/stock price, sports score, opening hours, 'what is X'), keep "
                "it casual: fix spelling/grammar and stop. Do NOT add words that demand "
                "precision or an official source ('präzise', 'genau', 'exakt', "
                "'verbindlich', 'offizielle Quelle', 'precise', 'exact', 'authoritative/"
                "official source', 'to N decimal places', 'real-time'), and do NOT impose "
                "a rigid output spec. Those raise the assistant's evidentiary bar so it "
                "REFUSES ordinary web results instead of just answering — the opposite "
                "of helpful. 'wie wird das wetter morgen in wien' → 'Wie wird das Wetter "
                "morgen in Wien?', NOT 'Gib eine präzise Wettervorhersage … aus "
                "offizieller Quelle'.\n"
                "- For simple requests output plain prose. For genuinely complex "
                "multi-part requests you MAY use <context>/<task>/<constraints> XML "
                "sections. No commentary outside the prompt."
            )
            if purpose == "scheduled_task":
                instructions += (
                    "\nThis prompt runs UNATTENDED on a schedule. Additionally, but "
                    "ONLY if the draft does not already cover them (do NOT restate "
                    "what's already there):\n"
                    "- If no stop/completion condition is stated, add a brief one.\n"
                    "- If the task performs a destructive action "
                    "(delete/overwrite/send/transfer) and has NO safeguard, add: "
                    "'Stop and report instead of acting if uncertain.'\n"
                    "- If it relies on info that may be missing, add: 'Report "
                    "instead of guessing if information is missing.'\n"
                    "Add nothing else. A well-scoped scheduled draft comes back "
                    "essentially unchanged."
                )
            instructions += ground_block + context_block
            request_line = (
                f"Rewrite this draft (output ONLY the rewritten prompt):\n\n{text}")
        else:
            instructions = (
                "You are a PROMPT REWRITER for an AI chat system. "
                "The user will give you a draft prompt/message they want to send to an AI assistant. "
                "Your job is to rewrite it into a better, clearer version of the SAME request. "
                "CRITICAL RULES:\n"
                "- Output ONLY the rewritten prompt, nothing else\n"
                "- Do NOT answer the question or fulfill the request — REWRITE it\n"
                "- Do NOT add explanations, analysis, alternatives, or commentary\n"
                "- Do NOT use markdown headings, bullet points, or formatting\n"
                "- The output replaces the user's input in a chat box — it must be a clean prompt\n"
                "- Fix grammar, spelling, punctuation\n"
                "- Make the request clearer and more specific using the context provided\n"
                "- Keep the same intent and language\n"
                "Example: Input: 'whats weather vienna' → Output: 'What is the weather like in Vienna today?'"
                + context_block
            )
            request_line = f"Rewrite this prompt (output ONLY the rewritten version):\n\n{text}"

        # Caveman (v9.120.0): caveman is OUTPUT-only — it never compresses the
        # system prompt or tool descriptions. The ONE place the INPUT query gets
        # caveman-compressed is here, during refinement: we (1) instruct the
        # refiner to write the rewrite in the requested terse style (the chat
        # style instruction is a legitimate instruction to the refiner, not a
        # compression of our own rules), and (2) deterministically compress the
        # REFINED TEXT the refiner returns (see _caveman_compress_text below), so
        # the query the user sends is itself caveman. The refiner's instructions
        # are left intact (readable) — only its OUTPUT is the target.
        if caveman in (1, 2, 3):
            instructions = instructions + engine.CAVEMAN_CHAT_PROMPTS.get(caveman, "")
        # Build the wire-level messages: prepend the (possibly compressed)
        # instructions to the user's request-line, since /v1/refine doesn't
        # use _build_system_prompt — the rules HAVE to ride in the user msg.
        wire_content = instructions + "\n\n" + request_line

        # GDPR policy gate. Both the user-typed `text` and the assembled
        # `chat_context` (last 5 messages of the session, chat-prompt path
        # only) flow into wire_content; profile/soul polishing paths still
        # carry the user's own free text. Scan once over the assembled
        # blob — the deanon callback rebuilds originals on the reply.
        _refine_deanon = engine._identity_deanon
        try:
            refine_model, (_pii_content,), _refine_deanon = engine.gdpr_pick_model_for_background(
                refine_model, [wire_content], purpose=f"refine_{purpose or 'chat_prompt'}_{tier}")
            wire_content = _pii_content
        except engine.GDPRBlockedError as e:
            self._send_json({"error": f"refine blocked by GDPR policy: {e}"}, 503)
            return

        messages = [{"role": "user", "content": wire_content}]

        try:
            from . import sidecar_proxy as _sidecar_proxy
            _res = _sidecar_proxy.background_call(
                messages=messages,
                model=refine_model,
                agent_id=agent_id,
                session_id=session_id or "",
                project=project or "",
                cost_purpose="refine",
                provider_resolver=self._resolve_provider,
            )
            result = _refine_deanon(_res.get("reply") or "")
            if _res.get("error") and not result:
                self._send_json({"error": str(_res["error"])}, 500)
                return
            # Deterministically compress the REFINED query text to the active
            # caveman level — this is the input-side compression that used to
            # (wrongly) live on the system prompt. The user's refined query lands
            # in the composer already in caveman form. Code/URLs/paths survive
            # (the rule-based pass leaves them intact).
            if result and caveman in (1, 2, 3):
                result = engine._caveman_compress_text(result, caveman)
            self._send_json({"refined": result or text, "model": refine_model,
                             "caveman": caveman, "tier": tier})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _refine_model_hint(self, model: str) -> str:
        """One short, non-drifting steering hint for the Engineer tier, derived
        from OUR model config (never a static external claim). Reasoning-native
        targets must not get CoT scaffolding; local/open-weight targets want
        flat explicit prompts. Returns "" when neither applies."""
        try:
            is_local = engine.is_model_local(model)
        except Exception:
            is_local = False
        m = ((engine._server_config() or {}).get("models", {}) or {}).get(model, {})
        inf = (m.get("inference", {}) or {}) if isinstance(m, dict) else {}
        name = (model or "").lower()
        reasoning = bool(inf.get("thinking_level")) or any(
            t in name for t in ("r1", "o3", "o4-mini", "qwen3", "reason", "think"))
        if reasoning:
            return ("Target model reasons internally — do NOT add 'think step by "
                    "step' or other reasoning scaffolding; state the goal and the "
                    "desired output cleanly.")
        if is_local:
            return ("Target is a local/open-weight model — keep the prompt flat "
                    "and explicit; avoid deep nesting.")
        return ""

    def _handle_soul_chat(self, path):
        """POST /v1/agents/<id>/soul-chat — chat to edit soul.md with LLM."""
        parts = path.split("/")
        agent_id = parts[3]
        body = self._read_json()
        message = body.get("message", "").strip()
        soul = body.get("soul", "")
        history = body.get("history", [])

        if not message:
            self._send_json({"error": "No message provided"}, 400)
            return

        # Resolve model (same policy as refine): refinement-tool override →
        # server default. No haiku/cheapest heuristics.
        tc = engine.get_tool_config()
        model = tc.get("refinement", {}).get("model", "")
        if not model or not engine._is_model_available(model):
            model = engine._background_model_default()
        if not model:
            self._send_json({"error": "No model available"}, 503)
            return

        provider = self._resolve_provider(model)

        system_block = (
            "You are a soul.md editor assistant. The user wants to modify an agent's soul.md file "
            "(system prompt that defines the agent's personality and behavior).\n\n"
            "CURRENT SOUL.MD:\n```\n" + soul + "\n```\n\n"
            "RULES:\n"
            "- Help the user edit, improve, or rewrite the soul.md based on their instructions\n"
            "- When you make changes, output the COMPLETE updated soul.md inside a ```soul\n...\n``` code block\n"
            "- You may also provide brief commentary outside the code block\n"
            "- If the user is just asking a question or discussing (not requesting changes), respond normally without a code block\n"
            "- Preserve existing structure and formatting unless asked to change it\n"
            "- Keep the same voice/style unless the user wants a different one\n"
        )

        messages = [{"role": "user", "content": system_block}, {"role": "assistant", "content": "I understand. I'm ready to help you edit this agent's soul.md. What changes would you like to make?"}]
        for h in history:
            messages.append({"role": h["role"], "content": h["content"]})
        messages.append({"role": "user", "content": message})

        try:
            from . import sidecar_proxy as _sidecar_proxy
            _res = _sidecar_proxy.background_call(
                messages=messages,
                model=model,
                agent_id=agent_id,
                cost_purpose="soul_chat",
                provider_resolver=self._resolve_provider,
            )
            if _res.get("error") and not _res.get("reply"):
                self._send_json({"error": str(_res["error"])}, 500)
                return
            self._send_json({"reply": _res.get("reply") or "", "model": model})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_services_status(self):
        """GET /v1/services — status of all managed services."""
        uptime = int(time.time() - _server_start_time)
        tg_running = self._is_telegram_running()

        self._send_json({
            "server": {
                "status": "running",
                "version": engine.VERSION,
                "version_date": engine.VERSION_DATE,
                "pid": os.getpid(),
                "uptime_seconds": uptime,
                "sessions": len(sessions.list_all()),
                "agents": engine.list_agents(),
                "scheduler_tasks": len(engine._scheduler.list_all()) if engine._scheduler else 0,
                "default_provider": next((name for name, p in server_config.get("providers", {}).items() if p.get("default_model") == server_config.get("default_model")), ""),
                "default_model": server_config.get("default_model", ""),
                "attachment_image_model": server_config.get("attachment_image_model", ""),
                "chat_summary_model": server_config.get("chat_summary_model", ""),
                "auto_route_classifier_mode": (server_config.get("auto_route", {}) or {}).get("classifier_mode", "keywords"),
                "gdpr_scanner": {
                    "enabled": bool(server_config.get("gdpr_scanner", {}).get("enabled", True)),
                    "server_log": bool(server_config.get("gdpr_scanner", {}).get("server_log", True)),
                    "name_precision_gate": bool(server_config.get("gdpr_scanner", {}).get("name_precision_gate", True)),
                    "block_unscannable_on_cloud": bool(server_config.get("gdpr_scanner", {}).get("block_unscannable_on_cloud", False)),
                    # Confidence-band thresholds (9.195.0) — replaced server_block.
                    "confidence_lower": float(server_config.get("gdpr_scanner", {}).get("confidence_lower", 0.50)),
                    "confidence_upper": float(server_config.get("gdpr_scanner", {}).get("confidence_upper", 0.85)),
                    "default_local_fallback_model": str(server_config.get("gdpr_scanner", {}).get("default_local_fallback_model") or ""),
                    "background_pii_action": (
                        server_config.get("gdpr_scanner", {}).get("background_pii_action")
                        if server_config.get("gdpr_scanner", {}).get("background_pii_action")
                            in ("anonymise", "swap_to_local", "skip", "abort")
                        else "anonymise"
                    ),
                    "background_ask_action": (
                        server_config.get("gdpr_scanner", {}).get("background_ask_action")
                        if server_config.get("gdpr_scanner", {}).get("background_ask_action")
                            in ("anonymise", "swap_to_local", "ignore")
                        else "anonymise"
                    ),
                    "background_anonymise_fail_action": (
                        server_config.get("gdpr_scanner", {}).get("background_anonymise_fail_action")
                        if server_config.get("gdpr_scanner", {}).get("background_anonymise_fail_action")
                            in ("swap_to_local", "abort")
                        else "swap_to_local"
                    ),
                    # Merge defaults UNDER saved so newly-added categories
                    # (e.g. business_id) always surface even when an older
                    # saved config predates them.
                    "categories": {
                        **{cat: {"action": act} for cat, act in engine.PII_DEFAULT_CATEGORY_ACTIONS.items()},
                        **(server_config.get("gdpr_scanner", {}).get("categories") or {}),
                    },
                    "rule_overrides": server_config.get("gdpr_scanner", {}).get("rule_overrides") or {},
                    # Per-rule min_occurrences — seeded from code defaults, with
                    # any saved overrides merged on top so the UI shows the
                    # effective thresholds (e.g. date=10) even before the admin
                    # touches them.
                    "min_occurrences": {
                        **dict(engine.PII_DEFAULT_MIN_OCCURRENCES),
                        **(server_config.get("gdpr_scanner", {}).get("min_occurrences") or {}),
                    },
                    # Per-rule count_points [lo,hi] — count→score calibration. The
                    # effective points (resolved via _pii_count_points, which
                    # seeds from min_occurrences) so the UI shows real values even
                    # before the admin edits them.
                    "count_points": {
                        rid: list(engine._pii_count_points(rid))
                        for rid in engine.PII_RULE_CATEGORIES
                    },
                    "email_allowlist": server_config.get("gdpr_scanner", {}).get("email_allowlist") or [],
                    # Static PII catalog (rule→category map, category labels,
                    # default actions, rule labels). Single source of truth for
                    # the Settings → GDPR panel + chat-view labels. Moved here
                    # from the deleted browser-side PIIScanner object (9.200.0)
                    # so the client never duplicates the rule catalog again.
                    "catalog": {
                        "rule_categories": dict(engine.PII_RULE_CATEGORIES),
                        "category_labels": dict(engine.PII_CATEGORY_LABELS),
                        "default_category_actions": dict(engine.PII_DEFAULT_CATEGORY_ACTIONS),
                        "rule_labels": dict(engine.PII_RULE_LABELS),
                    },
                },
                "available_tools": sorted(engine.TOOL_DISPATCH.keys()),
            },
            "telegram": {
                "status": "running" if tg_running else "stopped",
                "bot": _telegram_mod.telegram_service.bot_username if tg_running else "",
                "enabled": server_config.get("telegram_enabled", True),
            },
            "channels": _adapters_mod.channel_manager.status() if _adapters_mod.channel_manager else [],
            "nodes": self._get_nodes_summary(),
        })

    def _get_nodes_summary(self):
        """Get a summary of node statuses."""
        with _node_lock:
            total = len(_node_registry)
            connected = sum(1 for info in _node_registry.values() if info["status"] == "connected")
            return {"total": total, "connected": connected}

    def _handle_service_log(self):
        """GET /v1/services/log?name=server|qmd&lines=100 — tail a service log."""
        qs = self.path.split("?", 1)[1] if "?" in self.path else ""
        params = dict(p.split("=", 1) for p in qs.split("&") if "=" in p)
        name = params.get("name", "server")
        lines = min(int(params.get("lines", "100")), 500)

        log_paths = {
            "server": os.path.expanduser("~/.brain-agent/server.log"),
            "qmd": os.path.expanduser("~/.brain-agent/qmd.log"),
        }
        path = log_paths.get(name)
        if not path or not os.path.isfile(path):
            self._send_json({"name": name, "lines": [], "error": "Log file not found"})
            return

        try:
            with open(path, "r", errors="replace") as f:
                all_lines = f.readlines()
            tail = [l.rstrip("\n") for l in all_lines[-lines:]]
            self._send_json({"name": name, "lines": tail, "total": len(all_lines)})
        except Exception as e:
            self._send_json({"name": name, "lines": [], "error": str(e)})

    def _handle_telegram_action(self):
        """POST /v1/services/telegram — start/stop/restart/enable/disable Telegram."""
        body = self._read_json()
        action = body.get("action", "")
        svc = _telegram_mod.telegram_service

        if action == "start":
            ok = _start_telegram_service()
            self._send_json({"status": "started" if ok else "error",
                             "running": svc.running, "error": svc.error})

        elif action == "stop":
            svc.stop()
            self._send_json({"status": "stopped", "running": False})

        elif action == "restart":
            svc.stop()
            ok = _start_telegram_service()
            self._send_json({"status": "restarted" if ok else "error",
                             "running": svc.running, "error": svc.error})

        elif action == "enable":
            _set_telegram_enabled(True)
            if not svc.running:
                _start_telegram_service()
            self._send_json({"status": "enabled", "running": svc.running,
                             "enabled": True})

        elif action == "disable":
            _set_telegram_enabled(False)
            svc.stop()
            self._send_json({"status": "disabled", "running": False,
                             "enabled": False})

        else:
            self._send_json({"error": f"Unknown action: {action}"}, 400)

    def _handle_restart(self):
        """POST /v1/restart — restart the server process."""
        self._send_json({"status": "restarting"})
        # Schedule restart after response is sent
        def do_restart():
            time.sleep(0.5)
            os.execv(sys.executable, [sys.executable] + sys.argv)
        threading.Thread(target=do_restart, daemon=True).start()

    # --- Chat answer handler (interactive AskUserQuestion) ---

    def _handle_chat_answer(self):
        """POST /v1/chat/answer — deliver a user answer to a pending ask_user tool call.

        Body shapes:
          {session_id, answer: "..."}                             # single question
          {session_id, answers: {"<question>": "<answer>", ...}}  # batch
        """
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except Exception:
            self._send_json({"error": "invalid JSON body"}, 400)
            return
        session_id = (body.get("session_id") or "").strip()
        answer = body.get("answer")
        answers = body.get("answers")
        if not session_id or (answer is None and not isinstance(answers, dict)):
            self._send_json({"error": "session_id and answer/answers are required"}, 400)
            return
        if self._session_access_check(session_id) is None:
            return
        # Normalize answers dict values to strings
        if isinstance(answers, dict):
            answers = {str(k): str(v) for k, v in answers.items() if v is not None}
        from brain import deliver_ask_user_answer
        ok = deliver_ask_user_answer(
            session_id,
            answer=str(answer) if answer is not None else None,
            answers=answers if isinstance(answers, dict) and answers else None,
        )
        if not ok:
            self._send_json({"error": "no pending question for this session"}, 404)
            return
        self._send_json({"delivered": True, "session_id": session_id})

    # --- Notification handlers ---

    def _handle_notifications_list(self):
        """GET /v1/notifications — list recent notifications."""
        if not _notification_manager:
            self._send_json({"notifications": [], "unread": 0})
            return
        notifs = _notification_manager.get_notifications(limit=50)
        unread = _notification_manager.get_unread_count()
        self._send_json({"notifications": notifs, "unread": unread})

    def _handle_notifications_unread(self):
        """GET /v1/notifications/unread — get unread count."""
        count = _notification_manager.get_unread_count() if _notification_manager else 0
        self._send_json({"unread": count})

    def _handle_notifications_settings_post(self):
        """POST /v1/notifications/settings — save notification config."""
        body = self._read_json()
        if not _notification_manager:
            self._send_json({"error": "Notification manager not initialized"}, 500)
            return
        _notification_manager.update_config(body)
        # Persist to config.json
        try:
            config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.json")
            config = {}
            if os.path.exists(config_path):
                with open(config_path) as f:
                    config = json.load(f)
            config["notifications"] = body
            with open(config_path, "w") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            self._send_json({"error": f"Failed to save: {e}"}, 500)
            return
        self._send_json({"status": "saved"})

    def _handle_notifications_dismiss(self):
        """POST /v1/notifications/dismiss — dismiss notification(s)."""
        body = self._read_json()
        nid = body.get("id")
        if not _notification_manager:
            self._send_json({"error": "Not initialized"}, 500)
            return
        if nid == "all":
            _notification_manager.clear_all()
        elif nid:
            _notification_manager.dismiss(nid)
        self._send_json({"status": "dismissed"})

    def _handle_notifications_read(self):
        """POST /v1/notifications/read — mark notification(s) as read."""
        body = self._read_json()
        nid = body.get("id")  # None = mark all read
        if _notification_manager:
            _notification_manager.mark_read(nid)
        self._send_json({"status": "read"})

    # --- Backup / Restore handlers ---

    def _handle_backup_info(self):
        """GET /v1/backup/info — return what would be backed up."""
        import tarfile as _tarfile
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        agents_dir = os.path.join(base, "agents")
        agent_names = engine.list_agents()
        total_files = 0
        total_size = 0
        agent_info = []
        for aname in agent_names:
            adir = os.path.join(agents_dir, aname)
            mems = len([f for f in os.listdir(adir) if f.endswith(".md")]) if os.path.isdir(adir) else 0
            skills_dir = os.path.join(adir, "skills")
            skills = len(os.listdir(skills_dir)) if os.path.isdir(skills_dir) else 0
            agent_info.append({"name": aname, "memories": mems, "skills": skills})
            if os.path.isdir(adir):
                for root, dirs, files in os.walk(adir):
                    dirs[:] = [d for d in dirs if d != "__pycache__"]
                    for f in files:
                        if not f.endswith((".pyc", ".DS_Store")):
                            fp = os.path.join(root, f)
                            total_files += 1
                            try:
                                total_size += os.path.getsize(fp)
                            except OSError:
                                pass
        self._send_json({
            "agents": agent_info,
            "agent_count": len(agent_names),
            "total_files": total_files,
            "estimated_size_bytes": total_size,
        })

    def _handle_backup_create(self):
        """POST /v1/backup — create a tar.gz backup archive."""
        import tarfile as _tarfile
        import tempfile
        body = self._read_json()
        backup_type = body.get("type", "full")
        target_agent = body.get("agent")
        include_keys = body.get("include_keys", False)

        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        agents_dir = os.path.join(base, "agents")
        backup_dir = os.path.join(base, "backups")
        os.makedirs(backup_dir, exist_ok=True)

        _EXCLUDE = {"__pycache__", ".DS_Store", "node_modules"}
        _EXCLUDE_EXT = {".pyc", ".db-wal", ".db-shm"}

        def _should_exclude(name):
            base_name = os.path.basename(name)
            if base_name in _EXCLUDE:
                return True
            _, ext = os.path.splitext(base_name)
            if ext in _EXCLUDE_EXT:
                return True
            return False

        ts = time.strftime("%Y%m%dT%H%M%S")
        if backup_type == "agent" and target_agent:
            fname = f"{target_agent.lower()}-{ts}.brain-backup.tar.gz"
        else:
            fname = f"backup-{ts}.brain-backup.tar.gz"
        backup_path = os.path.join(backup_dir, fname)

        try:
            with _tarfile.open(backup_path, "w:gz") as tar:
                prefix = f"backup-{ts}"

                # Add config.json (with redacted keys)
                config_path = os.path.join(base, "config.json")
                if os.path.exists(config_path):
                    with open(config_path) as f:
                        config = json.load(f)
                    if not include_keys:
                        # Redact API keys
                        for pname, pcfg in config.get("providers", {}).items():
                            if "api_key" in pcfg:
                                pcfg["api_key"] = "REDACTED"
                        if "gmail" in config:
                            for k in list(config["gmail"].keys()):
                                if "password" in k.lower() or "secret" in k.lower():
                                    config["gmail"][k] = "REDACTED"
                    redacted_json = json.dumps(config, indent=2).encode("utf-8")
                    import io
                    info = _tarfile.TarInfo(name=f"{prefix}/config.json")
                    info.size = len(redacted_json)
                    tar.addfile(info, io.BytesIO(redacted_json))

                # Add agents
                agents_to_backup = [target_agent] if (backup_type == "agent" and target_agent) else engine.list_agents()
                for aname in agents_to_backup:
                    adir = os.path.join(agents_dir, aname)
                    if not os.path.isdir(adir):
                        continue
                    for root, dirs, files in os.walk(adir):
                        dirs[:] = [d for d in dirs if d not in _EXCLUDE]
                        for f in files:
                            if _should_exclude(f):
                                continue
                            fp = os.path.join(root, f)
                            arcname = f"{prefix}/agents/{aname}/{os.path.relpath(fp, adir)}"
                            try:
                                tar.add(fp, arcname=arcname)
                            except (OSError, PermissionError):
                                pass

                # Add databases (full backup only)
                if backup_type != "agent":
                    for db_name in ("chats.db", "scheduler.db", "costs.db"):
                        db_path = os.path.join(agents_dir, "main", db_name)
                        if os.path.exists(db_path):
                            # Safe SQLite copy using backup API
                            import sqlite3
                            tmp_db = os.path.join(backup_dir, f"_tmp_{db_name}")
                            try:
                                src = sqlite3.connect(db_path)
                                dst = sqlite3.connect(tmp_db)
                                src.backup(dst)
                                src.close()
                                dst.close()
                                tar.add(tmp_db, arcname=f"{prefix}/databases/{db_name}")
                            except Exception:
                                # Fallback: direct copy
                                tar.add(db_path, arcname=f"{prefix}/databases/{db_name}")
                            finally:
                                try:
                                    os.unlink(tmp_db)
                                except OSError:
                                    pass

            size = os.path.getsize(backup_path)
            self._send_json({
                "status": "created",
                "path": backup_path,
                "filename": fname,
                "size_bytes": size,
                "type": backup_type,
                "agents": agents_to_backup,
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._send_json({"error": str(e)}, 500)

    def _handle_restore(self):
        """POST /v1/restore — restore from a backup archive."""
        import tarfile as _tarfile
        body = self._read_json()
        backup_path = body.get("path", "")
        strategy = body.get("strategy", "merge")

        if not backup_path or not os.path.exists(backup_path):
            self._send_json({"error": f"Backup file not found: {backup_path}"}, 400)
            return

        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        agents_dir = os.path.join(base, "agents")

        try:
            imported = {"agents": [], "memories": 0, "files": 0}
            with _tarfile.open(backup_path, "r:gz") as tar:
                members = tar.getmembers()
                # Find the prefix (first directory component)
                prefix = ""
                for m in members:
                    parts = m.name.split("/")
                    if len(parts) > 1:
                        prefix = parts[0]
                        break

                for member in members:
                    if member.isdir():
                        continue
                    parts = member.name.split("/")
                    if len(parts) < 3:
                        continue
                    # Skip config.json on restore (security: may have redacted keys)
                    if parts[-1] == "config.json" and len(parts) == 2:
                        continue

                    if parts[1] == "agents" and len(parts) >= 3:
                        agent_name = parts[2]
                        rel_path = "/".join(parts[3:])
                        dest = os.path.join(agents_dir, agent_name, rel_path)

                        if strategy == "merge" and os.path.exists(dest):
                            continue  # Skip existing files in merge mode

                        os.makedirs(os.path.dirname(dest), exist_ok=True)
                        f = tar.extractfile(member)
                        if f:
                            with open(dest, "wb") as out:
                                out.write(f.read())
                            imported["files"] += 1
                            if rel_path.endswith(".md"):
                                imported["memories"] += 1
                            if agent_name not in imported["agents"]:
                                imported["agents"].append(agent_name)

                    elif parts[1] == "databases" and len(parts) >= 3:
                        db_name = parts[2]
                        if strategy == "merge":
                            continue  # Don't overwrite databases in merge mode
                        dest = os.path.join(agents_dir, "main", db_name)
                        f = tar.extractfile(member)
                        if f:
                            with open(dest, "wb") as out:
                                out.write(f.read())
                            imported["files"] += 1

            self._send_json({
                "restored": True,
                "strategy": strategy,
                "imported": imported,
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._send_json({"error": str(e)}, 500)

    # --- Nodes API handlers ---

    def _handle_workers_list(self):
        """GET /v1/workers — list workers, optionally filtered by session_id."""
        from execution import get_worker_registry
        qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        session_id = qs.get("session_id", [None])[0]
        registry = get_worker_registry()
        if session_id:
            workers = registry.list_session(session_id)
        else:
            workers = list(registry._workers.values())
        self._send_json({"workers": [registry.to_status_dict(w) for w in workers]})

    def _handle_workers_recent(self):
        """GET /v1/workers/recent — all workers across sessions (admin view)."""
        from execution import get_worker_registry
        qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        limit = int(qs.get("limit", [50])[0])
        registry = get_worker_registry()
        with registry._lock:
            all_workers = list(registry._workers.values())
        all_workers.sort(key=lambda w: w.started_at or 0, reverse=True)
        all_workers = all_workers[:limit]
        result = []
        for w in all_workers:
            d = registry.to_status_dict(w)
            d["session_id"] = w.session_id
            d["agent_id"] = w.agent_id
            d["duration"] = w.duration
            result.append(d)
        self._send_json({"workers": result, "total": len(registry._workers)})

    def _handle_worker_answer(self, path):
        """POST /v1/workers/{id}/answer — deliver answer to a worker question."""
        from execution import get_worker_registry
        parts = path.split("/")
        worker_id = parts[3] if len(parts) >= 5 else ""
        body = self._read_json_body()
        if not body:
            self._send_json({"error": "Missing body"}, 400)
            return
        answer = body.get("answer", "")
        if not answer:
            self._send_json({"error": "Missing 'answer' field"}, 400)
            return
        ok = get_worker_registry().answer(worker_id, answer)
        if not ok:
            self._send_json({"error": f"Worker '{worker_id}' not waiting for answer"}, 400)
            return
        self._send_json({"ok": True, "worker_id": worker_id})

    def _handle_nodes_list(self):
        """GET /v1/nodes — list all nodes with status."""
        nodes = []
        with _node_lock:
            for token, info in _node_registry.items():
                cfg = info.get("config", {})
                nodes.append({
                    "name": info["name"],
                    "description": cfg.get("description", ""),
                    "token": token,
                    "status": info["status"],
                    "paused": cfg.get("paused", False),
                    "hostname": info.get("hostname", ""),
                    "os": info.get("os", ""),
                    "tags": cfg.get("tags", []),
                    "allowed_tools": cfg.get("allowed_tools", []),
                    "max_concurrent": cfg.get("max_concurrent", 5),
                    "command_timeout": cfg.get("command_timeout", 300),
                    "last_heartbeat": info.get("last_heartbeat"),
                    "cpu_percent": info.get("cpu_percent"),
                    "mem_used_gb": info.get("mem_used_gb"),
                    "mem_total_gb": info.get("mem_total_gb"),
                    "disk_free_gb": info.get("disk_free_gb"),
                    "uptime_seconds": info.get("uptime_seconds"),
                    "active_commands": info.get("active_commands", 0),
                    "total_commands": info.get("total_commands", 0),
                    "connected_since": info.get("connected_since"),
                })
        self._send_json({"nodes": nodes})

    def _handle_node_poll(self):
        """GET /v1/nodes/poll?token=X — node polls for pending commands."""
        qs = self.path.split("?", 1)[1] if "?" in self.path else ""
        params = dict(p.split("=", 1) for p in qs.split("&") if "=" in p)
        token = params.get("token", "")

        with _node_lock:
            info = _node_registry.get(token)
            if not info:
                self._send_json({"error": "Invalid token"}, 401)
                return

            import urllib.parse
            info["status"] = "connected"
            info["last_heartbeat"] = time.time()
            info["hostname"] = urllib.parse.unquote(params.get("hostname", ""))
            info["os"] = urllib.parse.unquote(params.get("os", ""))
            try:
                info["cpu_percent"] = float(params.get("cpu_percent", 0))
                info["mem_used_gb"] = float(params.get("mem_used_gb", 0))
                info["mem_total_gb"] = float(params.get("mem_total_gb", 0))
                info["disk_free_gb"] = float(params.get("disk_free_gb", 0))
                info["uptime_seconds"] = int(params.get("uptime_seconds", 0))
                info["active_commands"] = int(params.get("active_commands", 0))
                info["total_commands"] = int(params.get("total_commands", 0))
            except (ValueError, TypeError):
                pass
            if not info.get("connected_since"):
                info["connected_since"] = time.time()

            if info.get("config", {}).get("paused"):
                self._send_json({"error": "Node is paused"}, 403)
                return

            pending = info.get("pending_commands", [])
            if pending:
                cmd = pending.pop(0)
                self._send_json({"command": cmd})
                return

        # Long-poll: wait up to 30s for a command
        deadline = time.time() + 30
        while time.time() < deadline:
            time.sleep(2)
            with _node_lock:
                info = _node_registry.get(token)
                if not info:
                    break
                pending = info.get("pending_commands", [])
                if pending:
                    cmd = pending.pop(0)
                    self._send_json({"command": cmd})
                    return

        self._send_json({"command": None})

    def _handle_node_result(self):
        """POST /v1/nodes/result — receive command result from node."""
        body = self._read_json()
        token = body.get("token", "")
        command_id = body.get("command_id", "")
        result = body.get("result", {})

        with _node_lock:
            if token not in _node_registry:
                self._send_json({"error": "Invalid token"}, 401)
                return
            entry = _node_commands.get(command_id)
            if entry:
                entry["result"] = result
                entry["result_event"].set()

        self._send_json({"status": "ok"})

    def _handle_nodes_action(self):
        """POST /v1/nodes — add/remove/pause/resume/update a node."""
        body = self._read_json()
        action = body.get("action", "")

        if action == "add":
            name = body.get("name", "")
            if not name:
                self._send_json({"error": "Missing name"}, 400)
                return
            import secrets
            token = f"nd_{secrets.token_hex(16)}"
            cfg = {
                "token": token,
                "description": body.get("description", ""),
                "allowed_tools": body.get("allowed_tools", ["execute_command", "read_file", "write_file", "list_directory"]),
                "tags": body.get("tags", []),
                "max_concurrent": body.get("max_concurrent", 5),
                "command_timeout": body.get("command_timeout", 300),
                "paused": False,
            }
            nodes_cfg = _load_node_config()
            nodes_cfg[name] = cfg
            _save_node_config(nodes_cfg)
            with _node_lock:
                _node_registry[token] = {
                    "name": name, "config": cfg, "status": "disconnected",
                    "last_heartbeat": None, "hostname": "", "os": "",
                    "cpu_percent": None, "mem_used_gb": None, "mem_total_gb": None,
                    "disk_free_gb": None, "uptime_seconds": None,
                    "active_commands": 0, "total_commands": 0,
                    "connected_since": None, "pending_commands": [],
                }
            port = server_config.get("port", 8420)
            install_cmd = f"python3 node.py --install --server http://SERVER_IP:{port} --token {token} --name {name}"
            self._send_json({"ok": True, "token": token, "install_command": install_cmd})

        elif action == "remove":
            name = body.get("name", "")
            nodes_cfg = _load_node_config()
            removed_token = None
            for n, cfg in nodes_cfg.items():
                if n == name:
                    removed_token = cfg.get("token")
                    break
            if name in nodes_cfg:
                del nodes_cfg[name]
                _save_node_config(nodes_cfg)
            if removed_token:
                with _node_lock:
                    _node_registry.pop(removed_token, None)
            self._send_json({"ok": True})

        elif action in ("pause", "resume"):
            name = body.get("name", "")
            paused = action == "pause"
            nodes_cfg = _load_node_config()
            if name in nodes_cfg:
                nodes_cfg[name]["paused"] = paused
                _save_node_config(nodes_cfg)
                with _node_lock:
                    for token, info in _node_registry.items():
                        if info["name"] == name:
                            info["config"]["paused"] = paused
                            break
            self._send_json({"ok": True, "paused": paused})

        elif action == "update":
            name = body.get("name", "")
            nodes_cfg = _load_node_config()
            if name in nodes_cfg:
                for key in ("description", "allowed_tools", "tags", "max_concurrent", "command_timeout"):
                    if key in body:
                        nodes_cfg[name][key] = body[key]
                _save_node_config(nodes_cfg)
                with _node_lock:
                    for token, info in _node_registry.items():
                        if info["name"] == name:
                            info["config"].update(nodes_cfg[name])
                            break
            self._send_json({"ok": True})
        else:
            self._send_json({"error": f"Unknown action: {action}"}, 400)

    def _handle_node_execute(self):
        """POST /v1/nodes/execute — submit command to a node (internal)."""
        body = self._read_json()
        node = body.get("node", "")
        tool = body.get("tool", "")
        params = body.get("params", {})
        if not node or not tool:
            self._send_json({"error": "Missing node or tool"}, 400)
            return
        result = _node_submit_command(node, tool, params)
        self._send_json(result)

    # --- Channels API handlers ---

    def _handle_channels_list(self):
        """GET /v1/channels — list all messaging channels."""
        mgr = _adapters_mod.channel_manager
        if not mgr:
            self._send_json({"channels": []})
            return
        self._send_json({"channels": mgr.status()})

    def _handle_channels_action(self):
        """POST /v1/channels — create/remove/update a channel."""
        body = self._read_json()
        action = body.get("action", "create")
        mgr = _adapters_mod.channel_manager
        if not mgr:
            self._send_json({"error": "Channel manager not initialized"}, 500)
            return

        if action == "create":
            ch_id = body.get("id", body.get("name", ""))
            if not ch_id:
                self._send_json({"error": "Missing channel id"}, 400)
                return
            try:
                channel = mgr.create_channel(ch_id, body)
                if body.get("enabled", True):
                    channel.start()
                self._save_channel_config(mgr)
                self._send_json({"ok": True, "channel": channel.status()})
            except Exception as e:
                self._send_json({"error": str(e)}, 400)

        elif action == "remove":
            ch_id = body.get("id", "")
            mgr.remove_channel(ch_id)
            self._save_channel_config(mgr)
            self._send_json({"ok": True})

        elif action == "update":
            ch_id = body.get("id", "")
            ch = mgr.channels.get(ch_id)
            if ch:
                for key in ("name", "agent_routing", "allowed_users", "default_model", "enabled"):
                    if key in body:
                        ch.config[key] = body[key]
                self._save_channel_config(mgr)
                self._send_json({"ok": True, "channel": ch.status()})
            else:
                self._send_json({"error": "Channel not found"}, 404)
        else:
            self._send_json({"error": f"Unknown action: {action}"}, 400)

    def _handle_channel_lifecycle(self, path: str, action: str):
        """POST /v1/channels/:id/start|stop|restart."""
        parts = path.split("/")
        ch_id = parts[3] if len(parts) > 3 else ""
        mgr = _adapters_mod.channel_manager
        if not mgr:
            self._send_json({"error": "Channel manager not initialized"}, 500)
            return
        ch = mgr.channels.get(ch_id)
        if not ch:
            self._send_json({"error": "Channel not found"}, 404)
            return
        if action == "stop":
            ch.stop()
        elif action == "start":
            ch.start()
        elif action == "restart":
            ch.stop()
            ch.start()
        self._send_json({"ok": True, "channel": ch.status()})

    def _save_channel_config(self, mgr):
        """Persist channel config to config.json."""
        config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.json")
        try:
            config = {}
            if os.path.exists(config_path):
                with open(config_path) as f:
                    config = json.load(f)
            channels = []
            for ch_id, ch in mgr.channels.items():
                channels.append({"id": ch_id, **ch.config})
            config["channels"] = channels
            with open(config_path, "w") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Failed to save channel config: {e}", flush=True)

    def _validate_file_path(self, file_path):
        """Validate that a file path is within allowed directories. Returns resolved path or None.
        Allows the cctest tree, agents/, cwd, AND any path under a project's
        input_folders[]. Project input folders are the user-explicit set of
        paths the project has been told to mine, so it's safe to serve files
        from there back to the same authenticated user via /v1/files/download
        — citations from `mempalace_query` / `mempalace_kg_*` resolve to
        absolute paths under those roots."""
        if not file_path:
            return None
        file_path = os.path.expanduser(file_path)
        resolved = os.path.realpath(file_path)
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        agents_dir = os.path.join(base, "agents")
        cwd = os.getcwd()
        # /tmp/brain-attachments/<session_id>/ — disk-saved chat attachments
        # (see chat.py: attach_dir). macOS resolves /tmp → /private/tmp, so
        # realpath both ends to make the prefix-match work.
        attach_root = os.path.realpath("/tmp/brain-attachments")
        allowed = [base, agents_dir, cwd, attach_root]
        if any(resolved.startswith(d) for d in allowed):
            return resolved
        # Project input folders — symlink-resolved, deduped. Any project
        # this user can see contributes its input_folder roots.
        try:
            auth_user = getattr(self, '_auth_user', None) or _auth_mod.SYNTHETIC_ADMIN
            uid = auth_user.get("id", "")
            # The synthetic-admin sentinel ('__system__') means "see everything"
            # — but list_projects(user_id='__system__') visibility-filters it down
            # to almost nothing. Normalise it to None so the admin/auth-disabled
            # path sees ALL projects (incl. code-mode working_dirs), matching the
            # rest of the admin surface. (This was why writing into a code-mode
            # working_dir failed under auth-disabled / synthetic-admin requests.)
            if uid == "__system__":
                uid = ""
            team_ids = []
            if uid and uid != "__system__":
                try:
                    team_ids = [t["id"] for t in _auth_mod.AuthDB.get_user_teams(uid)]
                except Exception:
                    pass
            for agent_id in os.listdir(engine.AGENTS_DIR):
                # Skip non-agent entries (.DS_Store, auth.db, .trash, …). Without
                # this guard a single bad entry's list_projects() can raise and —
                # since the whole loop is in one try/except — abort validation of
                # an otherwise-allowed project path (the bug that broke writing
                # into a code-mode working_dir).
                if agent_id.startswith(".") or not os.path.isdir(
                        os.path.join(engine.AGENTS_DIR, agent_id)):
                    continue
                try:
                    projects = engine.ProjectManager.list_projects(
                        agent_id, user_id=uid or None, user_team_ids=team_ids,
                    )
                except Exception:
                    continue
                for proj in projects:
                    for folder in (proj.get("input_folders") or []):
                        p = (folder or {}).get("path", "").strip()
                        if not p:
                            continue
                        root = os.path.realpath(os.path.expanduser(p))
                        if resolved.startswith(root):
                            return resolved
                    # Code-mode projects: the working_dir is the user-explicit
                    # root the project operates in → serve files from there too.
                    if proj.get("code_mode"):
                        wd = (proj.get("working_dir") or "").strip()
                        if wd:
                            wdroot = os.path.realpath(os.path.expanduser(wd))
                            if resolved == wdroot or resolved.startswith(wdroot + os.sep):
                                return resolved
        except Exception:
            pass
        return None

    def _resolve_project_basename(self, raw_path):
        """Best-effort lookup: given a bare basename or relative path
        (the shape MemPalace drawers carry as `source_file`), find a
        matching file under any project input_folders[] the authenticated
        user can see. Strips a trailing `.md` companion suffix
        automatically. Returns the absolute path of the first match, or
        None. First match wins; if multiple projects have a same-named
        file the user gets one of them — better than nothing, and the
        right-panel card already shows the basename so the user can tell.
        """
        if not raw_path or "/" in raw_path[:1]:
            return None
        # Normalise the lookup name: a) strip trailing .md if it sits on
        # top of a known binary extension; b) keep the raw name as-is
        # otherwise (e.g. .md sources are first-class).
        candidates = [raw_path]
        m = re.match(r"^(.+\.(pdf|docx|pptx|xlsx|xlsm|eml|msg))\.md$", raw_path, re.IGNORECASE)
        if m:
            candidates.insert(0, m.group(1))  # try original binary first
        try:
            auth_user = getattr(self, '_auth_user', None) or _auth_mod.SYNTHETIC_ADMIN
            uid = auth_user.get("id", "")
            # The synthetic-admin sentinel ('__system__') means "see everything"
            # — but list_projects(user_id='__system__') visibility-filters it down
            # to almost nothing. Normalise it to None so the admin/auth-disabled
            # path sees ALL projects (incl. code-mode working_dirs), matching the
            # rest of the admin surface. (This was why writing into a code-mode
            # working_dir failed under auth-disabled / synthetic-admin requests.)
            if uid == "__system__":
                uid = ""
            team_ids = []
            if uid and uid != "__system__":
                try:
                    team_ids = [t["id"] for t in _auth_mod.AuthDB.get_user_teams(uid)]
                except Exception:
                    pass
            roots = []
            for agent_id in os.listdir(engine.AGENTS_DIR):
                # Skip non-agent entries (.DS_Store, auth.db, .trash, …). Without
                # this guard a single bad entry's list_projects() can raise and —
                # since the whole loop is in one try/except — abort validation of
                # an otherwise-allowed project path (the bug that broke writing
                # into a code-mode working_dir).
                if agent_id.startswith(".") or not os.path.isdir(
                        os.path.join(engine.AGENTS_DIR, agent_id)):
                    continue
                try:
                    projects = engine.ProjectManager.list_projects(
                        agent_id, user_id=uid or None, user_team_ids=team_ids,
                    )
                except Exception:
                    continue
                for proj in projects:
                    for folder in (proj.get("input_folders") or []):
                        p = (folder or {}).get("path", "").strip()
                        if p:
                            roots.append(os.path.realpath(os.path.expanduser(p)))
            roots = list(dict.fromkeys(roots))  # dedupe, keep order
            for root in roots:
                if not os.path.isdir(root):
                    continue
                for cand in candidates:
                    base = os.path.basename(cand)
                    # First: cheap top-level glob
                    direct = os.path.join(root, cand)
                    if os.path.isfile(direct):
                        return os.path.realpath(direct)
                    # Then: recursive basename walk (capped to avoid runaway
                    # scans on misconfigured roots).
                    scanned = 0
                    for dirpath, _dirs, files in os.walk(root):
                        if base in files:
                            return os.path.realpath(os.path.join(dirpath, base))
                        scanned += 1
                        if scanned > 5000:  # safety
                            break
        except Exception:
            return None
        return None

    def _handle_file_download(self):
        """GET /v1/files/download?path=<absolute_path> — serve a file for download."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        file_path = qs.get("path", [""])[0]
        resolved = self._validate_file_path(file_path)
        # If the validator rejected (None) OR returned a path that doesn't
        # exist on disk (because the input was relative + got resolved
        # against the server's CWD into the cctest tree), try
        # project-input-folder basename resolution. MemPalace drawers
        # store `source_file` as a relative path (sometimes the bare
        # basename of a binary that's deeper in a project input folder).
        if not resolved or not os.path.isfile(resolved):
            looked_up = self._resolve_project_basename(file_path)
            if looked_up and os.path.isfile(looked_up):
                resolved = looked_up
        if not resolved:
            self._send_json({"error": "Invalid or disallowed file path"}, 403)
            return
        if not os.path.isfile(resolved):
            self._send_json({"error": "File not found"}, 404)
            return
        ext = resolved.rsplit(".", 1)[-1].lower() if "." in resolved else ""
        content_types = {
            "md": "text/markdown", "txt": "text/plain", "py": "text/x-python",
            "json": "application/json", "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "html": "text/html", "csv": "text/csv",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "js": "application/javascript", "ts": "text/typescript",
            "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "gif": "image/gif", "svg": "image/svg+xml",
        }
        ct = content_types.get(ext, "application/octet-stream")
        filename = os.path.basename(resolved)
        # Render PDFs and images inline so the browser opens them in a new
        # tab instead of force-downloading. Office-binary types stay
        # `attachment` because browsers can't render them — they'd just
        # download with a confusing blob:// URL otherwise.
        inline_exts = {"pdf", "png", "jpg", "jpeg", "gif", "svg",
                       "txt", "md", "html", "json", "csv"}
        disposition = "inline" if ext in inline_exts else "attachment"
        try:
            with open(resolved, "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type", ct)
            self.send_header("Content-Length", len(data))
            # Quote-escape the filename so non-ASCII (German umlauts) and
            # spaces don't break the header. RFC 5987 filename* takes a
            # UTF-8-encoded value.
            from urllib.parse import quote as _urlq
            self.send_header(
                "Content-Disposition",
                f"{disposition}; filename*=UTF-8''{_urlq(filename)}",
            )
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    _ZIP_SKIP_DIRS = {".git", "__pycache__", "node_modules", ".venv", "venv",
                      ".cbm-cache", ".brain-extracted", ".trash", "dist", "build",
                      ".idea", ".vscode"}

    def _handle_file_zip(self):
        """GET /v1/files/zip?path=<absolute dir> — zip a directory tree (skipping
        heavy/derived dirs) and stream it as application/zip. Path validated like
        file download, so only allowed roots (incl. code-mode working dirs) zip."""
        from urllib.parse import urlparse, parse_qs
        import io
        import zipfile
        qs = parse_qs(urlparse(self.path).query)
        dir_path = qs.get("path", [""])[0]
        resolved = self._validate_file_path(dir_path)
        if not resolved:
            self._send_json({"error": "Invalid or disallowed path"}, 403)
            return
        if not os.path.isdir(resolved):
            self._send_json({"error": "Not a directory"}, 404)
            return
        try:
            buf = io.BytesIO()
            root_name = os.path.basename(resolved.rstrip(os.sep)) or "project"
            total = 0
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for dirpath, dirnames, filenames in os.walk(resolved):
                    dirnames[:] = [d for d in dirnames
                                   if d not in self._ZIP_SKIP_DIRS and not d.startswith(".")]
                    for fn in filenames:
                        fp = os.path.join(dirpath, fn)
                        try:
                            if os.path.islink(fp) or os.path.getsize(fp) > 20 * 1024 * 1024:
                                continue  # skip symlinks + >20MB blobs
                        except OSError:
                            continue
                        arc = os.path.join(root_name, os.path.relpath(fp, resolved))
                        try:
                            zf.write(fp, arc)
                            total += 1
                        except OSError:
                            continue
            data = buf.getvalue()
            from urllib.parse import quote as _urlq
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Length", len(data))
            self.send_header("Content-Disposition",
                             f"attachment; filename*=UTF-8''{_urlq(root_name)}.zip")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_save(self):
        """POST /v1/files/save {path, content} — write a text file (create or
        overwrite). Same path validation as preview/download (allowed roots incl.
        code-mode working_dir). For a NEW file, the parent dir must already exist
        and be inside an allowed root. Returns {ok, path, size}."""
        body = self._read_json() or {}
        raw_path = (body.get("path") or "").strip()
        content = body.get("content")
        if not raw_path or content is None:
            self._send_json({"error": "path und content erforderlich"}, 400)
            return
        if not isinstance(content, str):
            self._send_json({"error": "content muss Text sein"}, 400)
            return
        # Validate the path. For a new file the target may not exist yet, so we
        # validate its PARENT dir (which must be a real, allowed directory) and
        # then re-validate the full path lands under the same allowed root.
        resolved = self._validate_file_path(raw_path)
        if not resolved:
            parent = os.path.dirname(os.path.realpath(os.path.expanduser(raw_path)))
            pv = self._validate_file_path(parent)
            if pv and os.path.isdir(pv):
                resolved = os.path.join(pv, os.path.basename(raw_path))
        if not resolved:
            self._send_json({"error": "Ungültiger oder nicht erlaubter Pfad"}, 403)
            return
        if len(content.encode("utf-8")) > 10 * 1024 * 1024:
            self._send_json({"error": "Datei zu groß (>10MB)"}, 400)
            return
        try:
            os.makedirs(os.path.dirname(resolved), exist_ok=True)
            with open(resolved, "w", encoding="utf-8") as f:
                f.write(content)
            self._send_json({"ok": True, "path": resolved,
                             "size": os.path.getsize(resolved)})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_rename(self):
        """POST /v1/files/rename {path, to} — rename OR move a file/folder. Both
        the source and the destination (and the destination's parent) must
        validate to the same allowed roots as save/preview. `to` may be an
        absolute path or a bare new name (resolved against the source's dir).
        Refuses to overwrite an existing target. Returns {ok, path}."""
        import shutil
        body = self._read_json() or {}
        raw_src = (body.get("path") or "").strip()
        raw_to = (body.get("to") or "").strip()
        if not raw_src or not raw_to:
            self._send_json({"error": "path und to erforderlich"}, 400)
            return
        src = self._validate_file_path(raw_src)
        if not src or not os.path.exists(src):
            self._send_json({"error": "Quelle ungültig oder nicht vorhanden"}, 403)
            return
        # Bare name (no separator) → keep it in the source's directory.
        if "/" not in raw_to and os.sep not in raw_to:
            dst_raw = os.path.join(os.path.dirname(src), raw_to)
        else:
            dst_raw = os.path.expanduser(raw_to)
        # Validate the destination via its parent (the target itself doesn't
        # exist yet), mirroring the save handler.
        dst = self._validate_file_path(dst_raw)
        if not dst:
            parent = os.path.dirname(os.path.realpath(dst_raw))
            pv = self._validate_file_path(parent)
            if pv and os.path.isdir(pv):
                dst = os.path.join(pv, os.path.basename(dst_raw))
        if not dst:
            self._send_json({"error": "Zielpfad ungültig oder nicht erlaubt"}, 403)
            return
        if os.path.exists(dst):
            self._send_json({"error": "Ziel existiert bereits"}, 409)
            return
        try:
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            shutil.move(src, dst)
            self._send_json({"ok": True, "path": dst})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_delete(self):
        """POST /v1/files/delete {path} — SOFT-delete a file/folder by moving it
        into a `.brain-trash/` folder at the root of its allowed tree (recoverable
        — never a hard rm, per the no-permanent-delete rule). Returns
        {ok, trashed}."""
        import shutil
        import time as _t
        body = self._read_json() or {}
        src = self._validate_file_path((body.get("path") or "").strip())
        if not src or not os.path.exists(src):
            self._send_json({"error": "Pfad ungültig oder nicht vorhanden"}, 403)
            return
        # Trash root: collect at the project working_dir the file sits under (so
        # all deletes for a project land in one .brain-trash), else fall back to
        # the file's own directory. We resolve the working_dir by re-validating
        # each ancestor until validation stops accepting it — the last accepted
        # ancestor is the allowed root. This keeps the trash INSIDE an allowed,
        # writable root (never walks up to '/').
        trash_dir = os.path.join(os.path.dirname(src), ".brain-trash")
        try:
            probe = os.path.dirname(src)
            last_ok = probe
            for _ in range(60):
                parent = os.path.dirname(probe)
                if parent == probe:
                    break
                # Stop at a project root marker (working_dir often == a git repo)
                if os.path.isdir(os.path.join(probe, ".git")):
                    last_ok = probe
                    break
                # Keep climbing only while the parent is still an allowed root.
                if self._validate_file_path(parent):
                    last_ok = parent
                    probe = parent
                else:
                    break
            trash_dir = os.path.join(last_ok, ".brain-trash")
        except Exception:
            pass
        try:
            os.makedirs(trash_dir, exist_ok=True)
            # Timestamp-prefix to avoid collisions on repeated deletes of a name.
            stamp = _t.strftime("%Y%m%d-%H%M%S")
            dest = os.path.join(trash_dir, f"{stamp}__{os.path.basename(src)}")
            shutil.move(src, dest)
            self._send_json({"ok": True, "trashed": dest})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_mkdir(self):
        """POST /v1/files/mkdir {path} — create a new folder. The parent must
        validate to an allowed root. Returns {ok, path}."""
        body = self._read_json() or {}
        raw = (body.get("path") or "").strip()
        if not raw:
            self._send_json({"error": "path erforderlich"}, 400)
            return
        target_raw = os.path.realpath(os.path.expanduser(raw))
        parent = os.path.dirname(target_raw)
        pv = self._validate_file_path(parent)
        if not pv or not os.path.isdir(pv):
            self._send_json({"error": "Übergeordneter Ordner ungültig oder nicht erlaubt"}, 403)
            return
        resolved = os.path.join(pv, os.path.basename(target_raw))
        if os.path.exists(resolved):
            self._send_json({"error": "Ordner existiert bereits"}, 409)
            return
        try:
            os.makedirs(resolved, exist_ok=False)
            self._send_json({"ok": True, "path": resolved})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_open_external(self):
        """POST /v1/files/open-external {path} — open a file in the host's default
        external application (Word/Excel/PowerPoint/Acrobat/…). Same path
        validation as preview/save (allowed roots incl. code-mode working_dir).
        Uses the OS opener (`open` on macOS, `xdg-open` on Linux, `os.startfile`
        on Windows). The file is launched DETACHED; no shell, args are a fixed
        list (the path is the only variable) so there's no injection surface.
        NOTE: this opens on the SERVER host — meaningful for a local single-user
        deploy (the daemon runs on the user's machine)."""
        import sys
        import subprocess
        body = self._read_json() or {}
        resolved = self._validate_file_path((body.get("path") or "").strip())
        if not resolved or not os.path.isfile(resolved):
            self._send_json({"error": "Ungültiger oder nicht erlaubter Pfad"}, 403)
            return
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", resolved])
            elif sys.platform.startswith("win"):
                os.startfile(resolved)  # noqa: S606 — Windows opener
            else:
                subprocess.Popen(["xdg-open", resolved])
            self._send_json({"ok": True, "path": resolved})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_preview(self):
        """GET /v1/files/preview?path=<absolute_path>&lines=100 — return file content for preview."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        file_path = qs.get("path", [""])[0]
        max_lines = int(qs.get("lines", ["100"])[0])
        resolved = self._validate_file_path(file_path)
        if not resolved:
            self._send_json({"error": "Invalid or disallowed file path"}, 403)
            return
        if not os.path.isfile(resolved):
            self._send_json({"error": "File not found"}, 404)
            return
        try:
            _stt = os.stat(resolved)
            size = _stt.st_size
            mtime = int(_stt.st_mtime)
            name = os.path.basename(resolved)
            ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
            image_exts = {"jpg", "jpeg", "png", "gif", "webp", "svg", "bmp", "ico"}
            office_exts = {"pdf", "docx", "xlsx", "pptx", "csv"}
            if ext in image_exts:
                self._send_json({
                    "path": resolved, "name": name, "size": size, "mtime": mtime,
                    "type": "image", "ext": ext,
                })
                return
            if ext in office_exts:
                try:
                    if ext == "pdf":
                        content = engine.DocumentParser.parse_pdf(resolved)
                    elif ext == "docx":
                        content = engine.DocumentParser.parse_docx(resolved)
                    elif ext in ("xlsx", "xls"):
                        content = engine.DocumentParser.parse_xlsx(resolved)
                    elif ext == "pptx":
                        content = engine.DocumentParser.parse_pptx(resolved)
                    elif ext == "csv":
                        with open(resolved, "r", errors="replace") as f:
                            content = f.read(50 * 1024)
                    else:
                        content = ""
                    all_lines = content.splitlines()
                    truncated = len(all_lines) > 200
                    self._send_json({
                        "path": resolved, "name": name, "size": size, "mtime": mtime,
                        "type": "document", "ext": ext,
                        "content": "\n".join(all_lines[:200]), "truncated": truncated,
                    })
                except Exception as e:
                    self._send_json({"error": f"Could not parse {ext.upper()}: {e}"}, 500)
                return
            # Plain text / code
            max_bytes = 50 * 1024
            with open(resolved, "r", errors="replace") as f:
                lines = []
                total_bytes = 0
                for i, line in enumerate(f):
                    if i >= max_lines or total_bytes >= max_bytes:
                        truncated = True
                        break
                    lines.append(line)
                    total_bytes += len(line.encode("utf-8"))
                else:
                    truncated = False
            self._send_json({
                "path": resolved, "name": name, "size": size, "mtime": mtime,
                "type": "text",
                "content": "".join(lines), "truncated": truncated,
            })
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_stat(self):
        """GET /v1/files/stat?path=<abs> → {mtime,size} only. Cheap poll target
        for the editor auto-reload (no content read)."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        resolved = self._validate_file_path(qs.get("path", [""])[0])
        if not resolved:
            self._send_json({"error": "Invalid or disallowed file path"}, 403)
            return
        if not os.path.isfile(resolved):
            self._send_json({"error": "File not found"}, 404)
            return
        try:
            st = os.stat(resolved)
            self._send_json({"mtime": int(st.st_mtime), "size": int(st.st_size)})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_file_xlsx_grid(self):
        """GET /v1/files/xlsx-grid?path=<abs>&sheet=&rows=500 — a spreadsheet
        as STRUCTURED grid JSON for the UI table preview (bottom-panel editor
        + artifacts fullview): {sheets: [{name, header, rows, total_rows,
        truncated}]}. Reuses the xlsx-toolset grid loader (header detection,
        placeholder trim, multi-table split, merged-header composition) so the
        preview shows exactly what the agent's xlsx tools see — unlike
        /v1/files/preview, which flattens to markdown text."""
        from urllib.parse import urlparse, parse_qs
        from engine.tools.xlsx_tools import _load_grids, _to_sql_value
        qs = parse_qs(urlparse(self.path).query)
        resolved = self._validate_file_path(qs.get("path", [""])[0])
        max_rows = min(2000, int(qs.get("rows", ["500"])[0] or 500))
        sheet = (qs.get("sheet", [""])[0] or None)
        if not resolved:
            self._send_json({"error": "Invalid or disallowed file path"}, 403)
            return
        if not os.path.isfile(resolved):
            self._send_json({"error": "File not found"}, 404)
            return
        ext = resolved.rsplit(".", 1)[-1].lower() if "." in resolved else ""
        if ext not in ("xlsx", "xlsm", "csv", "tsv"):
            self._send_json({"error": f"Not a spreadsheet: .{ext}"}, 400)
            return
        try:
            if os.path.getsize(resolved) > 30 * 1024 * 1024:
                self._send_json({"error": "Datei zu groß für die Vorschau (>30 MB)"}, 413)
                return
            def _js(v):
                v = _to_sql_value(v)
                return v if isinstance(v, (int, float, str)) or v is None else str(v)
            sheets = []
            for g in _load_grids(resolved, sheet=sheet):
                header = [str(h) for h in g["header"][:100]]
                rows = [[_js(v) for v in r[:100]]
                        for r in g["rows"][:max_rows]]
                sheets.append({
                    "name": g["name"], "header": header, "rows": rows,
                    "total_rows": len(g["rows"]),
                    "truncated": len(g["rows"]) > max_rows
                                 or len(g["header"]) > 100,
                    # v3 editable grid: the REAL worksheet title (block grids
                    # are named <sheet>_2 …) + per-row absolute sheet rows so
                    # the client can address cells for /v1/files/xlsx-cell.
                    "sheet_title": g.get("sheet_title", g["name"]),
                    "row_nums": (g.get("row_nums") or [])[:max_rows],
                })
            st = os.stat(resolved)
            self._send_json({"path": resolved,
                             "name": os.path.basename(resolved),
                             "size": int(st.st_size),
                             "mtime": int(st.st_mtime),
                             "sheets": sheets})
        except Exception as e:
            self._send_json({"error": f"Grid-Parse fehlgeschlagen: {e}"}, 500)

    def _handle_file_xlsx_cell(self):
        """POST /v1/files/xlsx-cell {path, sheet, row, col, value, mtime?} —
        write ONE cell of an existing workbook (the UI grid's inline edit,
        v9.264.0). row/col are 1-based absolute sheet coordinates (the grid
        endpoint returns row_nums for the mapping). `mtime` (as returned by
        xlsx-grid) enables a conflict check: 409 when the file changed since
        the grid was loaded. Formatting elsewhere is untouched (openpyxl
        in-place edit, keep_vba for .xlsm). No artifact-version churn — same
        policy as /v1/files/save."""
        body = self._read_json() or {}
        resolved = self._validate_file_path((body.get("path") or "").strip())
        if not resolved or not os.path.isfile(resolved):
            self._send_json({"error": "Ungültiger oder nicht erlaubter Pfad"}, 403)
            return
        if not resolved.lower().endswith((".xlsx", ".xlsm")):
            self._send_json({"error": "Nur .xlsx/.xlsm sind editierbar"}, 400)
            return
        sheet = body.get("sheet") or ""
        try:
            row = int(body.get("row"))
            col = int(body.get("col"))
        except (TypeError, ValueError):
            self._send_json({"error": "row/col (1-basiert) erforderlich"}, 400)
            return
        if row < 1 or col < 1:
            self._send_json({"error": "row/col müssen ≥ 1 sein"}, 400)
            return
        known_mtime = body.get("mtime")
        if known_mtime and int(os.stat(resolved).st_mtime) > int(known_mtime):
            self._send_json({"error": "Datei wurde zwischenzeitlich geändert — "
                                      "Ansicht neu laden"}, 409)
            return
        raw_val = body.get("value")
        # Coercion mirrors the grid's typing: empty → None, numbers typed,
        # '=' prefix stays a formula string, everything else text.
        value = raw_val
        if isinstance(raw_val, str):
            t = raw_val.strip()
            if t == "":
                value = None
            elif not t.startswith("="):
                try:
                    value = int(t)
                except ValueError:
                    try:
                        value = float(t)
                    except ValueError:
                        value = raw_val
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                resolved, keep_vba=resolved.lower().endswith(".xlsm"))
            if sheet not in wb.sheetnames:
                self._send_json({"error": f"Blatt '{sheet}' nicht gefunden — "
                                          f"vorhanden: {wb.sheetnames}"}, 404)
                return
            wb[sheet].cell(row=row, column=col, value=value)
            wb.save(resolved)
            st = os.stat(resolved)
            self._send_json({"ok": True, "mtime": int(st.st_mtime),
                             "size": int(st.st_size)})
        except Exception as e:
            self._send_json({"error": f"Zelle konnte nicht geschrieben werden: {e}"}, 500)

    def _handle_file_xlsm_vba(self):
        """GET /v1/files/xlsm-vba?path=<abs> — VBA module sources of a
        macro-enabled workbook as {modules: [{name, code}]} for the
        bottom-panel VBA viewer (v9.265.0). READ-ONLY by design: editing
        vbaProject.bin in place isn't safely possible without Excel (MS-OVBA
        compression + compiled P-code + offset directory), so the UI offers
        per-module .bas export instead of a fake save. Macros are never
        executed."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        resolved = self._validate_file_path(qs.get("path", [""])[0])
        if not resolved:
            self._send_json({"error": "Invalid or disallowed file path"}, 403)
            return
        if not os.path.isfile(resolved):
            self._send_json({"error": "File not found"}, 404)
            return
        if not resolved.lower().endswith((".xlsm", ".xls", ".xlsb",
                                          ".docm", ".pptm")):
            self._send_json({"error": "Kein makrofähiges Office-Format"}, 400)
            return
        try:
            from engine.doc_convert import list_vba_modules
            self._send_json({"path": resolved,
                             "modules": list_vba_modules(resolved)})
        except Exception as e:
            self._send_json({"error": f"VBA-Extraktion fehlgeschlagen: {e}"}, 500)

    # ── Code Mode Endpoints ──

    def _handle_file_tree(self):
        """GET /v1/files/tree?path=<dir>&depth=2 — return directory tree for Code mode."""
        from urllib.parse import urlparse, parse_qs, unquote
        qs = parse_qs(urlparse(self.path).query)
        dir_path = unquote(qs.get("path", [""])[0])
        max_depth = int(qs.get("depth", ["2"])[0])
        # Empty path defaults to the user's home dir, so the folder picker
        # doesn't need to know where to start.
        if not dir_path:
            dir_path = os.path.expanduser("~")
        else:
            dir_path = os.path.expanduser(dir_path)
        if not os.path.isdir(dir_path):
            self._send_json({"error": "Invalid or missing directory path"}, 400)
            return

        IGNORE = {".git", "node_modules", "__pycache__", ".venv", "venv", ".tox",
                  ".mypy_cache", ".pytest_cache", ".DS_Store", ".claude", "dist", "build"}

        def _scan(base, depth=0):
            items = []
            try:
                entries = sorted(os.scandir(base), key=lambda e: (not e.is_dir(), e.name.lower()))
            except PermissionError:
                return items
            for entry in entries:
                if entry.name in IGNORE or entry.name.startswith("."):
                    continue
                node = {"name": entry.name, "path": entry.path}
                if entry.is_dir():
                    node["type"] = "dir"
                    if depth < max_depth:
                        node["children"] = _scan(entry.path, depth + 1)
                    else:
                        node["children"] = []
                        node["truncated"] = True
                else:
                    node["type"] = "file"
                    try:
                        node["size"] = entry.stat().st_size
                    except OSError:
                        node["size"] = 0
                items.append(node)
            return items

        tree = _scan(dir_path)
        self._send_json({"path": dir_path, "tree": tree})

    # ── Artifact Endpoints ──

    def _handle_artifacts_list(self):
        """GET /v1/artifacts?session_id=X — list artifacts for a session."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        session_id = qs.get("session_id", [""])[0]
        if not session_id:
            self._send_json({"error": "session_id required"}, 400)
            return
        artifacts = ChatDB.get_artifacts(session_id)
        self._send_json({"artifacts": artifacts})

    def _handle_artifacts_browse(self):
        """GET /v1/artifacts/browse?agent_id=X&limit=N&source=chat|scheduled
        — browse all artifacts across sessions, tagged by source so the UI
        can split the view. Scheduled-task artifacts are identified by
        session_id matching `sched-<run_id>` (set by the scheduler's
        synthetic session context)."""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        agent_id = qs.get("agent_id", [None])[0]
        limit = int(qs.get("limit", ["100"])[0])
        source_filter = qs.get("source", [None])[0]  # chat | scheduled | None
        artifacts = ChatDB.get_all_artifacts(agent_id=agent_id, limit=limit)

        # Enrich: source tag + schedule-run summary for scheduled artifacts.
        # Batch-resolve run rows so we don't hit the scheduler DB per-artifact.
        # Translation jobs use synthetic session ids `tr<14 hex>` minted by
        # handlers/translate.py — tag them as "translation" so the browse UI
        # can split them off from regular chat artifacts.
        import re as _re
        _TR_SID_RE = _re.compile(r"^tr[0-9a-f]{14}$")
        run_ids_needed = set()
        for a in artifacts:
            sid = a.get("session_id") or ""
            if sid.startswith("sched-"):
                a["source"] = "scheduled"
                try:
                    a["run_id"] = int(sid.split("-", 1)[1])
                    run_ids_needed.add(a["run_id"])
                except (ValueError, IndexError):
                    a["run_id"] = None
            elif _TR_SID_RE.match(sid):
                a["source"] = "translation"
                a["run_id"] = None
            else:
                a["source"] = "chat"
                a["run_id"] = None

        run_map: dict = {}
        if run_ids_needed and engine._scheduler:
            for rid in run_ids_needed:
                row = engine._scheduler.get_run(rid)
                if row:
                    run_map[rid] = {
                        "run_id": rid,
                        "schedule_name": row.get("schedule_name"),
                        "status": row.get("status"),
                        "started_at": row.get("started_at"),
                    }
        for a in artifacts:
            if a.get("run_id") in run_map:
                a["schedule_run"] = run_map[a["run_id"]]

        if source_filter in ("chat", "scheduled", "translation"):
            artifacts = [a for a in artifacts if a.get("source") == source_filter]

        # Fetch text preview for each text-based artifact; flag image artifacts
        # that have a precomputed thumbnail so the browse grid can request it
        # (the small thumbnail) instead of re-fetching the full image bytes.
        binary_types = {"image", "document"}
        for a in artifacts:
            if a.get("type") not in binary_types:
                preview = ChatDB.get_artifact_preview(a["id"], max_chars=300)
                a["preview"] = preview
            else:
                a["preview"] = None
            if a.get("type") == "image":
                a["has_thumbnail"] = ChatDB.has_artifact_thumbnail(a["id"], a.get("latest_version") or 1)
        self._send_json({"artifacts": artifacts})

    def _handle_artifact_thumbnail(self, path):
        """GET /v1/artifacts/<id>/thumbnail?version=N — serve the precomputed
        WebP thumbnail bytes (image artifacts only). 404 when none exists so the
        client falls back to /content. Cacheable (immutable per id+version)."""
        from urllib.parse import urlparse, parse_qs
        parts = path.split("/")
        artifact_id = parts[3] if len(parts) >= 5 else ""
        qs = parse_qs(urlparse(self.path).query)
        version = qs.get("version", [None])[0]
        thumb = ChatDB.get_artifact_thumbnail(artifact_id, version)
        if not thumb:
            self._send_json({"error": "No thumbnail"}, 404)
            return
        self.send_response(200)
        self.send_header("Content-Type", "image/webp")
        self.send_header("Content-Length", str(len(thumb)))
        self.send_header("Cache-Control", "private, max-age=31536000, immutable")
        self.end_headers()
        self.wfile.write(thumb)

    def _handle_artifact_content(self, path):
        """GET /v1/artifacts/<id>/content?version=N — get artifact version content."""
        from urllib.parse import urlparse, parse_qs
        import base64
        parts = path.split("/")
        # /v1/artifacts/<id>/content
        artifact_id = parts[3] if len(parts) >= 5 else ""
        qs = parse_qs(urlparse(self.path).query)
        version = qs.get("version", [None])[0]

        artifact = ChatDB.get_artifact(artifact_id)
        if not artifact:
            self._send_json({"error": "Artifact not found"}, 404)
            return

        ver_data = ChatDB.get_artifact_content(artifact_id, version)
        if not ver_data:
            self._send_json({"error": "Version not found"}, 404)
            return

        content_raw = ver_data["content"]
        is_binary = artifact["type"] in ("image", "document", "audio")

        if content_raw is None:
            # Disk-only fallback (file was > 5MB)
            try:
                with open(artifact["path"], "rb") as f:
                    content_raw = f.read()
            except Exception:
                self._send_json({"error": "Content not available"}, 404)
                return

        if is_binary:
            content_str = base64.b64encode(content_raw if isinstance(content_raw, bytes) else content_raw.encode()).decode()
            encoding = "base64"
        else:
            content_str = content_raw.decode("utf-8", errors="replace") if isinstance(content_raw, bytes) else content_raw
            encoding = "text"

        self._send_json({
            "artifact_id": artifact_id,
            "name": artifact["name"],
            "type": artifact["type"],
            "version": ver_data["version"],
            "content": content_str,
            "encoding": encoding,
            "size": ver_data["size"],
        })

    def _handle_artifact_download(self, path):
        """GET /v1/artifacts/<id>/download?version=N — download artifact content."""
        from urllib.parse import urlparse, parse_qs
        parts = path.split("/")
        artifact_id = parts[3] if len(parts) >= 5 else ""
        qs = parse_qs(urlparse(self.path).query)
        version = qs.get("version", [None])[0]

        artifact = ChatDB.get_artifact(artifact_id)
        if not artifact:
            self._send_json({"error": "Artifact not found"}, 404)
            return

        filename = artifact["name"]
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        content_types = {
            "md": "text/markdown", "txt": "text/plain", "py": "text/x-python",
            "json": "application/json", "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "html": "text/html", "csv": "text/csv",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "js": "application/javascript", "ts": "text/typescript",
            "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "gif": "image/gif", "svg": "image/svg+xml",
            "mp3": "audio/mpeg", "wav": "audio/wav", "m4a": "audio/mp4", "ogg": "audio/ogg",
        }
        ct = content_types.get(ext, "application/octet-stream")

        # If no version specified, serve disk file
        if not version:
            try:
                with open(artifact["path"], "rb") as f:
                    data = f.read()
            except Exception:
                self._send_json({"error": "File not found on disk"}, 404)
                return
        else:
            ver_data = ChatDB.get_artifact_content(artifact_id, version)
            if not ver_data or ver_data["content"] is None:
                self._send_json({"error": "Version content not available"}, 404)
                return
            data = ver_data["content"] if isinstance(ver_data["content"], bytes) else ver_data["content"].encode()

        self.send_response(200)
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", len(data))
        # Audio plays inline in the Studio <audio> element; everything else
        # force-downloads as before.
        disposition = "inline" if ct.startswith("audio/") else "attachment"
        self.send_header("Content-Disposition", f'{disposition}; filename="{filename}"')
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(data)

    def _handle_tool_result_download(self):
        """GET /v1/tools/result?session_id=X&tool_use_id=Y — serve the complete,
        uncapped tool result text that _apply_tool_result_budget spilled to disk
        when it exceeded the in-context budget (>50KB). The client falls back to
        this when its in-DOM copy is the truncated preview stub (after reload)."""
        import glob as _glob
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        session_id = (qs.get("session_id", [""])[0] or "").strip()
        tool_use_id = (qs.get("tool_use_id", [""])[0] or "").strip()
        if not session_id or not tool_use_id:
            self._send_json({"error": "session_id and tool_use_id required"}, 400)
            return
        # Reject path-traversal in the ids before they hit a glob/join.
        if any(c in session_id + tool_use_id for c in ("/", "\\", "..")):
            self._send_json({"error": "invalid id"}, 400)
            return

        info = ChatDB.get_session_info(session_id)
        # Scheduled-run results live under a synthetic sched-<run> session that
        # has no sessions row; allow it through (folder is owned by the agent).
        is_sched = session_id.startswith("sched-")
        if not info and not is_sched:
            self._send_json({"error": "Session not found"}, 404)
            return

        # Ownership: non-admins may only fetch results for sessions they can see.
        user = self._get_auth_user()
        if user and not user.get("is_admin") and info:
            visible = set(_auth_mod.get_visible_user_ids(user) or [])
            owner = info.get("user_id") or ""
            if owner and owner not in visible:
                self._send_json({"error": "forbidden"}, 403)
                return

        agent_id = (info or {}).get("agent_id") or "main"
        # The spill folder is <date>_<session_id>; the date isn't recoverable
        # here, so glob across all dated folders for this session.
        pattern = os.path.join(engine.AGENTS_DIR, agent_id, "artifacts",
                               f"*_{session_id}", "tool-results", f"{tool_use_id}.txt")
        matches = _glob.glob(pattern)
        if not matches:
            self._send_json({"error": "No persisted result for this tool call"}, 404)
            return
        # If somehow >1 (re-run on a new day), serve the newest.
        filepath = max(matches, key=os.path.getmtime)
        try:
            with open(filepath, "rb") as f:
                data = f.read()
        except OSError:
            self._send_json({"error": "File not readable"}, 404)
            return

        filename = f"{tool_use_id}.txt"
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", len(data))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(data)

    # --- SearXNG supervisor (admin) ---

    def _handle_searxng_status(self):
        """GET /v1/searxng/status — current SearXNG process state."""
        from server_lib.sidecar_supervisor import searxng_supervisor
        self._send_json(searxng_supervisor.status())

    def _handle_searxng_restart(self):
        """POST /v1/searxng/restart — hard-restart the bundled SearXNG instance.
        Clears the circuit breaker so a manual restart recovers from it."""
        from server_lib.sidecar_supervisor import searxng_supervisor
        user = self._get_auth_user() or {}
        result = searxng_supervisor.restart(
            reason=f"manual by={user.get('username','')}")
        try:
            import engine as _eng
            if _eng._audit_log:
                _eng._audit_log.log_action(
                    agent="main",
                    action_type="searxng_restart",
                    tool_name="searxng",
                    args_summary=f"by={user.get('username','')}",
                    result_status="ok" if result.get("ok") else "error",
                )
        except Exception:
            pass
        status_code = 200 if result.get("ok") else 409
        self._send_json(result, status_code)

    # --- crawl4ai render-service supervisor (admin) ---

    def _handle_crawl4ai_status(self):
        """GET /v1/crawl4ai/status — current crawl4ai render-service state."""
        from server_lib.sidecar_supervisor import crawl4ai_supervisor
        self._send_json(crawl4ai_supervisor.status())

    def _handle_crawl4ai_restart(self):
        """POST /v1/crawl4ai/restart — hard-restart the render service.
        Clears the circuit breaker so a manual restart recovers from it."""
        from server_lib.sidecar_supervisor import crawl4ai_supervisor
        user = self._get_auth_user() or {}
        result = crawl4ai_supervisor.restart(
            reason=f"manual by={user.get('username','')}")
        try:
            import engine as _eng
            if _eng._audit_log:
                _eng._audit_log.log_action(
                    agent="main",
                    action_type="crawl4ai_restart",
                    tool_name="crawl4ai",
                    args_summary=f"by={user.get('username','')}",
                    result_status="ok" if result.get("ok") else "error",
                )
        except Exception:
            pass
        status_code = 200 if result.get("ok") else 409
        self._send_json(result, status_code)

    def _handle_searxng_engines(self):
        """GET /v1/searxng/engines — last per-engine health snapshot (state +
        latency per search engine), as gathered by the hourly probe or the
        manual 'Test now' button. Empty 'engines' until the first probe runs."""
        from server_lib import searxng_health
        self._send_json(searxng_health.last_snapshot())

    def _handle_searxng_test_engines(self):
        """POST /v1/searxng/test-engines — run the per-engine health probe NOW
        (synchronous) and return the fresh snapshot."""
        from server_lib import searxng_health
        import brain as _brain
        base = _brain._searxng_base_url()
        snap = searxng_health.run_health_check(base)
        self._send_json(snap)
